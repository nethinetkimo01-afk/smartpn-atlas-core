"""
DS-04 Progress Table Parser
Rules per Jim 2026-06-15:
- Sheets: 1部-12部 (incl 11部(2)); skip 說明表/外包/lay ra
- LEAN: col A 加七A组 → 7A
- Instruction cols: ALL cols scanned (G/P/W/AH and others)
- Model title row: non-MF, has parenthesised code containing /
- Instruction row: starts with MF
- Multi-order (MF...-01-02--56-36): SUM qtys → 92, keep combined order no
- 外包鞋面 A: inline in instruction string → Y
- 外包鞋面 B: col B standalone "外包鞋面" → all subsequent instr Y until next model title
- Skip pure-numeric cells (totals)
"""
import openpyxl, re, os, sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ── Chinese numeral conversion ────────────────────────────────────────────────
CN_TO_NUM = {
    '一': '1', '二': '2', '三': '3', '四': '4', '五': '5', '六': '6',
    '七': '7', '八': '8', '九': '9', '十': '10', '十一': '11', '十二': '12',
}

def lean_to_code(s):
    """'加七A组'→'7A', '加十一A1组'→'11A1'  (returns None if not a LEAN name)"""
    s = s.strip()
    m = re.match(r'^加(十[一二]?|[一二三四五六七八九])([A-Z]\d?)组$', s)
    if not m:
        return None
    num = CN_TO_NUM.get(m.group(1), m.group(1))
    return f"{num}{m.group(2)}"

# ── Model name detection ──────────────────────────────────────────────────────
# Model names have parenthesised codes like (71039/AD-71968/71967) or (41TUR/AD-66068)
MODEL_CODE_RE = re.compile(r'\([A-Z0-9]+(?:[/\-][A-Z0-9()]+)+\)')

def is_model_title(s):
    if not s or s.upper().startswith('MF'):
        return False
    # Must have forward-slash inside parentheses (distinguishes from date like (6/5))
    m = MODEL_CODE_RE.search(s)
    if not m:
        return False
    # Also require some alpha letters before the code
    before = s[:s.find('(')].strip()
    return bool(re.search(r'[A-Za-z]', before))

def extract_model_name(s):
    idx = s.find('(')
    return s[:idx].strip() if idx > 0 else s.strip()

# ── MF instruction parser ─────────────────────────────────────────────────────
# Handles: MF2606KH8402-01--873(6/5)  →  single
# Handles: MF2606KH8402-01-02--56-36(6/13)  →  multi, sum=92
MF_RE = re.compile(
    r'^(MF\d{4})([A-Z]{1,2}\d{4,6})((?:-\d+)+)--([\d]+(?:-[\d]+)*)\((\d+/\d+)\)(.*)',
    re.DOTALL
)

def parse_mf(s):
    """Returns (order_no, art, qty_str, date, waibao_Y) or None"""
    m = MF_RE.match(s.strip())
    if not m:
        return None
    prefix, art, suf_str, qty_str, date, note = m.groups()
    waibao = 'Y' if '外包鞋面' in note else ''
    qtys = [int(q) for q in re.findall(r'\d+', qty_str)]
    total_qty = sum(qtys)
    order_no = f"{prefix}{art}{suf_str}"
    return order_no, art, str(total_qty), date, waibao

# ── Sheet parser ──────────────────────────────────────────────────────────────
PURE_NUM_RE = re.compile(r'^\d+(\.\d+)?$')
DEPT_RE = re.compile(r'^\d+部')

def parse_sheet(ws, dept_name):
    records = []
    unparsed = []

    current_lean_code = None
    col_model  = {}   # col_idx(0-based) → model name str
    col_waibao = {}   # col_idx → bool  (column-level 外包鞋面 flag)
    global_waibao = False   # B-col 外包鞋面 flag

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        col_a = str(row[0] or '').strip() if row else ''
        col_b = str(row[1] or '').strip() if len(row) > 1 else ''

        # ── LEAN header row ───────────────────────────────────────────────
        lean_code = lean_to_code(col_a)
        if lean_code is not None:
            current_lean_code = lean_code
            col_model  = {}
            col_waibao = {}
            global_waibao = False
            # Model names may appear in same row (other cols)
            for j, v in enumerate(row[1:], 1):
                s = str(v or '').strip()
                if is_model_title(s):
                    col_model[j] = extract_model_name(s)
            continue

        if current_lean_code is None:
            continue

        # ── B-col 外包鞋面 label ──────────────────────────────────────────
        if col_b.strip().replace('　', '').replace(' ', '') == '外包鞋面' and not col_a:
            global_waibao = True
            continue

        # ── Scan all other columns ────────────────────────────────────────
        found_model_this_row = False
        for j, v in enumerate(row):
            if j < 2:
                continue  # skip A & B
            s = str(v or '').strip()
            if not s:
                continue
            # Skip pure numeric totals
            if PURE_NUM_RE.match(s):
                continue
            # Skip standalone 外包鞋面 in non-B cols (visual marker, instructions already inline)
            if s.replace(' ', '') == '外包鞋面':
                col_waibao[j] = True
                continue

            if is_model_title(s):
                col_model[j] = extract_model_name(s)
                col_waibao[j] = False
                found_model_this_row = True
            elif s.startswith('MF'):
                result = parse_mf(s)
                if result:
                    order_no, art, qty, date, waibao_inline = result
                    model = col_model.get(j, '')
                    waibao = waibao_inline or ('Y' if (global_waibao or col_waibao.get(j, False)) else '')
                    records.append({
                        '部門': dept_name,
                        'LEAN': current_lean_code,
                        '鞋型名稱': model,
                        'ART': art,
                        '訂單號': order_no,
                        '數量': qty,
                        '交期': date,
                        '外包鞋面': waibao,
                    })
                else:
                    unparsed.append(f"  [{dept_name}] {s!r}")

        if found_model_this_row:
            global_waibao = False  # reset on new model title

    return records, unparsed

# ── Main ──────────────────────────────────────────────────────────────────────
SRC = r'data\source_files\廠務編制\2026年6月份正式进度表 5 30.xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)

all_records = []
all_unparsed = []
dept_stats = {}

for sh_name in wb.sheetnames:
    clean = sh_name.strip()
    if not DEPT_RE.match(clean):
        continue
    ws = wb[sh_name]
    dept = clean
    recs, unparsed = parse_sheet(ws, dept)
    all_records.extend(recs)
    all_unparsed.extend(unparsed)
    waibao_cnt = sum(1 for r in recs if r['外包鞋面'] == 'Y')
    dept_stats[dept] = {'total': len(recs), 'waibao': waibao_cnt}
    print(f"  {dept:12s}: {len(recs):4d} rows, {waibao_cnt:3d} 外包鞋面")

total = len(all_records)
total_waibao = sum(1 for r in all_records if r['外包鞋面'] == 'Y')
print(f"\n  TOTAL: {total} records, {total_waibao} 外包鞋面, {len(all_unparsed)} unparsed")

# ── Write xlsx ────────────────────────────────────────────────────────────────
OUT_DIR = r'flask_backend\test_output'
os.makedirs(OUT_DIR, exist_ok=True)

try:
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'DS04解析'
    HEADERS = ['部門', 'LEAN', '鞋型名稱', 'ART', '訂單號', '數量', '交期', '外包鞋面']
    ws_out.append(HEADERS)
    for r in all_records:
        ws_out.append([r[h] for h in HEADERS])
    # Style header
    from openpyxl.styles import Font, PatternFill
    for cell in ws_out[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='DDEEFF')
    # Auto-width
    for col in ws_out.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws_out.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    xlsx_path = os.path.join(OUT_DIR, 'ds04_parsed.xlsx')
    wb_out.save(xlsx_path)
    print(f"\n  Saved: {xlsx_path}")
except Exception as e:
    print(f"  xlsx error: {e}")

# ── Write summary txt ─────────────────────────────────────────────────────────
try:
    txt_path = os.path.join(OUT_DIR, 'ds04_summary.txt')
    with open(txt_path, 'w', encoding='utf-8') as f:
        f.write("DS-04 進度表解析摘要\n")
        f.write(f"來源: {SRC}\n")
        f.write(f"總計: {total} 筆, {total_waibao} 外包鞋面\n\n")
        f.write(f"{'部門':10s} {'筆數':>6s} {'外包':>6s}\n")
        f.write('-' * 28 + '\n')
        for dept, s in dept_stats.items():
            f.write(f"{dept:10s} {s['total']:>6d} {s['waibao']:>6d}\n")
        if all_unparsed:
            f.write(f"\n\n無法解析清單 ({len(all_unparsed)} 筆):\n")
            for u in all_unparsed:
                f.write(u + '\n')
    print(f"  Saved: {txt_path}")
except Exception as e:
    print(f"  txt error: {e}")

# ── Print first 10 records as preview ────────────────────────────────────────
print("\n  Preview (first 10 records):")
print(f"  {'部門':6s} {'LEAN':4s} {'鞋型名稱':25s} {'ART':8s} {'訂單號':22s} {'數量':5s} {'交期':6s} {'外包'}")
print("  " + "-"*95)
for r in all_records[:10]:
    print(f"  {r['部門']:6s} {r['LEAN']:4s} {r['鞋型名稱'][:24]:25s} {r['ART']:8s} {r['訂單號'][:21]:22s} {r['數量']:5s} {r['交期']:6s} {r['外包鞋面']}")
