"""Migration lean2: Supplement ob_header.lean from 廠務編制 unit xlsx files.

Sources (data/source_files/廠務編制/):
  - 06月份同材共裁自动化編制 *.xlsx
  - 06月份打粗水洗照射 *.xlsx
  - 2026年06月份電腦針車+折邊 *.xlsx

Strategy:
  Per NULL-lean ob_header: intersect lean-sets of its found ARTs.
  - intersection == {single}  → update ob_header.lean + allocation_item.lean
  - intersection size > 1 or 0 → AMBIGUOUS / not-found (no write)
"""
import openpyxl, os, re, glob, sqlite3, sys
from functools import reduce

sys.stdout.reconfigure(encoding='utf-8')

DB_PATH  = 'flask_backend/data/atlas.db'
SRC_DIR  = r'D:\smartpn-atlas-core\data\source_files\廠務編制'
ART_RE   = re.compile(r'\b([A-Z]{1,2}\d{4,6})\b')
LEAN_RE  = re.compile(r'\b(\d+[A-Z]\d*)\b')

# ── Locate xlsx files ─────────────────────────────────────────────────────────
def find_xlsx(keyword):
    return [m for m in glob.glob(os.path.join(SRC_DIR, f'*{keyword}*.xlsx'))
            if not os.path.basename(m).startswith('~$')]

file_groups = {
    '同材共裁': find_xlsx('同材共裁'),
    '打粗水洗': find_xlsx('打粗水洗'),
    '電腦針車': find_xlsx('电脑针车'),
}
for k, v in file_groups.items():
    print(f'{k}: {[os.path.basename(x) for x in v]}')

# ── Step 1: Build art → set(lean) from all xlsx sheets ───────────────────────
art_lean_map = {}

def scan_sheet(ws):
    rows_data = list(ws.iter_rows(max_row=400, max_col=35, values_only=True))

    # Find header row with LEAN column
    lean_col = art_col = None
    for row in rows_data[:10]:
        cells = [str(c or '').strip() for c in row]
        for j, c in enumerate(cells):
            if re.search(r'\blean\b', c, re.I) and lean_col is None:
                lean_col = j
            if re.search(r'\bart\b', c, re.I) and lean_col is not None and art_col is None:
                art_col = j
        if lean_col is not None:
            break
    if lean_col is None:
        return 0

    count = 0
    current_leans = set()
    for row in rows_data:
        cells = [str(c or '').strip() for c in row]
        lean_cell = cells[lean_col] if lean_col < len(cells) else ''
        new_leans = set(LEAN_RE.findall(lean_cell))
        if new_leans:
            current_leans = new_leans
        if not current_leans:
            continue

        # Scan ART column + neighbours
        start = max(0, (art_col or 3) - 1)
        end   = min(len(cells), (art_col or 3) + 6)
        for cell in cells[start:end]:
            for art in ART_RE.findall(cell.upper()):
                art_lean_map.setdefault(art, set()).update(current_leans)
                count += 1
    return count

for category, fpaths in file_groups.items():
    for fpath in fpaths:
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        except Exception as e:
            print(f'  WARN {os.path.basename(fpath)}: {e}')
            continue
        for sname in wb.sheetnames:
            ws = wb[sname]
            n = scan_sheet(ws)
            if n:
                print(f'  [{category}] "{sname}": {n} mappings')
        wb.close()

print(f'\nTotal unique ARTs in mapping: {len(art_lean_map)}')

# ── Step 2: Load NULL-lean headers ───────────────────────────────────────────
conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

null_headers = cur.execute('''
    SELECT h.id, h.model_name, GROUP_CONCAT(a.art, ',') as arts
    FROM ob_header h
    JOIN ob_articles a ON a.header_id=h.id
    WHERE h.lean IS NULL
    GROUP BY h.id
''').fetchall()
print(f'NULL-lean headers: {len(null_headers)}')

# ── Step 3: Resolve via intersection ─────────────────────────────────────────
updated   = 0
ambiguous = []
not_found = 0

for row in null_headers:
    hid   = row['id']
    arts  = [a.strip() for a in (row['arts'] or '').split(',') if a.strip()]
    model = (row['model_name'] or '')[:40]

    found_sets = [art_lean_map[a] for a in arts if a in art_lean_map]
    if not found_sets:
        not_found += 1
        continue

    lean_set = reduce(lambda a, b: a & b, found_sets)  # intersection

    if len(lean_set) == 1:
        lean_val = next(iter(lean_set))
        cur.execute('UPDATE ob_header SET lean=? WHERE id=?', (lean_val, hid))
        cur.execute('UPDATE allocation_item SET lean=? WHERE header_id=?', (lean_val, hid))
        updated += 1
        print(f'  UPDATE header {hid:4d} → {lean_val:5s}  [{model}]')
    else:
        union_all = sorted(reduce(lambda a, b: a | b, found_sets))
        ambiguous.append((hid, model, arts, union_all, sorted(lean_set)))

conn.commit()

# ── Step 4: Report ────────────────────────────────────────────────────────────
print(f'\n=== RESULT ===')
print(f'Updated (single LEAN):       {updated}')
print(f'No data in xlsx:             {not_found}')
print(f'AMBIGUOUS (needs Jim review):{len(ambiguous)}')

if ambiguous:
    print('\n--- AMBIGUOUS LIST ---')
    for hid, model, arts, union_all, inter in ambiguous:
        print(f'  header {hid:4d}  {model}')
        print(f'    ARTs:        {arts}')
        print(f'    Union LEANs: {union_all}')
        print(f'    Intersection:{inter or "(empty — conflicting units)"}')

print('\n--- ob_header.lean distribution ---')
for r in cur.execute('SELECT lean, COUNT(*) n FROM ob_header GROUP BY lean ORDER BY lean'):
    print(f'  {(r["lean"] or "NULL"):10s}  {r["n"]}')

conn.close()
print('\nDone.')
