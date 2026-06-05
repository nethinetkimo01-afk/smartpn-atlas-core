#!/usr/bin/env python3
"""Inspect DS-04 sheet structure to find LEAN group title rows."""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl

SCHEDULE = r'C:\Users\user\AppData\Local\Temp\sched_tmp.xlsx'
_ART_RE  = re.compile(r'[A-Z]{2}\d{4,6}')

wb = openpyxl.load_workbook(SCHEDULE, data_only=True, read_only=True)
print(f'Sheets ({len(wb.sheetnames)}): {wb.sheetnames}')
print()

for ws in wb.worksheets:
    title = ws.title.strip()
    print(f'=== Sheet: {title!r} ===')
    row_count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 60:
            break
        non_empty = [(ci+1, str(v or '').strip()) for ci, v in enumerate(row)
                     if str(v or '').strip()]
        if non_empty:
            # Print rows that look like headers or group titles (not pure order cells)
            has_art = any(_ART_RE.search(v) and '--' in v for _, v in non_empty)
            if not has_art:
                print(f'  R{i:3d}: ' + '  |  '.join(f'col{c}={v[:40]}' for c, v in non_empty[:6]))
    print()

wb.close()
