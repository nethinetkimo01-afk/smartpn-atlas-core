#!/usr/bin/env python3
"""
DS-05 大底課進度表 (Sole Department Progress Sheet) Analyzer.

Source file structure (column-A driven):
  Column A cell: "T1\\n5月:20人\\n6月:22人"   ← T-group header
                 "T1+T2\\n5月:40人\\n6月:44人" ← combined group
  Within each T-group row range:
    Model header cell (any column): "ADICHILL AD-69148"  ← AD code + name
                                    or "ADICHILL\\nAD-69148" (multiline)
    Order cells:  MF2604KJ0835-01--500(5/15)             ← MF order format

Parsing rules:
  1. Scan column A for T-group markers (starts with T + digit or "+").
  2. Within each T-group, scan every cell for AD-codes (format: AD-\d{5}).
  3. MF orders are assigned to the nearest AD-code above them (row-order).
  4. Same AD-code within a T-group → merge model entries, sum quantities.
  5. ADICHILL fix: when a cell contains "ADICHILL" without an AD-code on the
     same line, look at the very next non-empty cell in the same column or the
     cell immediately below for a matching AD-xxxxx pattern.

Usage:
    python analyze_ds05.py <progress_sheet.xlsx>
    python analyze_ds05.py <progress_sheet.xlsx> --group T1
    python analyze_ds05.py <progress_sheet.xlsx> --dry-run
"""
import sys, os, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── Regex patterns ────────────────────────────────────────────────────────────

# T-group marker in column A: T1 / T2 / T1+T2 / T1+T2+T3 (with optional whitespace)
_T_GROUP_RE = re.compile(r'^(T\d+(?:\+T\d+)*)', re.IGNORECASE)

# Headcount: "5月:20人" or "6月:22人" anywhere in the T-header cell text
_HEAD_RE = re.compile(r'(\d+)月[:\s：]*(\d+)\s*人')

# AD-code: AD-12345 (5 digits)
_AD_RE = re.compile(r'AD-(\d{5})', re.IGNORECASE)

# MF order number: MFyymmART-seq--qty(date)
# Examples:  MF2604KJ0835-01--500(5/15)
#            MF2506IH1650-02--1200(6/20)
_MF_RE = re.compile(r'MF\d{4}([A-Z]{2}\d{4,6})-\d+--(\d+)\((\d+/\d+)\)',
                    re.IGNORECASE)

# ART code anywhere (fallback)
_ART_RE = re.compile(r'[A-Z]{2}\d{4,6}')


# ── Helpers ───────────────────────────────────────────────────────────────────

def _cell_text(val) -> str:
    """Return stripped string from a cell value (handles None + multiline)."""
    return str(val).strip() if val is not None else ''


def _parse_t_header(text: str):
    """
    Parse a T-group header cell.
    Returns (group_name, headcount_dict) or (None, {}).

    group_name: exactly as in source, e.g. "T1+T2+T3"
    headcount:  {month_str: count}  e.g. {"5月": 20, "6月": 22}
    """
    m = _T_GROUP_RE.match(text.strip())
    if not m:
        return None, {}
    group_name = m.group(1).upper()
    headcount = {}
    for mo, cnt in _HEAD_RE.findall(text):
        headcount[f'{mo}月'] = int(cnt)
    return group_name, headcount


def _parse_mf_orders(text: str) -> list:
    """
    Extract all MF orders from a cell text.
    Returns list of {'art': str, 'qty': int, 'deadline': str}.
    """
    return [
        {'art': art.upper(), 'qty': int(qty), 'deadline': dl}
        for art, qty, dl in _MF_RE.findall(text)
    ]


def _extract_ad(text: str):
    """Return first AD-code string from text, or None."""
    m = _AD_RE.search(text)
    return f'AD-{m.group(1)}' if m else None


def _is_adichill(text: str) -> bool:
    return 'ADICHILL' in text.upper()


# ── Core parser ───────────────────────────────────────────────────────────────

def parse_sheet(ws) -> list:
    """
    Parse a single worksheet into a list of T-group dicts.

    Return schema per group:
    {
      'group_name': str,           # e.g. "T1+T2+T3"
      'headcount': {str: int},     # e.g. {"5月": 20, "6月": 22}
      'start_row': int,
      'end_row': int,
      'models': [                  # ordered by first appearance in the sheet
        {
          'ad_code': str,          # "AD-69148"
          'model_name': str,       # "ADICHILL" (text before AD-code)
          'orders': [{'art', 'qty', 'deadline'}],
          'total_qty': int,
        }
      ]
    }
    """
    # ── Pass 1: collect all rows as {row_idx: {col_idx: text}} ───────────────
    rows_data: dict[int, dict[int, str]] = {}
    for row in ws.iter_rows(values_only=True):
        rn = row[0]  # placeholder; will re-iterate with row numbers
        break

    all_rows: dict[int, dict[int, str]] = {}
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_cells = {}
        for col_idx, val in enumerate(row):
            t = _cell_text(val)
            if t:
                row_cells[col_idx] = t
        if row_cells:
            all_rows[row_idx] = row_cells

    # ── Pass 2: identify T-group boundaries from column A (col_idx=0) ────────
    t_groups_raw: list[tuple[int, str, dict]] = []  # (row_idx, name, headcount)
    for row_idx, cols in sorted(all_rows.items()):
        col_a = cols.get(0, '')
        name, hc = _parse_t_header(col_a)
        if name:
            t_groups_raw.append((row_idx, name, hc))

    if not t_groups_raw:
        return []

    # Add sentinel at end
    max_row = max(all_rows.keys()) + 1
    t_groups_raw.append((max_row, '__END__', {}))

    # ── Pass 3: for each T-group range, extract models and MF orders ─────────
    results = []

    for i, (start_row, group_name, headcount) in enumerate(t_groups_raw[:-1]):
        end_row = t_groups_raw[i + 1][0] - 1

        # Collect all (row, col, text) triples in this range
        cells_in_range: list[tuple[int, int, str]] = []
        for row_idx in range(start_row, end_row + 1):
            for col_idx, text in all_rows.get(row_idx, {}).items():
                cells_in_range.append((row_idx, col_idx, text))

        # ── Find model headers (cells containing AD-code) ─────────────────
        # Also handle ADICHILL fix: model name in one cell, AD-code in the
        # next non-empty cell of the same column (within ±2 rows).
        model_headers: list[tuple[int, str, str]] = []
        # (row_idx, ad_code, model_name)

        # Pre-build column text index for ADICHILL lookahead
        col_texts: dict[int, list[tuple[int, str]]] = {}  # col → [(row, text)]
        for r, c, t in cells_in_range:
            col_texts.setdefault(c, []).append((r, t))
        for c in col_texts:
            col_texts[c].sort()

        seen_ad_rows: set[int] = set()

        for row_idx, col_idx, text in cells_in_range:
            ad = _extract_ad(text)
            if ad:
                # Extract model name: everything before the AD- token
                model_name = _AD_RE.split(text, maxsplit=1)[0].strip(' \t\n\r-_').strip()
                model_headers.append((row_idx, ad, model_name))
                seen_ad_rows.add(row_idx)
            elif _is_adichill(text) and row_idx not in seen_ad_rows:
                # ADICHILL fix: look for AD-code in the next 1–2 rows of the
                # same column, or in any cell on the very next row.
                ad_found = None
                model_name = text.strip()
                # Same column, next rows
                col_list = col_texts.get(col_idx, [])
                for r, t in col_list:
                    if r <= row_idx:
                        continue
                    if r > row_idx + 2:
                        break
                    ad_found = _extract_ad(t)
                    if ad_found:
                        break
                # Fallback: any cell on row_idx+1
                if not ad_found:
                    for r, c, t in cells_in_range:
                        if r == row_idx + 1:
                            ad_found = _extract_ad(t)
                            if ad_found:
                                break
                if ad_found:
                    model_headers.append((row_idx, ad_found, model_name))
                    seen_ad_rows.add(row_idx)
                    seen_ad_rows.add(row_idx + 1)

        # Sort model headers by row
        model_headers.sort(key=lambda x: x[0])

        # ── Assign MF orders to models ────────────────────────────────────
        # Each MF order belongs to the last model header at or above its row.
        # Build intervals: model i owns rows [header_row_i .. header_row_{i+1}-1]
        # (or until end_row for the last model).

        order_map: dict[str, list] = {ad: [] for _, ad, _ in model_headers}
        model_name_map: dict[str, str] = {}

        for _, ad, mname in model_headers:
            if ad not in model_name_map or mname:
                model_name_map[ad] = mname

        header_rows = [r for r, _, _ in model_headers]

        def _owning_ad(row_idx: int):
            """Return AD-code for the model that owns this row, or None."""
            best = None
            for idx, (r, ad, _) in enumerate(model_headers):
                if r <= row_idx:
                    best = ad
                else:
                    break
            return best

        for row_idx, col_idx, text in cells_in_range:
            orders = _parse_mf_orders(text)
            if not orders:
                continue
            ad = _owning_ad(row_idx)
            if ad is None:
                continue  # orders before any model header → skip
            order_map[ad].extend(orders)

        # ── Aggregate orders per AD-code ──────────────────────────────────
        models_out = []
        seen_ads_in_group: dict[str, dict] = {}

        for _, ad, mname in model_headers:
            if ad in seen_ads_in_group:
                # Merge: sum orders
                entry = seen_ads_in_group[ad]
                entry['orders'].extend(order_map.get(ad, []))
                entry['total_qty'] += sum(o['qty'] for o in order_map.get(ad, []))
                if not entry['model_name'] and mname:
                    entry['model_name'] = mname
            else:
                orders = order_map.get(ad, [])
                entry = {
                    'ad_code': ad,
                    'model_name': mname or model_name_map.get(ad, ''),
                    'orders': orders,
                    'total_qty': sum(o['qty'] for o in orders),
                }
                seen_ads_in_group[ad] = entry
                models_out.append(entry)

        results.append({
            'group_name': group_name,
            'headcount': headcount,
            'start_row': start_row,
            'end_row': end_row,
            'models': models_out,
        })

    return results


# ── Main analyze entry point ──────────────────────────────────────────────────

def analyze(file_path: str, group_filter: str = '') -> dict:
    """
    Run DS-05 analysis on a 大底課進度表 Excel file.

    Args:
        file_path:    Absolute or relative path to the .xlsx/.xls file.
        group_filter: If set, return only the T-group whose name contains
                      this string (case-insensitive).  Empty = all groups.

    Returns dict with keys:
        ok, file, sheet_used, groups, total_groups, total_models, total_qty
    """
    if not HAS_OPENPYXL:
        return {'ok': False, 'error': 'openpyxl not installed'}
    if not os.path.exists(file_path):
        return {'ok': False, 'error': f'File not found: {file_path}'}

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return {'ok': False, 'error': f'Cannot open file: {e}'}

    # Use first sheet by default; prefer a sheet whose name hints at 大底
    ws = wb.worksheets[0]
    for s in wb.worksheets:
        if any(k in s.title for k in ['大底', '底', 'sole', 'Sole']):
            ws = s
            break

    groups = parse_sheet(ws)
    wb.close()

    # Filter by group name if requested
    if group_filter:
        gf = group_filter.upper().strip()
        groups = [g for g in groups if gf in g['group_name'].upper()]

    total_models = sum(len(g['models']) for g in groups)
    total_qty = sum(
        m['total_qty']
        for g in groups
        for m in g['models']
    )

    return {
        'ok': True,
        'file': os.path.basename(file_path),
        'sheet_used': ws.title,
        'total_groups': len(groups),
        'total_models': total_models,
        'total_qty': total_qty,
        'groups': groups,
    }


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import argparse, json

    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    parser = argparse.ArgumentParser(description='DS-05 大底課進度表 Analyzer')
    parser.add_argument('file', help='Progress sheet Excel path (.xlsx)')
    parser.add_argument('--group', default='', help='Filter by T-group name (e.g. T1)')
    parser.add_argument('--dry-run', action='store_true',
                        help='Print summary only, no DB writes')
    args = parser.parse_args()

    result = analyze(args.file, args.group)

    if args.dry_run and result.get('ok'):
        # Print compact summary
        print(f"File:   {result['file']}")
        print(f"Sheet:  {result['sheet_used']}")
        print(f"Groups: {result['total_groups']}  "
              f"Models: {result['total_models']}  "
              f"Total qty: {result['total_qty']:,}")
        print()
        for g in result['groups']:
            hc = ', '.join(f"{m}:{v}人" for m, v in g['headcount'].items())
            print(f"  [{g['group_name']}]  {hc}  ({len(g['models'])} models)")
            for m in g['models']:
                arts = ', '.join(sorted({o['art'] for o in m['orders']}))
                print(f"    {m['ad_code']:12s}  {m['model_name'][:30]:30s}  "
                      f"qty={m['total_qty']:6d}  arts=[{arts}]")
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
