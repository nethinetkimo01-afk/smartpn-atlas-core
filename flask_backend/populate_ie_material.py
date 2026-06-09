#!/usr/bin/env python3
"""
One-time script: scan all IE Excel files, extract material type from SUM.C2B sheet
row 4, then populate ob_header.material for matching records.

SUM.C2B row 4 format:
  "Loại nguyên liệu  Material 材料类 : <value>"

Run: python flask_backend/populate_ie_material.py
"""
import sys, io, os, re, sqlite3
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

DB_PATH = os.path.join(os.path.dirname(__file__), 'data', 'atlas.db')

SCAN_DIRS = [
    r'C:\Users\user\OneDrive\Desktop\IE',
    r'C:\Users\user\OneDrive\Desktop\Biên chế\Jun\IE',
]

ART_RE = re.compile(r'\b([A-Z]{2}\d{4,6})\b')


def extract_art_and_material(fp):
    """Return (art, material_str) from the IE Excel file, or (None, None)."""
    try:
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    except Exception:
        return None, None

    art = None
    material = None

    # Find SUM.C2B sheet (shortest match)
    c2b_sheet = None
    best_len = float('inf')
    for sn in wb.sheetnames:
        sk = sn.lower().replace(' ', '')
        if 'sum' in sk and 'c2b' in sk and len(sn) < best_len:
            c2b_sheet = sn
            best_len = len(sn)

    if c2b_sheet:
        ws = wb[c2b_sheet]
        for row in ws.iter_rows(max_row=10, values_only=True):
            for cell in row:
                if cell is None:
                    continue
                s = str(cell).strip()
                sl = s.lower()
                # ART
                if art is None and ('article' in sl or 'mã số' in sl) and ':' in s:
                    m = ART_RE.search(s.split(':',1)[1])
                    if m:
                        art = m.group(1)
                # Material: "Loại nguyên liệu  Material 材料类 : ..."
                if material is None and ('nguyên liệu' in sl or 'material' in sl or '材料' in s):
                    if ':' in s:
                        mat = s.split(':',1)[1].strip()
                        if mat:
                            material = mat

    wb.close()
    return art, material


def run():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    updated = 0
    skipped_no_art = 0
    skipped_no_mat = 0
    skipped_no_match = 0

    for base_dir in SCAN_DIRS:
        if not os.path.isdir(base_dir):
            print(f'[material] SKIP (not found): {base_dir}')
            continue
        all_files = []
        for root, dirs, files in os.walk(base_dir):
            dirs.sort()
            for fn in sorted(files):
                if fn.lower().endswith(('.xlsx', '.xlsm')) and not fn.startswith('~'):
                    all_files.append(os.path.join(root, fn))
        print(f'[material] Scanning {len(all_files)} files in {base_dir} (recursive)')
        for fp in all_files:
            fn = os.path.basename(fp)
            art, material = extract_art_and_material(fp)
            if not art:
                skipped_no_art += 1
                continue
            if not material:
                skipped_no_mat += 1
                continue
            # Find header_id via art
            row = conn.execute(
                'SELECT header_id FROM ob_articles WHERE art=? LIMIT 1', (art,)
            ).fetchone()
            if not row:
                skipped_no_match += 1
                continue
            header_id = row['header_id']
            # Update all headers with the same model_name (EOLR=60 and 120 should share same material)
            model_row = conn.execute('SELECT model_name FROM ob_header WHERE id=?', (header_id,)).fetchone()
            if model_row:
                conn.execute(
                    "UPDATE ob_header SET material=? WHERE model_name=? AND (material IS NULL OR material='')",
                    (material, model_row['model_name'])
                )
                affected = conn.execute(
                    'SELECT changes()'
                ).fetchone()[0]
                updated += affected

    conn.commit()

    total = conn.execute('SELECT COUNT(*) FROM ob_header').fetchone()[0]
    filled = conn.execute("SELECT COUNT(*) FROM ob_header WHERE material IS NOT NULL AND material!=''").fetchone()[0]
    conn.close()

    print(f'\n[material] Done.')
    print(f'  Updated rows: {updated}')
    print(f'  Skip (no ART extracted): {skipped_no_art}')
    print(f'  Skip (no material extracted): {skipped_no_mat}')
    print(f'  Skip (ART not in DB): {skipped_no_match}')
    print(f'  ob_header with material: {filled}/{total}')


if __name__ == '__main__':
    run()
