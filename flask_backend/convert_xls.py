#!/usr/bin/env python3
"""
Convert legacy .xls files to .xlsx format.
Requires: xlrd==1.2.0 (pip install xlrd==1.2.0)

Usage:
    python convert_xls.py <folder>          # convert all .xls in folder
    python convert_xls.py <file.xls>        # convert single file
    python convert_xls.py <folder> --dry-run
"""
import sys, os, glob

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import win32com.client
    HAS_WIN32 = True
except ImportError:
    HAS_WIN32 = False


def _convert_via_excel_com(xls_path, xlsx_path):
    """Use Excel COM automation to convert .xls → .xlsx (most reliable on Windows)."""
    import pythoncom
    excel = None
    try:
        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx('Excel.Application')
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(os.path.abspath(xls_path))
        wb.SaveAs(os.path.abspath(xlsx_path), FileFormat=51)
        wb.Close(False)
        return True, None
    except Exception as e:
        return False, str(e)
    finally:
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def _convert_via_xlrd(xls_path, xlsx_path):
    """Convert .xls → .xlsx using xlrd (pure Python, no Excel required)."""
    if not HAS_XLRD:
        return False, 'xlrd not installed'
    if not HAS_OPENPYXL:
        return False, 'openpyxl not installed'
    try:
        xls_wb = xlrd.open_workbook(xls_path)
    except Exception as e:
        return False, f'Cannot open .xls: {e}'
    try:
        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)
        for sheet_idx in range(xls_wb.nsheets):
            xls_ws = xls_wb.sheet_by_index(sheet_idx)
            new_ws = new_wb.create_sheet(title=xls_ws.name[:31])
            for row_idx in range(xls_ws.nrows):
                for col_idx in range(xls_ws.ncols):
                    cell = xls_ws.cell(row_idx, col_idx)
                    if cell.ctype == 0:
                        val = None
                    elif cell.ctype == 3:
                        try:
                            import datetime
                            dt = xlrd.xldate_as_datetime(cell.value, xls_wb.datemode)
                            val = dt.strftime('%Y-%m-%d')
                        except Exception:
                            val = cell.value
                    elif cell.ctype == 4:
                        val = bool(cell.value)
                    else:
                        val = cell.value
                    new_ws.cell(row=row_idx + 1, column=col_idx + 1, value=val)
        new_wb.save(xlsx_path)
        return True, None
    except Exception as e:
        return False, f'Conversion failed: {e}'


def convert_file(xls_path, overwrite=False):
    """Convert a single .xls file to .xlsx. Tries xlrd first, then Excel COM."""
    xlsx_path = os.path.splitext(xls_path)[0] + '.xlsx'
    if os.path.exists(xlsx_path) and not overwrite:
        return {'ok': True, 'skipped': True, 'path': xlsx_path, 'reason': 'already exists'}

    # Try xlrd first (no Excel needed)
    ok, err = _convert_via_xlrd(xls_path, xlsx_path)
    if ok:
        return {'ok': True, 'skipped': False, 'path': xlsx_path, 'method': 'xlrd'}

    # Fall back to Excel COM (requires Excel installed)
    if HAS_WIN32:
        ok2, err2 = _convert_via_excel_com(xls_path, xlsx_path)
        if ok2:
            return {'ok': True, 'skipped': False, 'path': xlsx_path, 'method': 'excel-com'}
        return {'ok': False, 'error': f'xlrd: {err} | excel-com: {err2}'}

    return {'ok': False, 'error': f'xlrd failed: {err}. Install pywin32 for Excel COM fallback.'}


def batch_convert(folder, dry_run=False, overwrite=False):
    files = glob.glob(os.path.join(folder, '**', '*.xls'), recursive=True)
    files = [f for f in files if not f.endswith('.xlsx')
             and not os.path.basename(f).startswith('~$')]

    if not files:
        print(f'No .xls files found in: {folder}')
        return 0, 0

    print(f'Found {len(files)} .xls file(s)')
    if dry_run:
        print('DRY RUN — no files written\n')

    ok_count = err_count = skip_count = 0

    for path in files:
        rel = os.path.relpath(path, folder)
        if dry_run:
            xlsx_exists = os.path.exists(os.path.splitext(path)[0] + '.xlsx')
            status = 'SKIP (xlsx exists)' if xlsx_exists and not overwrite else 'WOULD CONVERT'
            print(f'  {status}: {rel}')
            continue

        result = convert_file(path, overwrite=overwrite)
        if result.get('skipped'):
            skip_count += 1
            print(f'  SKIP: {rel} (xlsx exists)')
        elif result['ok']:
            ok_count += 1
            print(f'  OK: {rel} → {os.path.basename(result["path"])} '
                  f'({result.get("sheets",0)} sheets, {result.get("rows",0)} rows)')
        else:
            err_count += 1
            print(f'  ERROR: {rel} — {result["error"]}')

    if not dry_run:
        print(f'\nDone: {ok_count} converted, {skip_count} skipped, {err_count} errors')

    return ok_count, err_count


if __name__ == '__main__':
    if len(sys.argv) < 2 or sys.argv[1] in ('-h', '--help'):
        print(__doc__)
        sys.exit(0)

    if not HAS_XLRD:
        print('ERROR: xlrd not installed. Run: pip install xlrd==1.2.0')
        sys.exit(1)

    target = sys.argv[1]
    dry_run = '--dry-run' in sys.argv
    overwrite = '--overwrite' in sys.argv

    if os.path.isfile(target):
        if dry_run:
            print(f'DRY RUN: would convert {target}')
        else:
            r = convert_file(target, overwrite=overwrite)
            if r['ok'] and not r.get('skipped'):
                print(f'Converted: {r["path"]}')
            elif r.get('skipped'):
                print(f'Skipped: {r["path"]} (already exists)')
            else:
                print(f'Error: {r["error"]}')
                sys.exit(1)
    elif os.path.isdir(target):
        batch_convert(target, dry_run=dry_run, overwrite=overwrite)
    else:
        print(f'Not found: {target}')
        sys.exit(1)
