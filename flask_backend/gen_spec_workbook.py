#!/usr/bin/env python3
"""
IE Sheet Spec Workbook Generator
Output: sheet規格定義_待Jim審核.xlsx + formula_variants.txt
"""
import sqlite3, sys, os, re
from collections import defaultdict, Counter
from datetime import datetime

if hasattr(sys.stdout,'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB       = r"D:\smartpn-atlas-core\flask_backend\data\atlas.db"
OUT_XL   = r"D:\smartpn-atlas-core\flask_backend\test_output\sheet規格定義_待Jim審核.xlsx"
OUT_TXT  = r"D:\smartpn-atlas-core\flask_backend\test_output\formula_variants.txt"
TARGET   = 32   # header_id for LA TRAINER OG

COMMON_TYPES = ['tongcai','cutting','atom','computer_stitch','stitching',
                'assembly_1','assembly_2','sum_c2b','stock','dacui','xishi']

# ── helpers ───────────────────────────────────────────────────────────────────

def cletter(n):
    r=''
    while n:
        n,rem=divmod(n-1,26)
        r=chr(65+rem)+r
    return r

def normalize_f(f):
    """Replace row numbers in cell refs so =F10/I10/222 → =FN/IN/222"""
    return re.sub(r'(?<=[A-Za-z\$])(\d+)(?![\d])', 'N', f) if f else ''

def thin_border():
    s=Side(style='thin',color='CCCCCC')
    return Border(left=s,right=s,top=s,bottom=s)

def fill(hex_):
    return PatternFill('solid', fgColor=hex_)

def font(bold=False,color='000000',size=10):
    return Font(bold=bold,color=color,size=size)

# ── DB helpers ────────────────────────────────────────────────────────────────

def get_conn():
    conn=sqlite3.connect(DB); conn.row_factory=sqlite3.Row
    return conn

def load_grid(conn, header_id, type_key):
    """Return (sheet_name, grid {row:{col:{v,f,t}}}) for first matching sheet."""
    sn=conn.execute(
        '''SELECT DISTINCT sd.sheet_name FROM ie_sheet_data sd
           JOIN sheet_name_map snm ON snm.raw_name=sd.sheet_name
           WHERE sd.header_id=? AND snm.type_key=? LIMIT 1''',
        (header_id,type_key)
    ).fetchone()
    if not sn: return None,None
    sname=sn[0]
    rows=conn.execute(
        'SELECT row,col,value,formula FROM ie_sheet_data WHERE header_id=? AND sheet_name=? ORDER BY row,col',
        (header_id,sname)
    ).fetchall()
    grid={}
    for r in rows:
        if r['row'] not in grid: grid[r['row']]={}
        grid[r['row']][r['col']]={'v':r['value'],'f':r['formula']}
    return sname, grid

def best_alt_header(conn, type_key):
    """Find header_id with most cells for this type (excluding TARGET)."""
    r=conn.execute(
        '''SELECT sd.header_id, COUNT(*) c FROM ie_sheet_data sd
           JOIN sheet_name_map snm ON snm.raw_name=sd.sheet_name
           WHERE snm.type_key=? AND sd.header_id!=?
           GROUP BY sd.header_id ORDER BY c DESC LIMIT 1''',
        (type_key,TARGET)
    ).fetchone()
    return r['header_id'] if r else None

def detect_header_row(grid):
    """Find the row containing the column header labels."""
    # Look for a cell matching STT/序號/Seq.No in first 20 rows
    for rn in range(1,21):
        if rn not in grid: continue
        for cell in grid[rn].values():
            v=str(cell.get('v') or '').strip()
            if re.search(r'^\s*STT\s*$|序[號号]|Seq\.\s*No|^No\.?\s*$', v, re.I):
                return rn
    # Fallback: row with most distinct text values
    best,best_score=7,0
    for rn in range(1,18):
        if rn not in grid: continue
        score=sum(1 for c in grid[rn].values()
                  if c.get('v') and not str(c['v']).replace('.','').replace('-','').lstrip().isdigit()
                  and len(str(c['v']))>2)
        if score>best_score: best_score,best=score,rn
    return best

def detect_data_start(grid, header_row):
    """First row after header where first column has a numeric value."""
    for rn in range(header_row+1, header_row+6):
        if rn not in grid: continue
        first_col=min(grid[rn].keys())
        v=str(grid[rn][first_col].get('v') or '').strip()
        if v and re.match(r'^\d+$', v): return rn
    return header_row+2

# ── Field analysis ─────────────────────────────────────────────────────────────

def infer_input_type(header_text, has_formula, sample_val):
    h=header_text.lower()
    if has_formula: return '公式(鎖定)'
    if re.search(r'^\s*stt\s*$|序[號号]|^seq|^no\.', h, re.I): return '自動帶入'
    if re.search(r'model|art\b|eolr|target output|mục tiêu sản lượng|目標產能|鞋型名', h, re.I):
        return '自動帶入'
    if re.search(r'remark|ghi chú|備?注|note', h, re.I): return '人工輸入'
    return '人工輸入'

def is_title_content(header_text):
    h=header_text.lower()
    return any(k in h for k in ['biểu phân tích','operation breakdown','mã số văn kiện',
                                  'tên hình thể','model name','鞋型名称','目標產能','document no'])

def analyze_fields(grid, header_row, data_start):
    fields=[]
    # Title rows
    for rn in range(1, header_row):
        if rn not in grid: continue
        vals=[]
        for col in sorted(grid[rn].keys()):
            v=str(grid[rn][col].get('v') or '').strip()
            if v: vals.append(v[:60])
        if vals:
            text=' | '.join(vals[:4])
            fields.append(dict(is_title=True, pos=f'第{rn}行', name=text[:80],
                               sample='', formula='', keep='不要', itype='—',
                               col_num=-rn, vc=0))
    # Column headers from header_row (may span 2 rows for bilingual)
    col_names={}
    for sub in range(header_row, min(header_row+3, data_start)):
        if sub not in grid: continue
        for col,cell in grid[sub].items():
            v=str(cell.get('v') or '').strip()
            if not v: continue
            if col not in col_names:
                col_names[col]=v
            elif len(col_names[col])<50:
                col_names[col]+=' / '+v[:25]
    # Data columns
    # collect all columns that appear in data rows
    data_cols=set(col_names.keys())
    for rn in range(data_start, min(data_start+15, max(grid.keys() if grid else [data_start])+1)):
        if rn in grid: data_cols.update(grid[rn].keys())
    for col in sorted(data_cols):
        htext=col_names.get(col, f'(無標題 {cletter(col)})')
        # collect from data rows
        samples,formulas=[],[]
        for rn in range(data_start, data_start+60):
            if rn not in grid or col not in grid[rn]: continue
            cell=grid[rn][col]
            if cell.get('f'):
                formulas.append(cell['f'])
            elif cell.get('v'):
                samples.append(str(cell['v']))
        if not samples and not formulas: continue
        formula1=formulas[0] if formulas else ''
        sample1=(samples[0] if samples else formula1)[:50]
        itype=infer_input_type(htext, bool(formula1), sample1)
        keep = '要'
        fields.append(dict(is_title=False, pos=cletter(col), name=htext[:60],
                            sample=sample1, formula=formula1[:120], keep=keep,
                            itype=itype, col_num=col, vc=0))
    return fields

# ── Appearance rates ───────────────────────────────────────────────────────────

def appearance_rates(conn, type_key, header_row):
    total=conn.execute(
        '''SELECT COUNT(DISTINCT sd.header_id) FROM ie_sheet_data sd
           JOIN sheet_name_map snm ON snm.raw_name=sd.sheet_name
           WHERE snm.type_key=?''', (type_key,)
    ).fetchone()[0]
    if not total: return {}, 0
    col_cnts=conn.execute(
        '''SELECT sd.col, COUNT(DISTINCT sd.header_id) cnt
           FROM ie_sheet_data sd
           JOIN sheet_name_map snm ON snm.raw_name=sd.sheet_name
           WHERE snm.type_key=? AND sd.row > ?
           GROUP BY sd.col''', (type_key, header_row)
    ).fetchall()
    return {r['col']: round(r['cnt']/total*100) for r in col_cnts}, total

# ── Formula variants ──────────────────────────────────────────────────────────

def formula_variants(conn, type_key, header_row):
    """Return {col: [(norm_f, count), ...]} for cols with >1 variant."""
    # Get one formula per (header_id, col) from first data row
    rows=conn.execute(
        '''SELECT sd.header_id, sd.col, sd.formula
           FROM ie_sheet_data sd
           JOIN sheet_name_map snm ON snm.raw_name=sd.sheet_name
           WHERE snm.type_key=? AND sd.formula IS NOT NULL AND sd.row > ? AND sd.row <= ?+5
           ORDER BY sd.col, sd.header_id, sd.row''',
        (type_key, header_row, header_row)
    ).fetchall()
    # One formula per (header_id, col)
    seen={}
    for r in rows:
        key=(r['header_id'],r['col'])
        if key not in seen:
            seen[key]=normalize_f(r['formula'])
    col_variants=defaultdict(Counter)
    for (hid,col),nf in seen.items():
        col_variants[col][nf]+=1
    return {col: sorted(cnt.items(), key=lambda x:-x[1])
            for col,cnt in col_variants.items() if len(cnt)>1}

# ── Excel writer ──────────────────────────────────────────────────────────────

HDR_COLS=['A: 原表位置','B: 欄位名稱/表頭文字','C: 樣本值',
          'D: 公式原文','E: 判定 要/不要','F: 判定 輸入方式',
          'G: 出現率%','H: 【Jim反饋欄】← 空=同意，有字=修正意見']
COL_WIDTHS=[14,38,32,52,14,18,12,42]

F_HEADER = fill('1A2744')
F_META   = fill('E8ECF5')
F_TITLE  = fill('F0F0F0')
F_FORM   = fill('D6EAF8')
F_AUTO   = fill('D5F5E3')
F_MANUAL = fill('FDFEFE')
F_JIM    = fill('FFFDE7')
F_WARN   = fill('FFF3CD')

def write_ws(ws, type_key, display, sname, hid, fields, rates, total_files, variants):
    # Tab color
    tab_colors={'tongcai':'70AD47','cutting':'FF0000','atom':'7030A0',
                'computer_stitch':'4472C4','stitching':'ED7D31',
                'assembly_1':'FFC000','assembly_2':'FF8C00',
                'sum_c2b':'5B9BD5','stock':'70AD47','dacui':'A9A9A9','xishi':'00B0F0'}
    ws.sheet_properties.tabColor=tab_colors.get(type_key,'888888')

    # Row 1: sheet meta
    ws.row_dimensions[1].height=18
    meta=f'類型: {display}  |  樣本: header_id={hid} sheet="{sname}"  |  總檔數: {total_files}  |  生成: {datetime.now():%Y-%m-%d}'
    ws.merge_cells('A1:H1')
    c=ws['A1']; c.value=meta; c.font=font(bold=True,color='FFFFFF')
    c.fill=F_HEADER; c.alignment=Alignment(horizontal='left',vertical='center')

    # Row 2: column headers
    ws.row_dimensions[2].height=32
    for i,(hdr,w) in enumerate(zip(HDR_COLS,COL_WIDTHS),1):
        c=ws.cell(2,i,hdr)
        c.font=font(bold=True,color='FFFFFF')
        c.fill=fill('2E4057')
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width=w

    # Data rows
    for ri,f in enumerate(fields,3):
        ws.row_dimensions[ri].height=15
        pos=f['pos']; name=f['name']; sample=f['sample']
        formula=f['formula']; keep=f['keep']; itype=f['itype']
        rate=rates.get(f['col_num'],0) if not f['is_title'] else ''
        variant_note=f' ⚠ {f["vc"]}種變體' if f['vc']>0 else ''
        formula_disp=(formula+variant_note) if formula else (variant_note or '')

        if f['is_title']:
            row_fill=F_TITLE
        elif '公式' in itype:
            row_fill=F_FORM
        elif '自動' in itype:
            row_fill=F_AUTO
        else:
            row_fill=F_MANUAL
        if f['vc']>0: row_fill=F_WARN

        vals=[pos, name, sample, formula_disp, keep, itype,
              f'{rate}%' if rate!='' else '', '']
        for ci,v in enumerate(vals,1):
            c=ws.cell(ri,ci,v)
            c.fill = F_JIM if ci==8 else row_fill
            c.font=font(size=10)
            c.border=thin_border()
            c.alignment=Alignment(vertical='center', wrap_text=(ci in (2,4)))
            if ci==5:
                c.font=font(bold=(keep=='要'),color='155724' if keep=='要' else '721C24')
            if ci==7 and rate!='':
                c.font=font(color='155724' if int(rate or 0)>=80 else ('856404' if int(rate or 0)>=50 else 'C0392B'))

    # Freeze header rows
    ws.freeze_panes='A3'

# ── main ─────────────────────────────────────────────────────────────────────

def main():
    conn=get_conn()
    wb=openpyxl.Workbook()
    wb.remove(wb.active)

    type_display={r['type_key']:r['display_name']
                  for r in conn.execute('SELECT type_key,display_name FROM sheet_types').fetchall()}

    all_variants={}
    report_lines=[]
    report_lines.append('IE Sheet 規格定義工作簿 — 生成報告')
    report_lines.append(f'生成時間: {datetime.now():%Y-%m-%d %H:%M}')
    report_lines.append('='*80)
    report_lines.append(f'{"分頁":<22} {"欄位數":>6} {"含公式":>8} {"公式變體警告":>12} {"樣本來源":<20}')
    report_lines.append('-'*80)

    for tk in COMMON_TYPES:
        display=type_display.get(tk,tk)
        # Load grid
        sname,grid=load_grid(conn,TARGET,tk)
        used_hid=TARGET
        if grid is None:
            alt=best_alt_header(conn,tk)
            if alt:
                used_hid=alt
                sname,grid=load_grid(conn,used_hid,tk)
        if grid is None:
            report_lines.append(f'  {display:<22} {"(無資料跳過)":>40}')
            continue

        header_row=detect_header_row(grid)
        data_start=detect_data_start(grid,header_row)
        fields=analyze_fields(grid,header_row,data_start)

        rates,total=appearance_rates(conn,tk,header_row)
        # Apply rates to fields
        for f in fields:
            if not f['is_title']:
                f['rate_pct']=rates.get(f['col_num'],0)

        # Formula variants
        fvars=formula_variants(conn,tk,header_row)
        all_variants[tk]={'display':display,'variants':fvars,'header_row':header_row}
        # Annotate fields
        var_count=0
        for f in fields:
            if not f['is_title'] and f['col_num'] in fvars:
                f['vc']=len(fvars[f['col_num']])
                var_count+=1

        safe_title=re.sub(r'[/\\?\*\[\]:]','',display)[:31]
        ws=wb.create_sheet(title=safe_title)
        write_ws(ws,tk,display,sname,used_hid,fields,rates,total,fvars)

        data_fields=[f for f in fields if not f['is_title']]
        formula_fields=[f for f in data_fields if '公式' in f['itype']]
        src='主樣本 hid=32' if used_hid==TARGET else f'替代 hid={used_hid}'
        report_lines.append(f'  {display:<22} {len(data_fields):>6} {len(formula_fields):>8} {var_count:>12}  {src}')

    # ── formula_variants.txt ──────────────────────────────────────────────────
    with open(OUT_TXT,'w',encoding='utf-8') as f:
        f.write('formula_variants.txt — 公式變體明細\n')
        f.write(f'生成時間: {datetime.now():%Y-%m-%d %H:%M}\n')
        f.write('='*80+'\n\n')
        for tk in COMMON_TYPES:
            info=all_variants.get(tk)
            if not info or not info['variants']: continue
            f.write(f'【{info["display"]}】({tk})\n')
            f.write('-'*60+'\n')
            for col_n, variants in sorted(info['variants'].items()):
                total_files=sum(c for _,c in variants)
                f.write(f'  欄 {cletter(col_n)}: {len(variants)} 種變體 (共 {total_files} 檔)\n')
                for nf,cnt in variants:
                    pct=round(cnt/total_files*100)
                    f.write(f'    [{pct:>3}% / {cnt:>3}檔] {nf}\n')
            f.write('\n')

    wb.save(OUT_XL)
    conn.close()

    report_lines.append('='*80)
    report_lines.append(f'輸出: {OUT_XL}')
    report_lines.append(f'附檔: {OUT_TXT}')
    print('\n'.join(report_lines))

main()
