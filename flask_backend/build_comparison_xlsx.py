#!/usr/bin/env python3
"""
Build comparison_table.xlsx from existing result_full.txt + missing/mismatch files.
Does NOT require locking the 廠務組織編制表 file.

Output: flask_backend/test_output/comparison_table.xlsx
"""
import sys, os, re

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEST_DIR = os.path.join(os.path.dirname(__file__), 'test_output')
RESULT_TXT  = os.path.join(TEST_DIR, 'result_full.txt')
MISSING_TXT = os.path.join(TEST_DIR, 'missing_ie_files.txt')
MISMATCH_TXT= os.path.join(TEST_DIR, 'mp_mismatch.txt')
OUTPUT      = os.path.join(TEST_DIR, 'comparison_table.xlsx')

RED_FILL    = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
GREEN_FILL  = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
GREY_FILL   = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
HEADER_FILL = PatternFill(start_color='17375E', end_color='17375E', fill_type='solid')
WHITE_FONT  = Font(color='FFFFFF', bold=True, size=10)
BOLD_FONT   = Font(bold=True)

THIN = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = [
    'LEAN', '鞋型', 'ART', '訂單',
    '計算Cut', '結果表Cut',
    '計算Stitch', '結果表Stitch',
    '計算Asm', '結果表Asm',
    '狀態'
]
COL_WIDTHS = [7, 32, 12, 8, 10, 10, 12, 12, 10, 10, 12]


def parse_missing(path):
    """Return {art: {lean, qty, ref_cut, ref_s, ref_asm, model}} from missing_ie_files.txt."""
    data = {}
    if not os.path.exists(path):
        return data
    with open(path, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#') or line.startswith('ART') or line.startswith('-'):
                continue
            parts = line.split()
            if len(parts) < 3:
                continue
            art = parts[0]
            lean = parts[1]
            try: qty = int(parts[2])
            except: qty = 0
            def _fv(s):
                try: return float(s)
                except: return None
            ref_cut = _fv(parts[3]) if len(parts) > 3 else None
            ref_s   = _fv(parts[4]) if len(parts) > 4 else None
            ref_asm = _fv(parts[5]) if len(parts) > 5 else None
            model = ' '.join(parts[6:]) if len(parts) > 6 else ''
            data[art] = {
                'lean': lean, 'qty': qty,
                'ref_cut': ref_cut, 'ref_s': ref_s, 'ref_asm': ref_asm,
                'model': model,
                'calc_cut': None, 'calc_s': None, 'calc_asm': None,
                'status': 'MISSING_IE',
            }
    return data


def parse_mismatch(path):
    """Return {art: {lean, qty, calc_cut, ref_cut, ...}} from mp_mismatch.txt."""
    data = {}
    if not os.path.exists(path):
        return data
    with open(path, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#') or line.startswith('ART') or line.startswith('-'):
                continue
            parts = line.split()
            if len(parts) < 5:
                continue
            art = parts[0]
            lean = parts[1]
            try: qty = int(parts[2])
            except: qty = 0
            def _fv(s):
                try: return float(s) if s != '-' else None
                except: return None
            cc  = _fv(parts[3]) if len(parts) > 3 else None
            rc  = _fv(parts[4]) if len(parts) > 4 else None
            cs  = _fv(parts[6]) if len(parts) > 6 else None
            rs  = _fv(parts[7]) if len(parts) > 7 else None
            ca  = _fv(parts[9]) if len(parts) > 9 else None
            ra  = _fv(parts[10]) if len(parts) > 10 else None
            model = ' '.join(parts[12:]) if len(parts) > 12 else ''
            # Avoid duplicate (same ART may appear in multiple groups)
            key = f'{art}|{lean}'
            data[key] = {
                'lean': lean, 'art': art, 'qty': qty,
                'calc_cut': cc, 'ref_cut': rc,
                'calc_s': cs, 'ref_s': rs,
                'calc_asm': ca, 'ref_asm': ra,
                'model': model,
                'status': 'MISMATCH',
            }
    return data


def parse_result_full(path):
    """Parse CSA section from result_full.txt for OK rows."""
    ok_rows = {}
    if not os.path.exists(path):
        return ok_rows
    in_csa = False
    in_diff = False
    with open(path, encoding='utf-8') as f:
        for line in f:
            line = line.rstrip()
            if '=== CSA' in line:
                in_csa = True
                continue
            if '=== 差異清單' in line or '===' in line and in_csa:
                if '差異' in line:
                    in_diff = True
                    in_csa = False
                continue
            if not in_csa:
                continue
            # Data rows: Sheet  ART  Model  Qty  Cut  Stitch  Asm  Total  MP?
            parts = line.split()
            if len(parts) < 8:
                continue
            # Skip header/separator
            if parts[0] in ('Sheet', '---'):
                continue
            # Check if MP column is Y (ok)
            if parts[-1] not in ('Y', 'N', '✓', '✗'):
                continue
            mp_ok = parts[-1] in ('Y', '✓')
            if not mp_ok:
                continue
            # Find the ART (second token is ART if it matches pattern)
            art_re = re.compile(r'^[A-Z]{2}\d{4,6}$')
            if len(parts) >= 9 and art_re.match(parts[1]):
                art = parts[1]
                def _fv(s):
                    try: return float(s)
                    except: return None
                ok_rows[art] = {
                    'sheet': parts[0],
                    'art': art,
                    'calc_cut': _fv(parts[4]),
                    'calc_s':   _fv(parts[5]),
                    'calc_asm': _fv(parts[6]),
                }
    return ok_rows


def main():
    missing = parse_missing(MISSING_TXT)   # art → row info
    mismatch = parse_mismatch(MISMATCH_TXT) # art|lean → row info

    print(f'Missing IE: {len(missing)} ARTs')
    print(f'Mismatch:   {len(mismatch)} entries')

    # Build unified row list
    rows = []

    # 1. MISMATCH rows (have IE but values differ)
    for key, r in mismatch.items():
        rows.append({
            'lean':        r['lean'],
            'model':       r.get('model', ''),
            'art':         r['art'],
            'qty':         r['qty'],
            'calc_cut':    r.get('calc_cut'),
            'ref_cut':     r.get('ref_cut'),
            'calc_stitch': r.get('calc_s'),
            'ref_stitch':  r.get('ref_s'),
            'calc_asm':    r.get('calc_asm'),
            'ref_asm':     r.get('ref_asm'),
            'status':      'MISMATCH',
        })

    # 2. MISSING_IE rows (no IE file → no calculated MP)
    for art, r in missing.items():
        # Check if this art now has ob_epph (was just imported)
        import database as db
        db.init_db()
        conn = db.get_conn()
        epph = conn.execute(
            '''SELECT e.cutting, e.stitching, e.assembly
               FROM ob_header h JOIN ob_epph e ON e.header_id=h.id
               WHERE h.art=? ORDER BY h.id DESC LIMIT 1''', (art,)
        ).fetchone()
        conn.close()

        if epph and any(v for v in epph):
            status = 'MISMATCH'  # Now has data but still differs from result table
            calc_cut, calc_s, calc_asm = epph[0], epph[1], epph[2]
        else:
            status = 'MISSING_IE'
            calc_cut = calc_s = calc_asm = None

        rows.append({
            'lean':        r['lean'],
            'model':       r.get('model', ''),
            'art':         art,
            'qty':         r['qty'],
            'calc_cut':    calc_cut,
            'ref_cut':     r.get('ref_cut'),
            'calc_stitch': calc_s,
            'ref_stitch':  r.get('ref_s'),
            'calc_asm':    calc_asm,
            'ref_asm':     r.get('ref_asm'),
            'status':      status,
        })

    # Sort by LEAN then ART
    def sort_key(r):
        lean = r.get('lean', '')
        art  = r.get('art', '')
        # Extract LEAN number for sorting
        m = re.match(r'(\d+)([A-Z]?)', lean)
        lean_num = (int(m.group(1)), m.group(2) or '') if m else (99, lean)
        return (lean_num, art)

    rows.sort(key=sort_key)

    # Write Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '比對表'

    # Header
    for col, (hdr, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, r in enumerate(rows, 2):
        status = r['status']
        values = [
            r.get('lean', ''),
            (r.get('model') or '')[:40],
            r.get('art', ''),
            r.get('qty') or 0,
            r.get('calc_cut'),
            r.get('ref_cut'),
            r.get('calc_stitch'),
            r.get('ref_stitch'),
            r.get('calc_asm'),
            r.get('ref_asm'),
            status,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='left' if col == 2 else 'center')

            if status == 'MISSING_IE':
                cell.fill = YELLOW_FILL
            elif status == 'MISMATCH':
                # Highlight only mismatched value cells
                pairs = {5: ('calc_cut', 6), 7: ('calc_stitch', 8), 9: ('calc_asm', 10)}
                ref_pairs = {6: ('ref_cut', 5), 8: ('ref_stitch', 7), 10: ('ref_asm', 9)}
                if col in pairs:
                    dept, ref_col = pairs[col]
                    cv = r.get(dept)
                    rv = r.get({6:'ref_cut',8:'ref_stitch',10:'ref_asm'}.get(ref_col, ''))
                    if cv is None or (rv is not None and abs((cv or 0) - (rv or 0)) > 0.5):
                        cell.fill = RED_FILL
                elif col in ref_pairs:
                    dept, calc_col = ref_pairs[col]
                    rv = r.get(dept)
                    cv = r.get({5:'calc_cut',7:'calc_stitch',9:'calc_asm'}.get(calc_col, ''))
                    if rv is None or (cv is not None and abs((cv or 0) - (rv or 0)) > 0.5):
                        cell.fill = RED_FILL
                elif col == 11:
                    cell.fill = RED_FILL
            elif status == 'NO_REF':
                cell.fill = GREY_FILL

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}1'

    # Summary sheet
    ws2 = wb.create_sheet('摘要')
    sc = {}
    for r in rows:
        sc[r['status']] = sc.get(r['status'], 0) + 1

    ws2.append(['狀態', '筆數', '說明'])
    ws2.append(['MISMATCH',   sc.get('MISMATCH', 0),   '有IE文件但計算值≠結果表（待Jim確認分配邏輯）'])
    ws2.append(['MISSING_IE', sc.get('MISSING_IE', 0), '仍無IE文件對應'])
    ws2.append(['合計',       len(rows), ''])
    ws2.append([])
    ws2.append(['注意', '結果表 MP 為各組分配值，計算值為整條產線 MP，邏輯待確認'])
    ws2.append(['生成時間', '2026-06-04'])

    for col in (1, 2, 3):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    wb.save(OUTPUT)
    print(f'Written: {OUTPUT}')
    print(f'Total rows: {len(rows)}')
    for k in ['MISMATCH', 'MISSING_IE', 'OK', 'NO_REF']:
        if k in sc:
            print(f'  {k}: {sc[k]}')


if __name__ == '__main__':
    main()
