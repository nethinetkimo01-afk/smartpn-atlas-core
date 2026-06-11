#!/usr/bin/env python3
"""
IE Stage 1 — 全量導入 + sheet 正規化 + ART×Sheet 矩陣
Part 1: schema 擴充
Part 2: 雙讀導入引擎 (formula + value)
Part 3: sheet 名稱正規化
Part 4: ART×Sheet 三態矩陣
Part 5: 報告輸出
"""
import sys, os, re, glob, sqlite3, shutil, traceback
from datetime import datetime
from statistics import mean

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

DB_PATH    = os.path.join(os.path.dirname(__file__), 'data', 'atlas.db')
IE_DIRS    = [
    r"C:\Users\user\OneDrive\Desktop\IE",
    r"C:\Users\user\OneDrive\Desktop\Biên chế\Jun\IE",
]
OUT_DIR    = os.path.join(os.path.dirname(__file__), 'test_output')
FAIL_LOG   = os.path.join(OUT_DIR, 'ie_import_failures.txt')
UNMAP_LOG  = os.path.join(OUT_DIR, 'sheet_unmapped.txt')
MATRIX_XL  = os.path.join(OUT_DIR, 'ie_matrix.xlsx')
REPORT_OUT = os.path.join(OUT_DIR, 'ie_stage1_report.txt')

_ART_RE    = re.compile(r'[A-Z]{2}\d{4,6}')
_ART_SUFFIX_RE = re.compile(
    r'[_\s]+(?:[A-Z]{2}\d{4,6}(?:-\d+)*(?:[_,\s]+[A-Z]{2}\d{4,6}(?:-\d+)*)*)\s*$',
    re.IGNORECASE
)

_SKIP_FILENAMES = {
    '120双_FW26_GHOST SPRINT W_HQ3330..xlsx',
    '120双 LA TRAINER OG-KH9682-83-84- vải lưới+da lộn+da giả (Recovered).xlsx',
    '60双_SS25 TOKYO MJ_KI5323_da thật (da cá)_thêm 1 ng mài thô mg - Copy.xlsx',
}

os.makedirs(OUT_DIR, exist_ok=True)

# ── DB helpers ────────────────────────────────────────────────────────────────

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')
    conn.execute('PRAGMA journal_mode = WAL')
    return conn

# ─────────────────────────────────────────────────────────────────────────────
# PART 1: Schema migration
# ─────────────────────────────────────────────────────────────────────────────

def part1_schema():
    print("\n=== PART 1: Schema migration ===")
    conn = get_conn()

    # Add cell_type to ie_sheet_data if missing
    cols = [r[1] for r in conn.execute("PRAGMA table_info(ie_sheet_data)").fetchall()]
    if 'cell_type' not in cols:
        conn.execute("ALTER TABLE ie_sheet_data ADD COLUMN cell_type TEXT")
        print("  Added cell_type column to ie_sheet_data")
    else:
        print("  cell_type already exists")

    # sheet_types table
    conn.execute('''CREATE TABLE IF NOT EXISTS sheet_types (
        id           INTEGER PRIMARY KEY AUTOINCREMENT,
        type_key     TEXT NOT NULL UNIQUE,
        display_name TEXT NOT NULL,
        sort_order   INTEGER DEFAULT 99
    )''')

    # sheet_name_map: raw_name → type_key
    conn.execute('''CREATE TABLE IF NOT EXISTS sheet_name_map (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        raw_name   TEXT NOT NULL UNIQUE,
        normalized TEXT NOT NULL,
        type_key   TEXT,
        FOREIGN KEY(type_key) REFERENCES sheet_types(type_key)
    )''')

    # art_sheet_status table
    conn.execute('''CREATE TABLE IF NOT EXISTS art_sheet_status (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        header_id  INTEGER NOT NULL REFERENCES ob_header(id) ON DELETE CASCADE,
        art        TEXT NOT NULL,
        sheet_type TEXT NOT NULL,
        status     TEXT NOT NULL DEFAULT 'has_data',
        UNIQUE(header_id, sheet_type)
    )''')

    # Indexes
    conn.execute('CREATE INDEX IF NOT EXISTS idx_sheet_name_map_raw ON sheet_name_map(raw_name)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_art_sheet_status_hdr ON art_sheet_status(header_id)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_art_sheet_status_art ON art_sheet_status(art)')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_ie_sheet_data_type ON ie_sheet_data(cell_type)')

    conn.commit()
    conn.close()
    print("  Schema migration done.")

# ─────────────────────────────────────────────────────────────────────────────
# PART 2: Dual-read import engine
# ─────────────────────────────────────────────────────────────────────────────

def fn_eolr(filename):
    base = os.path.basename(filename)
    m = re.search(r'(\d+)双', base)
    if m:
        return int(m.group(1))
    return None


def find_header_for_arts(conn, arts, eolr_hint):
    """Return (header_id, matched_art) using arts list + eolr_hint.
    Prefer eolr match. Fall back to any header for these arts.
    Returns (None, None) if nothing found.
    """
    for art in arts:
        rows = conn.execute(
            '''SELECT h.id, h.eolr FROM ob_articles a
               JOIN ob_header h ON h.id=a.header_id
               WHERE a.art=? ORDER BY h.updated_at DESC, h.id DESC''',
            (art,)
        ).fetchall()
        if not rows:
            continue
        if eolr_hint:
            exact = [r for r in rows if r['eolr'] == eolr_hint]
            if exact:
                return exact[0]['id'], art
        return rows[0]['id'], art
    return None, None


def import_one_file(xlsx_path, conn, stats):
    """Import one file. Returns header_id or None on failure."""
    import openpyxl
    fname = os.path.basename(xlsx_path)

    arts = _ART_RE.findall(fname)
    if not arts:
        stats['orphan_no_art'] += 1
        stats['failures'].append((fname, 'no ART in filename'))
        return None

    eolr_hint = fn_eolr(fname)
    header_id, matched_art = find_header_for_arts(conn, arts, eolr_hint)

    if header_id is None:
        stats['orphan_no_header'] += 1
        stats['orphans'].append({'file': fname, 'arts': arts, 'eolr': eolr_hint})
        return None

    # Delete existing data for this header_id (re-import)
    conn.execute('DELETE FROM ie_sheet_data WHERE header_id=?', (header_id,))

    # Pass 1: formulas
    try:
        wb_f = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=False)
        sheet_names = list(wb_f.sheetnames)
        formula_map = {}  # (sheet, row, col) -> formula
        for sn in sheet_names:
            ws = wb_f[sn]
            for row in ws.iter_rows():
                for cell in row:
                    if hasattr(cell, 'value') and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_map[(sn, cell.row, cell.column)] = cell.value
        wb_f.close()
    except Exception as e:
        stats['failures'].append((fname, f'Pass1 error: {e}'))
        return None

    # Pass 2: computed values
    try:
        wb_v = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet_rows = []
        for sn in sheet_names:
            if sn not in wb_v.sheetnames:
                continue
            ws = wb_v[sn]
            for row in ws.iter_rows():
                for cell in row:
                    if not hasattr(cell, 'row'):
                        continue
                    v = cell.value
                    formula = formula_map.get((sn, cell.row, cell.column))
                    if v is None and formula is None:
                        continue
                    val_str = None if v is None else str(v)
                    if formula:
                        cell_type = 'formula'
                    elif val_str is not None:
                        cell_type = 'manual'
                    else:
                        cell_type = 'empty'
                    sheet_rows.append((header_id, sn, cell.row, cell.column, val_str, formula, cell_type))
        wb_v.close()
    except Exception as e:
        stats['failures'].append((fname, f'Pass2 error: {e}'))
        return None

    # Write to DB in single transaction
    try:
        conn.executemany(
            'INSERT INTO ie_sheet_data (header_id, sheet_name, row, col, value, formula, cell_type) VALUES (?,?,?,?,?,?,?)',
            sheet_rows
        )
        stats['ok_files'] += 1
        stats['total_cells'] += len(sheet_rows)
        stats['formula_cells'] += sum(1 for r in sheet_rows if r[6] == 'formula')
        stats['manual_cells'] += sum(1 for r in sheet_rows if r[6] == 'manual')
        stats['imported_headers'].add(header_id)
    except Exception as e:
        conn.rollback()
        stats['failures'].append((fname, f'DB write error: {e}'))
        return None

    return header_id


def part2_import():
    print("\n=== PART 2: Full IE import ===")
    import openpyxl

    # Collect all xlsx files
    all_files = []
    for d in IE_DIRS:
        if os.path.isdir(d):
            for f in glob.glob(os.path.join(d, '**', '*.xlsx'), recursive=True):
                bn = os.path.basename(f)
                if not bn.startswith('~$') and bn not in _SKIP_FILENAMES:
                    all_files.append(f)

    all_files.sort()
    print(f"  Total xlsx files: {len(all_files)}")

    stats = {
        'ok_files': 0,
        'orphan_no_art': 0,
        'orphan_no_header': 0,
        'total_cells': 0,
        'formula_cells': 0,
        'manual_cells': 0,
        'failures': [],
        'orphans': [],
        'imported_headers': set(),
    }

    conn = get_conn()
    for i, fpath in enumerate(all_files):
        if (i + 1) % 20 == 0:
            print(f"  [{i+1}/{len(all_files)}] processing...")
        try:
            import_one_file(fpath, conn, stats)
        except Exception as e:
            fname = os.path.basename(fpath)
            stats['failures'].append((fname, f'unexpected: {e}\n{traceback.format_exc()[:300]}'))

    conn.commit()
    conn.close()

    print(f"  OK: {stats['ok_files']} files, {len(stats['imported_headers'])} headers")
    print(f"  Orphan (no ART): {stats['orphan_no_art']}, Orphan (no header): {stats['orphan_no_header']}")
    print(f"  Failures: {len(stats['failures'])}")
    print(f"  Cells: total={stats['total_cells']} formula={stats['formula_cells']} manual={stats['manual_cells']}")

    # Write failure log
    with open(FAIL_LOG, 'w', encoding='utf-8') as f:
        f.write(f"IE Import Failures — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Total: {len(stats['failures'])} failures\n\n")
        for fname, reason in stats['failures']:
            f.write(f"{fname}|{reason}\n")
        if stats['orphans']:
            f.write(f"\n--- ORPHANS (no matching header) ---\n")
            for o in stats['orphans']:
                f.write(f"{o['file']}|ARTs={o['arts']}|eolr={o['eolr']}\n")

    return stats

# ─────────────────────────────────────────────────────────────────────────────
# PART 3: Sheet name normalization
# ─────────────────────────────────────────────────────────────────────────────

# Standard sheet types with matching keywords
SHEET_TYPE_RULES = [
    # (type_key, display_name, sort_order, keyword_list)
    ('tongcai',       '同材共裁',          1,  ['同材', 'tongcai', 'tong cai', '共裁']),
    ('cutting',       'Cutting 裁斷',      2,  ['cutting', 'cut', 'cắt', '裁', 'da thật', 'da lộn', 'da giả', 'vải', 'tổng hợp', 'mesh', 'thật']),
    ('atom',          'ATOM 自動化',        3,  ['atom', '自动', '自動', 'automat', 'tự động']),
    ('computer_stitch','電腦針車 AC',       4,  ['ac', '电脑', '電腦', 'computer', 'needle', 'máy', 'cs']),
    ('stitching',     'Stitching 針車',    5,  ['stitch', 'may', '針車', '縫', '针车', 'sewing']),
    ('sub_stitching', 'Sub-Stitching 支流', 6, ['sub.stitch', 'sub_stitch', 'sub stitch', '支流', 'zhi', 'tributary']),
    ('assembly_1',    'Assembly.1 成型1',  7,  ['assembly.1', 'assembly 1', 'assembly1', 'lắp ráp 1', 'ráp 1', 'assembly_1']),
    ('assembly_2',    'Assembly.2 成型2',  8,  ['assembly.2', 'assembly 2', 'assembly2', 'lắp ráp 2', 'ráp 2', 'assembly_2']),
    ('assembly',      'Assembly 成型',     9,  ['assembly', 'lắp ráp', 'ráp ráp', 'chenxing', '成型']),
    ('sum_c2b',       'SUM C2B',          10,  ['sum.c2b', 'sum_c2b', 'sumc2b', 'sum c2b']),
    ('stock',         'Stock/Lót',        11,  ['stock', 'lót', 'lot', 'sock', 'inner']),
    ('dacui',         '打粗',             12,  ['打粗', 'da cui', 'rough', 'dacui']),
    ('xishi',         '水洗',             13,  ['水洗', 'xi shi', 'wash', 'xishi']),
    ('tiedadi',       '貼大底',           14,  ['贴大底', '貼大底', 'sole', 'tiedadi', 'cement']),
    ('zhaoshe',       '照射',             15,  ['照射', 'irradiat', 'uv', 'zhaoshe']),
]

def normalize_sheet_name(raw):
    """Strip whitespace, fullwidth chars, ART suffixes. Return normalized lowercase."""
    s = raw.strip()
    # Remove fullwidth spaces
    s = s.replace('　', ' ').strip()
    # Remove ART suffixes like _KH9682-83, _IH1651, etc.
    s = re.sub(r'[_\s]+[A-Z]{2}\d{4,6}(?:[_\-,]\d+)*(?:[_\s]+[A-Z]{2}\d{4,6}(?:[_\-,]\d+)*)*\s*$', '', s)
    # Remove parenthetical material notes that are common
    s = re.sub(r'\s*\(da thật\)', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s*\(da lộn\)', '', s, flags=re.IGNORECASE)
    # Normalize SUM.C2B / SUM_C2B → SUM.C2B
    s = re.sub(r'SUM[_\.]C2B', 'SUM.C2B', s, flags=re.IGNORECASE)
    # Normalize Sub.Stitching / Sub Stitching
    s = re.sub(r'Sub[\._\s]Stitch', 'Sub.Stitch', s, flags=re.IGNORECASE)
    # Strip trailing space again
    s = s.strip()
    return s


def classify_sheet(normalized):
    """Return type_key for a normalized sheet name, or None."""
    nl = normalized.lower()
    for type_key, _dn, _so, keywords in SHEET_TYPE_RULES:
        for kw in keywords:
            if kw.lower() in nl:
                return type_key
    return None


def part3_normalize():
    print("\n=== PART 3: Sheet name normalization ===")
    conn = get_conn()

    # Upsert standard sheet types
    for type_key, display_name, sort_order, _ in SHEET_TYPE_RULES:
        conn.execute(
            '''INSERT INTO sheet_types (type_key, display_name, sort_order)
               VALUES (?,?,?)
               ON CONFLICT(type_key) DO UPDATE SET display_name=excluded.display_name, sort_order=excluded.sort_order''',
            (type_key, display_name, sort_order)
        )

    # Get all distinct sheet names from ie_sheet_data
    rows = conn.execute(
        'SELECT DISTINCT sheet_name FROM ie_sheet_data ORDER BY sheet_name'
    ).fetchall()
    raw_names = [r['sheet_name'] for r in rows]
    print(f"  Distinct raw sheet names: {len(raw_names)}")

    mapped   = {}   # raw -> (normalized, type_key)
    unmapped = []

    for raw in raw_names:
        norm = normalize_sheet_name(raw)
        tk   = classify_sheet(norm) or classify_sheet(raw)
        if tk is None:
            unmapped.append(raw)
        mapped[raw] = (norm, tk)

    # Upsert into sheet_name_map
    for raw, (norm, tk) in mapped.items():
        conn.execute(
            '''INSERT INTO sheet_name_map (raw_name, normalized, type_key)
               VALUES (?,?,?)
               ON CONFLICT(raw_name) DO UPDATE SET normalized=excluded.normalized, type_key=excluded.type_key''',
            (raw, norm, tk)
        )

    conn.commit()
    conn.close()

    # Write unmapped log
    with open(UNMAP_LOG, 'w', encoding='utf-8') as f:
        f.write(f"Sheet names that could not be auto-classified ({len(unmapped)}):\n")
        f.write("Please assign a type_key from sheet_types table.\n\n")
        for n in sorted(unmapped):
            f.write(f"  {n!r}\n")

    print(f"  Mapped: {len(mapped) - len(unmapped)}/{len(mapped)}")
    print(f"  Unmapped: {len(unmapped)} → {UNMAP_LOG}")

    # Coverage per type
    conn2 = get_conn()
    coverage = {}
    total_headers = conn2.execute('SELECT COUNT(DISTINCT id) FROM ob_header').fetchone()[0]
    type_rows = conn2.execute('SELECT type_key, display_name FROM sheet_types ORDER BY sort_order').fetchall()
    for tr in type_rows:
        tk = tr['type_key']
        cnt = conn2.execute(
            '''SELECT COUNT(DISTINCT sd.header_id) FROM ie_sheet_data sd
               JOIN sheet_name_map snm ON snm.raw_name=sd.sheet_name
               WHERE snm.type_key=?''',
            (tk,)
        ).fetchone()[0]
        coverage[tk] = cnt
    conn2.close()

    return mapped, unmapped, coverage, total_headers


# ─────────────────────────────────────────────────────────────────────────────
# PART 4: ART × Sheet status matrix
# ─────────────────────────────────────────────────────────────────────────────

def part4_matrix(coverage, total_headers):
    print("\n=== PART 4: ART×Sheet matrix ===")
    conn = get_conn()

    # Determine "common" sheet types (>50% coverage across headers that have ANY data)
    headers_with_data = conn.execute(
        'SELECT COUNT(DISTINCT header_id) FROM ie_sheet_data'
    ).fetchone()[0]

    common_types = []
    for tk, cnt in coverage.items():
        if headers_with_data > 0 and cnt / headers_with_data > 0.50:
            common_types.append(tk)

    print(f"  Headers with data: {headers_with_data}")
    print(f"  Common sheet types (>50%): {common_types}")

    # Clear and rebuild art_sheet_status
    conn.execute('DELETE FROM art_sheet_status')

    # For each header with data, build status
    h_rows = conn.execute(
        '''SELECT DISTINCT sd.header_id, a.art
           FROM ie_sheet_data sd
           JOIN ob_articles a ON a.header_id=sd.header_id
           ORDER BY sd.header_id, a.art'''
    ).fetchall()

    # Get which sheet_types each header actually has
    hdr_types = {}  # header_id -> set of type_keys
    type_rows = conn.execute(
        '''SELECT sd.header_id, snm.type_key, COUNT(*) as manual_cnt
           FROM ie_sheet_data sd
           JOIN sheet_name_map snm ON snm.raw_name=sd.sheet_name
           WHERE snm.type_key IS NOT NULL AND sd.cell_type='manual'
           GROUP BY sd.header_id, snm.type_key'''
    ).fetchall()
    for tr in type_rows:
        hid = tr['header_id']
        if hid not in hdr_types:
            hdr_types[hid] = {}
        hdr_types[hid][tr['type_key']] = tr['manual_cnt']

    # Also track has_data (any cells) per type
    hdr_has = {}  # header_id -> set of type_keys
    has_rows = conn.execute(
        '''SELECT DISTINCT sd.header_id, snm.type_key
           FROM ie_sheet_data sd
           JOIN sheet_name_map snm ON snm.raw_name=sd.sheet_name
           WHERE snm.type_key IS NOT NULL'''
    ).fetchall()
    for hr in has_rows:
        hid = hr['header_id']
        if hid not in hdr_has:
            hdr_has[hid] = set()
        hdr_has[hid].add(hr['type_key'])

    # Get unique headers
    headers_done = {}  # header_id -> set of arts
    for row in h_rows:
        hid = row['header_id']
        if hid not in headers_done:
            headers_done[hid] = set()
        headers_done[hid].add(row['art'])

    inserts = []
    missing_count = 0
    for hid, arts in headers_done.items():
        for art in arts:
            for tk in common_types:
                has = tk in hdr_has.get(hid, set())
                if has:
                    status = 'has_data'
                else:
                    status = 'missing'
                    missing_count += 1
                inserts.append((hid, art, tk, status))

    conn.executemany(
        'INSERT OR IGNORE INTO art_sheet_status (header_id, art, sheet_type, status) VALUES (?,?,?,?)',
        inserts
    )
    conn.commit()

    print(f"  art_sheet_status rows: {len(inserts)}")
    print(f"  missing entries: {missing_count}")

    # Build matrix Excel
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'ART×Sheet Matrix'

        # Column order: arts sorted, then sheet types
        all_types_ordered = conn.execute(
            'SELECT type_key, display_name FROM sheet_types ORDER BY sort_order'
        ).fetchall()

        hdr_fill   = PatternFill('solid', fgColor='1A2744')
        hdr_font   = Font(bold=True, color='FFFFFF', size=9)
        ok_fill    = PatternFill('solid', fgColor='D4EDDA')
        miss_fill  = PatternFill('solid', fgColor='F8D7DA')
        gray_fill  = PatternFill('solid', fgColor='F5F5F5')
        thin       = Side(style='thin', color='CCCCCC')
        bord       = Border(left=thin, right=thin, top=thin, bottom=thin)
        center_al  = Alignment(horizontal='center', vertical='center')

        # Headers row
        ws.cell(row=1, column=1, value='ART').font = hdr_font
        ws.cell(row=1, column=1).fill = hdr_fill
        ws.cell(row=1, column=1).alignment = center_al
        ws.cell(row=1, column=2, value='header_id').font = hdr_font
        ws.cell(row=1, column=2).fill = hdr_fill
        ws.cell(row=1, column=2).alignment = center_al

        col_map = {}  # type_key -> col index
        for ci, tr in enumerate(all_types_ordered, 3):
            tk = tr['type_key']
            dn = tr['display_name']
            c = ws.cell(row=1, column=ci, value=dn)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = center_al
            c.border = bord
            col_map[tk] = ci

        # Data rows
        # Get status for each (hid, art, type_key)
        status_map = {}  # (hid, art, tk) -> manual_cnt or 'missing'
        status_rows = conn.execute(
            'SELECT header_id, art, sheet_type, status FROM art_sheet_status'
        ).fetchall()
        for sr in status_rows:
            status_map[(sr['header_id'], sr['art'], sr['sheet_type'])] = sr['status']
        # manual_cnt map
        manual_map = {}  # (hid, tk) -> count
        for hid, types in hdr_types.items():
            for tk, cnt in types.items():
                manual_map[(hid, tk)] = cnt

        # Get all arts sorted
        arts_list = sorted(
            [(hid, art) for hid, arts in headers_done.items() for art in arts],
            key=lambda x: x[1]
        )

        for ri, (hid, art) in enumerate(arts_list, 2):
            ws.cell(row=ri, column=1, value=art)
            ws.cell(row=ri, column=2, value=hid)
            for tk, ci in col_map.items():
                if tk not in common_types:
                    # gray out non-common types
                    c = ws.cell(row=ri, column=ci, value='')
                    c.fill = gray_fill
                    c.border = bord
                    continue
                st = status_map.get((hid, art, tk))
                if st == 'has_data':
                    mc = manual_map.get((hid, tk), 0)
                    c = ws.cell(row=ri, column=ci, value=mc if mc > 0 else '✓')
                    c.fill = ok_fill
                elif st == 'missing':
                    c = ws.cell(row=ri, column=ci, value='✗')
                    c.fill = miss_fill
                else:
                    c = ws.cell(row=ri, column=ci, value='')
                    c.fill = gray_fill
                c.border = bord
                c.alignment = center_al

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 9
        for ci in range(3, len(all_types_ordered) + 3):
            col_letter = ws.cell(1, ci).column_letter
            ws.column_dimensions[col_letter].width = 14

        # Freeze panes
        ws.freeze_panes = 'C2'

        wb.save(MATRIX_XL)
        print(f"  Matrix saved: {MATRIX_XL}")
    except Exception as e:
        print(f"  WARNING: Excel matrix failed: {e}")

    conn.close()
    return missing_count, common_types


# ─────────────────────────────────────────────────────────────────────────────
# PART 5: Add /ie/matrix API + HTML routes
# ─────────────────────────────────────────────────────────────────────────────

def part5_report(import_stats, mapped, unmapped, coverage, total_headers, missing_count, common_types):
    print("\n=== PART 5: Report ===")
    conn = get_conn()

    total_cells = conn.execute('SELECT COUNT(*) FROM ie_sheet_data').fetchone()[0]
    formula_cnt = conn.execute("SELECT COUNT(*) FROM ie_sheet_data WHERE cell_type='formula'").fetchone()[0]
    manual_cnt  = conn.execute("SELECT COUNT(*) FROM ie_sheet_data WHERE cell_type='manual'").fetchone()[0]
    headers_cnt = conn.execute('SELECT COUNT(DISTINCT header_id) FROM ie_sheet_data').fetchone()[0]
    conn.close()

    lines = []
    W = 90
    lines.append('=' * W)
    lines.append('  IE Stage 1 全量導入報告')
    lines.append(f'  生成時間: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
    lines.append('=' * W)
    lines.append('')

    lines.append('─── Part 2: 導入結果 ─────────────────────────────────────────')
    lines.append(f'  成功導入檔案:     {import_stats["ok_files"]}')
    lines.append(f'  無 ART 孤兒:      {import_stats["orphan_no_art"]}')
    lines.append(f'  無 header 孤兒:   {import_stats["orphan_no_header"]}')
    lines.append(f'  失敗:             {len(import_stats["failures"])}')
    lines.append(f'  有資料的 header:  {headers_cnt} / {total_headers}')
    lines.append('')
    lines.append(f'  ie_sheet_data 總行數:   {total_cells:,}')
    lines.append(f'  formula 格數:           {formula_cnt:,}  ({formula_cnt/total_cells*100:.1f}% )' if total_cells else '  formula 格數: 0')
    lines.append(f'  manual 格數:            {manual_cnt:,}  ({manual_cnt/total_cells*100:.1f}% )' if total_cells else '  manual 格數: 0')
    lines.append('')

    lines.append('─── Part 3: sheet 類型覆蓋率 ─────────────────────────────────')
    lines.append(f'  總 ob_header 數: {total_headers}')
    lines.append(f'  有 IE 資料 header 數: {headers_cnt}')
    lines.append('')
    lines.append(f'  {"type_key":<20}  {"display":<20}  {"header_cnt":>10}  {"覆蓋率":>8}')
    lines.append('  ' + '─' * 65)
    for tk, cnt in sorted(coverage.items(), key=lambda x: -x[1]):
        is_common = '★' if tk in common_types else ' '
        base = headers_cnt if headers_cnt > 0 else 1
        pct = cnt / base * 100
        lines.append(f'  {is_common} {tk:<19}  {"":<20}  {cnt:>10}  {pct:>7.1f}%')
    lines.append('  (★ = 覆蓋率 >50%，列入矩陣 missing 判定)')
    lines.append('')

    lines.append('─── Part 4: 矩陣統計 ─────────────────────────────────────────')
    lines.append(f'  missing 總數:  {missing_count}  (= 員工待補工作量初估)')
    lines.append(f'  矩陣檔案:      {MATRIX_XL}')
    lines.append('')

    lines.append('─── 待 Jim 處理清單 ──────────────────────────────────────────')
    lines.append(f'  1. ie_import_failures.txt — {len(import_stats["failures"])} 筆失敗 + {import_stats["orphan_no_header"]} 筆孤兒')
    lines.append(f'  2. sheet_unmapped.txt — {len(unmapped)} 個無法自動分類的 sheet 名稱')
    lines.append(f'     → 請在 sheet_name_map 表手動設 type_key')
    lines.append(f'  3. art_sheet_status → missing 的 header 待補 IE 資料')
    lines.append('')
    lines.append('  Note: 純讀取 + 寫入 ie_sheet_data，未改 ob_epph / ob_header。')
    lines.append('=' * W)

    text = '\n'.join(lines)
    with open(REPORT_OUT, 'w', encoding='utf-8') as f:
        f.write(text + '\n')
    print(text)
    print(f'\n[Report written to {REPORT_OUT}]')


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print(f"IE Stage 1 — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"DB: {DB_PATH}")

    part1_schema()
    import_stats = part2_import()
    mapped, unmapped, coverage, total_headers = part3_normalize()
    missing_count, common_types = part4_matrix(coverage, total_headers)
    part5_report(import_stats, mapped, unmapped, coverage, total_headers, missing_count, common_types)


if __name__ == '__main__':
    main()
