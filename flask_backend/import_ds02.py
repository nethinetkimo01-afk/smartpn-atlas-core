#!/usr/bin/env python3
"""
DS-02 FOB Price List Excel importer (CLI).
Usage:
    python import_ds02.py <path.xlsx>
    python import_ds02.py <path.xlsx> --dry-run

Maps columns by name (case-insensitive).
Reports: new / updated / unchanged / errors.
"""
import sys, os, json

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import database as db

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

COL_MAP = {
    'article #':'art','article#':'art','article no':'art','article number':'art',
    'art':'art','art #':'art',
    'model #':'model_no','model#':'model_no','model no':'model_no','model name':'model_name',
    'silhouette number':'silhouette_no','silhouette no':'silhouette_no',
    'sil. no':'silhouette_no','silhouette#':'silhouette_no',
    'factory':'factory','season':'season','category':'category',
    'lc total':'lc_total','lc_total':'lc_total',
    'lc ctb':'lc_ctb','lc_ctb':'lc_ctb','ctb':'lc_ctb',
    'cutting':'lc_cutting','lc cutting':'lc_cutting','lc_cutting':'lc_cutting',
    'stitching':'lc_stitching','lc stitching':'lc_stitching','lc_stitching':'lc_stitching',
    'stockfitting':'lc_stockfitting','lc stockfitting':'lc_stockfitting',
    'stock fitting':'lc_stockfitting','lc_stockfitting':'lc_stockfitting',
    'assembly':'lc_assembly','lc assembly':'lc_assembly','lc_assembly':'lc_assembly',
    'stage':'stage','valid from':'valid_from','valid_from':'valid_from',
}


def read_fob_excel(path):
    if not HAS_OPENPYXL:
        return None, 'openpyxl not installed. Run: pip install openpyxl'

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Find header row (first row with ≥2 mapped column names)
    header_row_idx, headers = None, []
    for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
        cells = [str(c or '').strip().lower() for c in row]
        if sum(1 for c in cells if c in COL_MAP) >= 2:
            header_row_idx = i + 1
            headers = [str(c or '').strip() for c in row]
            break

    if header_row_idx is None:
        return None, 'Cannot find header row with Article # and LC cost columns'

    col_to_field, unmapped = {}, []
    for idx, h in enumerate(headers):
        key = h.lower().strip()
        if key in COL_MAP:
            col_to_field[idx] = COL_MAP[key]
        elif key:
            unmapped.append(h)

    if 'art' not in col_to_field.values():
        return None, f'Article # column not found. Headers: {headers[:20]}'

    records = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(c is not None and str(c).strip() for c in row):
            continue
        rec, raw = {}, {}
        for idx, val in enumerate(row):
            hdr = headers[idx] if idx < len(headers) else f'col_{idx}'
            raw[hdr] = val
            if idx in col_to_field:
                rec[col_to_field[idx]] = val
        if not str(rec.get('art', '') or '').strip():
            continue
        rec['raw_data'] = json.dumps(raw, default=str)
        records.append(rec)

    return records, {'total': len(records), 'unmapped_cols': unmapped}


def import_file(path):
    records, info = read_fob_excel(path)
    if records is None:
        return {'ok': False, 'error': info}
    result = db.import_ds02_records(records)
    result['file_info'] = info
    return result


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: python import_ds02.py <path.xlsx> [--dry-run]')
        sys.exit(1)

    path = sys.argv[1]
    if not os.path.exists(path):
        print(f'File not found: {path}')
        sys.exit(1)

    if '--dry-run' in sys.argv:
        records, info = read_fob_excel(path)
        if records is None:
            print(f'Error: {info}')
            sys.exit(1)
        print(f'DRY RUN — {len(records)} records found')
        print(f'Unmapped columns: {info.get("unmapped_cols", [])}')
        if records:
            print(f'Sample fields: {list(k for k in records[0] if k != "raw_data")}')
            print(f'First ART: {records[0].get("art")}')
        sys.exit(0)

    db.init_db()
    result = import_file(path)

    if result['ok']:
        s  = result.get('stats', {})
        fi = result.get('file_info', {})
        print(f'Import complete: {os.path.basename(path)}')
        print(f'  New:       {s.get("new", 0)}')
        print(f'  Updated:   {s.get("updated", 0)}')
        print(f'  Unchanged: {s.get("unchanged", 0)}')
        print(f'  Errors:    {s.get("errors", 0)}')
        if fi.get('unmapped_cols'):
            print(f'  Unmapped columns: {fi["unmapped_cols"]}')
        for e in s.get('error_details', [])[:5]:
            print(f'  ERR: {e}')
    else:
        print(f'Import failed: {result.get("error")}')
        sys.exit(1)
