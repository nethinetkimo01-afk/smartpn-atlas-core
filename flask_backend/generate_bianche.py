#!/usr/bin/env python3
"""
generate_bianche.py
自動生成廠務組織編制表 (auto_bianche.xlsx)

CSA section: 由 DS-04 進度表產生 (LEAN/鞋型/ART/訂單)
OCS/RB/QC:  直接從現行廠務組織編制表讀取
MP欄位 (裁斷/針車/成型/協理給/合計/編制): 留空

輸出: test_output/auto_bianche.xlsx
"""
import sys, os, re, shutil, collections
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import build_result_table as brt
import database as db
from full_compare_report import load_bianche

BIANCHE_ORIG = r'C:\Users\user\OneDrive\Desktop\Biên chế\Jun\2026年6月份廠務组织編制 20260524.xlsx'
BIANCHE_TMP  = r'C:\Users\user\AppData\Local\Temp\bianche_gen.xlsx'
OUT_PATH     = r'D:\smartpn-atlas-core\flask_backend\test_output\auto_bianche.xlsx'

# Patch brt file paths to temp copies (source files may be locked by Excel)
import tempfile
def _tmp(orig):
    dst = os.path.join(tempfile.gettempdir(), os.path.basename(orig))
    shutil.copy2(orig, dst)
    return dst
brt.SCHEDULE = _tmp(brt.SCHEDULE)
brt.BIANCHE  = _tmp(brt.BIANCHE)
MONTH_SH     = '6.2026'
_ART_RE      = re.compile(r'[A-Z]{2}\d{4,6}')

# ── Styles ────────────────────────────────────────────────────────────────────
THIN      = Side(style='thin')
THIN_BDR  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL  = PatternFill('solid', fgColor='4472C4')
LEAN_FILL = PatternFill('solid', fgColor='FFF2CC')
SEC_FILL  = PatternFill('solid', fgColor='D9D9D9')
BLANK_FILL= PatternFill('solid', fgColor='FFE2CC')   # orange tint = no data
HDR_FONT  = Font(bold=True, color='FFFFFF', size=10)
BOLD      = Font(bold=True, size=10)
NORM      = Font(size=10)
CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT      = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT     = Alignment(horizontal='right', vertical='center')

OCS_SECTIONS = ['組底配套', '自動化', '電腦針車', '印刷', '設備工程']
RB_QC_SECTIONS = ['RB', 'QC']


def _safe_bianche():
    try:
        with open(BIANCHE_ORIG, 'rb'): pass
        return BIANCHE_ORIG
    except (PermissionError, FileNotFoundError):
        if not os.path.exists(BIANCHE_TMP):
            shutil.copy2(BIANCHE_ORIG, BIANCHE_TMP)
        return BIANCHE_TMP


def get_lc(art):
    """Return (cut, stitch, asm, stock) tuple from DS-02, or None if not found."""
    try:
        conn = db.get_conn()
        row = conn.execute(
            'SELECT lc_cutting, lc_stitching, lc_assembly, lc_stockfitting '
            'FROM ds02_fob WHERE art=? LIMIT 1', (art,)
        ).fetchone()
        conn.close()
        if row and any(v is not None for v in row):
            return tuple(row)
        return None
    except Exception:
        return None


def get_ob_mp(art):
    """Return (cutting, stitching, assembly) from ob_epph via ob_articles, or None."""
    try:
        conn = db.get_conn()
        row = conn.execute(
            '''SELECT e.cutting, e.stitching, e.assembly
               FROM ob_epph e
               JOIN ob_header h ON h.id = e.header_id
               JOIN ob_articles a ON a.header_id = h.id
               WHERE a.art = ?
               ORDER BY (h.eolr=120) DESC, h.id DESC LIMIT 1''',
            (art,)
        ).fetchone()
        conn.close()
        if row and any(v for v in row):
            return float(row[0] or 0), float(row[1] or 0), float(row[2] or 0)
        return None
    except Exception:
        return None


def merge_lean_rows(art_rows):
    """Within a LEAN group: same model_name + same LC key → merge, sum qty.
    Different LC or no DS-02 data → separate row.
    Returns list of {'label': str, 'qty': int, 'arts': [str]}.
    """
    # Use ordered dict to preserve first-seen order per group
    groups = {}   # key (model, lc_tuple) -> {'model', 'lc', 'arts': [], 'qty': 0}
    order  = []   # insertion order of keys

    for r in art_rows:
        art   = r['art']
        qty   = r['qty']
        model = brt.get_model(art) or ''
        lc    = get_lc(art)

        if model and lc:
            key = (model, lc)
        else:
            # No match possible — treat as unique singleton
            key = ('', None, art)   # unique key per ART

        if key not in groups:
            groups[key] = {'model': model, 'lc': lc, 'arts': [], 'qty': 0}
            order.append(key)
        groups[key]['arts'].append(art)
        groups[key]['qty'] += qty

    result = []
    for key in order:
        g = groups[key]
        arts_sorted = sorted(g['arts'])
        arts_joined = ' / '.join(arts_sorted)
        if g['model']:
            label = g['model'] + ' ( ' + arts_joined + ' )'
        else:
            label = '( ' + arts_joined + ' )'
        result.append({'label': label, 'qty': g['qty'], 'arts': arts_sorted})
    return result


def _lean_sort_key(lean):
    m = re.match(r'^(\d+)([A-Z])(\d*)$', lean)
    if m:
        return (int(m.group(1)), m.group(2), int(m.group(3)) if m.group(3) else 0)
    return (999, lean, 0)


def load_ds04_by_lean():
    """Each DS-04 sheet is independent — no cross-sheet aggregation.
    Returns {lean: [{art, qty, sheet}]} where each (sheet, lean, art) is a separate row.
    If the same (lean, art) appears in multiple sheets → multiple rows in that LEAN group.
    """
    print('Loading DS-04...')
    sheets = brt.load_schedule()
    lean_rows = collections.defaultdict(list)  # lean -> [{art, qty, sheet}]
    for s in sheets:
        sheet_title = s['sheet']
        for r in s['rows']:
            lean = r['lean'].strip()
            if not lean:
                continue
            lean_rows[lean].append({'art': r['art'], 'qty': r['qty'], 'sheet': sheet_title})

    result = {}
    for lean in sorted(lean_rows, key=_lean_sort_key):
        rows = sorted(lean_rows[lean], key=lambda x: (x['art'], x['sheet']))
        result[lean] = rows

    total = sum(len(v) for v in result.values())
    print(f'  {len(result)} LEAN groups, {total} (sheet,lean,art) rows')
    return result


def load_schedule_detail():
    """Individual MF order records from DS-04 (not aggregated).
    Includes 成型进度 AND 外包鞋面 sections; skips 针车进度 only.
    Returns list of {lean, sheet, art, mf_order, section, qty, date}.
    """
    wb = openpyxl.load_workbook(brt.SCHEDULE, data_only=True)
    records = []
    _SKIP_SECT = frozenset({'针车进度', '针車進度'})

    for ws in wb.worksheets:
        title = ws.title.strip()
        rows_data = list(ws.iter_rows(values_only=True))

        # Pass 1: which LEANs have 成型进度 sub-section
        lean_has_cx = set()
        tmp_lean = ''
        for row in rows_data:
            col1 = str(row[0] or '').strip() if row else ''
            col2 = str(row[1] or '').strip() if len(row) > 1 else ''
            lp = brt._parse_lean_title(col1)
            if lp:
                tmp_lean = lp
            if col2 in brt._SECTION_CHENGXING and tmp_lean:
                lean_has_cx.add(tmp_lean)

        # Pass 2: extract individual orders
        seen_mf = set()
        current_lean = ''
        current_section = ''

        for row in rows_data:
            col1 = str(row[0] or '').strip() if row else ''
            col2 = str(row[1] or '').strip() if len(row) > 1 else ''
            lp = brt._parse_lean_title(col1)
            if lp:
                current_lean = lp
                current_section = ''
                continue
            if col2 in brt._ALL_SECTIONS:
                current_section = col2
                continue
            if current_section in _SKIP_SECT:
                continue
            # For LEANs with 成型进度: skip main body unless in 外包鞋面
            if current_lean in lean_has_cx and current_section not in brt._SECTION_CHENGXING and current_section != '外包鞋面':
                continue

            for cell in row:
                if not brt._is_order_cell(cell):
                    continue
                s = str(cell or '').strip()
                mf_m = brt._MF_ORDER_RE.match(s)
                mf_order = mf_m.group(1) if mf_m else ''
                if mf_order:
                    mf_key = (current_lean, mf_order)
                    if mf_key in seen_mf:
                        continue
                    seen_mf.add(mf_key)

                qty_m = re.search(r'--(.+?)\(', s)
                qty = sum(int(x) for x in re.findall(r'\d+', qty_m.group(1))) if qty_m else 0
                date_m = re.search(r'\(([^)]+)\)', s)
                date = date_m.group(1) if date_m else ''
                arts = [a for a in _ART_RE.findall(s) if not a.startswith('MF')]
                if not arts:
                    continue
                sect = '成型进度' if current_section in brt._SECTION_CHENGXING else (current_section or '成型进度')
                for art in arts:
                    records.append({
                        'lean': current_lean, 'sheet': title,
                        'art': art, 'mf_order': mf_order,
                        'section': sect, 'qty': qty, 'date': date,
                    })

    wb.close()
    return records


def load_ocs_rb_qc(bianche_path):
    """Read OCS fixed units and RB/QC sections from factory table."""
    print('Loading OCS/RB/QC from factory table...')
    wb = openpyxl.load_workbook(bianche_path, data_only=True)
    ws = next((wb[s] for s in wb.sheetnames if s.strip() == MONTH_SH), None)
    if ws is None:
        wb.close()
        return {}, {}

    def _hc(row, *idxs):
        for idx in idxs:
            v = row[idx] if len(row) > idx else None
            if v is not None and str(v).strip() not in ('', '.', '編制'):
                try: return int(float(v))
                except: pass
        return None

    ocs = {s: [] for s in OCS_SECTIONS}
    rbqc = {s: [] for s in RB_QC_SECTIONS}
    cur_ocs = None
    cur_rbqc = None

    for row in ws.iter_rows(values_only=True):
        col1 = str(row[0] or '').strip() if row else ''
        col2 = str(row[1] or '').strip() if len(row) > 1 else ''

        if col1 in OCS_SECTIONS:
            cur_ocs = col1; cur_rbqc = None
            if col2 and col2 not in ('編制', ''):
                ocs[cur_ocs].append({'unit': col2, 'hc': _hc(row, 12, 10, 14)})
        elif col1 in RB_QC_SECTIONS:
            cur_rbqc = col1; cur_ocs = None
            if col2 and col2 not in ('編制', ''):
                last_m = _hc(row, 12)
                this_m = _hc(row, 14)
                rbqc[cur_rbqc].append({'unit': col2, 'last_month': last_m, 'this_month': this_m or last_m})
        elif col1 and col1 not in OCS_SECTIONS and col1 not in RB_QC_SECTIONS:
            cur_ocs = None
            if col1 not in ('設備工程', '副總室', '現場技轉/ KTHT'):
                cur_rbqc = None
        else:
            if cur_ocs and not col1 and col2 and col2 not in ('編制', ''):
                ocs[cur_ocs].append({'unit': col2, 'hc': _hc(row, 12, 10, 14)})
            elif cur_rbqc and not col1 and col2 and col2 not in ('編制', ''):
                last_m = _hc(row, 12)
                this_m = _hc(row, 14)
                rbqc[cur_rbqc].append({'unit': col2, 'last_month': last_m, 'this_month': this_m or last_m})

    wb.close()
    for s in OCS_SECTIONS:
        print(f'  OCS {s}: {len(ocs[s])} units')
    for s in RB_QC_SECTIONS:
        print(f'  {s}: {len(rbqc[s])} units')
    return ocs, rbqc


def wc(ws, r, c, val, font=None, fill=None, align=None, border=None, num_fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if border: cell.border = border
    if num_fmt: cell.number_format = num_fmt
    return cell


MP_FILL   = PatternFill('solid', fgColor='E2EFDA')   # green tint = has IE MP data

# Column layout (1-indexed):
# 1:LEAN  2-5:鞋型+ART  6:訂單  7:裁斷IE MP  8:針車IE MP  9:成型IE MP
# 10:自動化MP  11:訂單編制MP  12:裁斷MP  13:針車MP  14:成型MP  15:協理給  16:合計
_NCOLS = 16


def build_sheet(ws, lean_by_art, ocs, rbqc):
    # Column widths
    widths = {1: 8, 2: 45, 3: 5, 4: 5, 5: 5, 6: 10,
              7: 9, 8: 9, 9: 9, 10: 8, 11: 9,
              12: 8, 13: 8, 14: 8, 15: 8, 16: 8}
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    row = 1

    # ── Title ────────────────────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:P{row}')
    wc(ws, row, 1, f'廠務組織編制表（自動產生 auto_bianche）— {MONTH_SH}',
       font=Font(bold=True, size=13), align=CENTER)
    row += 1

    ws.merge_cells(f'A{row}:P{row}')
    wc(ws, row, 1,
       '★ 橙色欄位 = 本月無資料（MP、協理給等待 Jim 填入）| G/H/I = DS-03 IE MP | 此表由 DS-04 自動產生',
       font=Font(size=9, italic=True, color='CC4400'), align=LEFT)
    row += 2

    # ── Column headers ───────────────────────────────────────────────────────
    hdr_row = ['LEAN', '鞋型 + ART', '', '', '', '訂單',
               '裁斷IE MP', '針車IE MP', '成型IE MP',
               '自動化MP', '訂單編制MP',
               '裁斷MP', '針車MP', '成型MP', '協理給', '合計']
    ws.row_dimensions[row].height = 36
    for ci, h in enumerate(hdr_row, 1):
        wc(ws, row, ci, h, font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=THIN_BDR)
    row += 1
    ws.freeze_panes = f'A{row}'

    # ── CSA section ──────────────────────────────────────────────────────────
    lean_num_groups = {}
    for lean in lean_by_art:
        m = re.match(r'^(\d+)', lean)
        n = int(m.group(1)) if m else 99
        lean_num_groups.setdefault(n, []).append(lean)

    for dept_num in sorted(lean_num_groups):
        ws.merge_cells(f'A{row}:P{row}')
        wc(ws, row, 1, f'LEAN {dept_num} （加{dept_num}部）',
           font=Font(bold=True, size=10, color='1F4E79'), fill=SEC_FILL, align=LEFT)
        row += 1

        for lean in sorted(lean_num_groups[dept_num], key=_lean_sort_key):
            arts = lean_by_art[lean]

            # LEAN group header row
            ws.row_dimensions[row].height = 18
            wc(ws, row, 1, lean, font=BOLD, fill=LEAN_FILL, align=CENTER, border=THIN_BDR)
            ws.merge_cells(f'B{row}:E{row}')
            wc(ws, row, 2, '鞋型 + ART', font=BOLD, fill=LEAN_FILL, align=CENTER, border=THIN_BDR)
            wc(ws, row, 6, '訂單', font=BOLD, fill=LEAN_FILL, align=CENTER, border=THIN_BDR)
            for ci in range(7, _NCOLS + 1):
                wc(ws, row, ci, None, fill=BLANK_FILL, border=THIN_BDR)
            row += 1

            # Merge rows: same model_name + same LC → one row, sum qty
            merged = merge_lean_rows(arts)
            for mg in merged:
                ws.row_dimensions[row].height = 15
                wc(ws, row, 1, None, border=THIN_BDR)
                ws.merge_cells(f'B{row}:E{row}')
                wc(ws, row, 2, mg['label'], font=NORM, align=LEFT, border=THIN_BDR)
                wc(ws, row, 6, mg['qty'], font=NORM, align=RIGHT, border=THIN_BDR, num_fmt='#,##0')

                # Fill G/H/I from ob_epph (try each ART in the group)
                ob_mp = None
                for a in mg['arts']:
                    ob_mp = get_ob_mp(a)
                    if ob_mp:
                        break

                if ob_mp:
                    cut_v, stch_v, asm_v = ob_mp
                    wc(ws, row, 7,  cut_v,  font=NORM, fill=MP_FILL, align=CENTER, border=THIN_BDR, num_fmt='0.0')
                    wc(ws, row, 8,  stch_v, font=NORM, fill=MP_FILL, align=CENTER, border=THIN_BDR, num_fmt='0.0')
                    wc(ws, row, 9,  asm_v,  font=NORM, fill=MP_FILL, align=CENTER, border=THIN_BDR, num_fmt='0.0')
                else:
                    for ci in [7, 8, 9]:
                        wc(ws, row, ci, None, fill=BLANK_FILL, border=THIN_BDR)

                for ci in range(10, _NCOLS + 1):
                    wc(ws, row, ci, None, fill=BLANK_FILL, border=THIN_BDR)
                row += 1

    row += 1  # spacer

    # ── OCS sections ─────────────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:P{row}')
    wc(ws, row, 1, 'OCS 固定單位（直接從廠務組織編制表讀入）',
       font=Font(bold=True, size=11, color='1F4E79'), fill=SEC_FILL, align=LEFT)
    row += 1

    for sec in OCS_SECTIONS:
        units = ocs.get(sec, [])
        ws.merge_cells(f'A{row}:P{row}')
        wc(ws, row, 1, sec, font=BOLD, fill=LEAN_FILL, align=LEFT, border=THIN_BDR)
        row += 1
        for u in units:
            wc(ws, row, 1, None, border=THIN_BDR)
            ws.merge_cells(f'B{row}:O{row}')
            wc(ws, row, 2,  u['unit'], font=NORM, align=LEFT, border=THIN_BDR)
            wc(ws, row, 16, u['hc'],   font=NORM, align=RIGHT, border=THIN_BDR, num_fmt='#,##0')
            row += 1
        row += 1

    # ── RB / QC sections ─────────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:P{row}')
    wc(ws, row, 1, 'RB / QC（直接從廠務組織編制表讀入）',
       font=Font(bold=True, size=11, color='1F4E79'), fill=SEC_FILL, align=LEFT)
    row += 1

    for sec in RB_QC_SECTIONS:
        units = rbqc.get(sec, [])
        ws.merge_cells(f'A{row}:P{row}')
        wc(ws, row, 1, sec, font=BOLD, fill=LEAN_FILL, align=LEFT, border=THIN_BDR)
        row += 1

        # Sub-headers
        ws.row_dimensions[row].height = 18
        ws.merge_cells(f'A{row}:N{row}')
        wc(ws, row, 1,  '單位',    font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=THIN_BDR)
        wc(ws, row, 15, '上月人數', font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=THIN_BDR)
        wc(ws, row, 16, '本月人數', font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=THIN_BDR)
        row += 1

        for u in units:
            ws.merge_cells(f'A{row}:N{row}')
            wc(ws, row, 1,  u['unit'],       font=NORM, align=LEFT,  border=THIN_BDR)
            wc(ws, row, 15, u['last_month'], font=NORM, align=RIGHT, border=THIN_BDR, num_fmt='#,##0')
            wc(ws, row, 16, u['this_month'], font=NORM, align=RIGHT, border=THIN_BDR, num_fmt='#,##0')
            row += 1
        row += 1


def build_detail_sheet(ws, detail_records, art_to_ref):
    """製令明細 sheet: one row per MF order + subtotal vs factory qty."""
    from itertools import groupby
    RED_D  = PatternFill('solid', fgColor='FFCCCC')
    YELL_D = PatternFill('solid', fgColor='FFF2CC')

    col_w = {1: 8, 2: 26, 3: 12, 4: 28, 5: 10, 6: 11, 7: 8, 8: 11, 9: 11, 10: 11}
    for ci, w in col_w.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.merge_cells('A1:J1')
    wc(ws, 1, 1, f'製令明細 — DS-04 vs 廠務組織編制表 {MONTH_SH}',
       font=Font(bold=True, size=13), align=CENTER)

    hdrs = ['LEAN', '製令號碼', 'ART', '鞋型', '段落', 'DS-04訂單量', '交期', '廠務訂單', '差異', '外包鞋面']
    for ci, h in enumerate(hdrs, 1):
        wc(ws, 2, ci, h, font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=THIN_BDR)
    ws.freeze_panes = 'A3'

    _WBMX = frozenset({'外包鞋面'})

    recs = sorted(detail_records,
                  key=lambda r: (_lean_sort_key(r['lean']), r['art'], r['mf_order']))
    row = 3

    for (lean, art), grp_iter in groupby(recs, key=lambda r: (r['lean'], r['art'])):
        buf = list(grp_iter)
        model      = brt.get_model(art) or ''
        ref        = art_to_ref.get(art, {})
        fac_qty    = ref.get('order')
        ds04_total = sum(o['qty'] for o in buf)
        wbmx_total = sum(o['qty'] for o in buf if o['section'] in _WBMX)
        diff = (ds04_total - int(fac_qty)) if fac_qty is not None else None

        # Individual order rows
        for i, o in enumerate(buf):
            is_first  = (i == 0)
            is_wbmx   = o['section'] in _WBMX
            wc(ws, row, 1,  lean if is_first else '', font=NORM, align=CENTER, border=THIN_BDR)
            wc(ws, row, 2,  o['mf_order'], font=Font(size=9), align=LEFT, border=THIN_BDR)
            wc(ws, row, 3,  art if is_first else '', font=NORM, align=LEFT, border=THIN_BDR)
            wc(ws, row, 4,  model if is_first else '', font=Font(size=9), align=LEFT, border=THIN_BDR)
            wc(ws, row, 5,  o['section'], font=Font(size=9), align=CENTER, border=THIN_BDR)
            wc(ws, row, 6,  o['qty'], font=NORM, align=RIGHT, border=THIN_BDR, num_fmt='#,##0')
            wc(ws, row, 7,  o['date'], font=Font(size=9), align=CENTER, border=THIN_BDR)
            wc(ws, row, 8,  None, border=THIN_BDR)
            wc(ws, row, 9,  None, border=THIN_BDR)
            wc(ws, row, 10, o['qty'] if is_wbmx else None,
               font=Font(size=9, color='7030A0'), align=RIGHT, border=THIN_BDR,
               num_fmt='#,##0')
            row += 1

        # Subtotal / summary row
        label = f'小計 ({len(buf)} 製令)' if len(buf) > 1 else '合計'
        wc(ws, row, 1, lean,  font=BOLD, fill=LEAN_FILL, align=CENTER, border=THIN_BDR)
        wc(ws, row, 2, label, font=Font(bold=True, size=9, italic=True), fill=LEAN_FILL, align=LEFT, border=THIN_BDR)
        wc(ws, row, 3, art,   font=BOLD, fill=LEAN_FILL, align=LEFT,   border=THIN_BDR)
        for ci in [4, 5, 7]:
            wc(ws, row, ci, None, fill=LEAN_FILL, border=THIN_BDR)
        wc(ws, row, 6, ds04_total, font=BOLD, fill=LEAN_FILL, align=RIGHT, border=THIN_BDR, num_fmt='#,##0')

        if fac_qty is not None:
            wc(ws, row, 8, int(fac_qty), font=BOLD, fill=LEAN_FILL, align=RIGHT, border=THIN_BDR, num_fmt='#,##0')
            threshold = max(int(fac_qty) * 0.2, 50)
            if diff is not None and abs(diff) > threshold:
                d_fill = RED_D
                d_font = Font(bold=True, size=10, color='CC0000')
            elif diff:
                d_fill = YELL_D
                d_font = BOLD
            else:
                d_fill = LEAN_FILL
                d_font = BOLD
            wc(ws, row, 9, diff, font=d_font, fill=d_fill, align=RIGHT, border=THIN_BDR, num_fmt='+#,##0;-#,##0;0')
        else:
            wc(ws, row, 8, '—', font=NORM, fill=LEAN_FILL, align=CENTER, border=THIN_BDR)
            wc(ws, row, 9, '—', font=NORM, fill=LEAN_FILL, align=CENTER, border=THIN_BDR)

        wc(ws, row, 10, wbmx_total if wbmx_total else None,
           font=Font(bold=True, color='7030A0'), fill=LEAN_FILL, align=RIGHT,
           border=THIN_BDR, num_fmt='#,##0')
        row += 1


def run():
    bianche_path = _safe_bianche()
    lean_by_art  = load_ds04_by_lean()
    ocs, rbqc    = load_ocs_rb_qc(bianche_path)

    print('Loading factory reference for 製令明細...')
    try:
        art_to_ref, _ = load_bianche(bianche_path)
    except Exception as e:
        print(f'  Warning: {e}')
        art_to_ref = {}

    print('Loading individual MF order records...')
    detail_records = load_schedule_detail()
    print(f'  {len(detail_records)} individual order records')

    print('Building auto_bianche.xlsx...')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = MONTH_SH
    build_sheet(ws, lean_by_art, ocs, rbqc)

    ws_detail = wb.create_sheet('製令明細')
    build_detail_sheet(ws_detail, detail_records, art_to_ref)

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f'Saved: {OUT_PATH}')

    # Verify HP4218 in 8B
    hp_recs = [r for r in detail_records if r['art'] == 'HP4218']
    print()
    print('=== 驗證：HP4218 所有製令 ===')
    if hp_recs:
        for r in sorted(hp_recs, key=lambda x: (x['lean'], x['mf_order'])):
            print(f"  LEAN={r['lean']:6s}  {r['mf_order']:26s}  {r['section']:6s}  {r['qty']:>6,}  交期:{r['date']}")
        recs_8b   = [r for r in hp_recs if r['lean'] == '8B']
        total_8b  = sum(r['qty'] for r in recs_8b)
        fac       = art_to_ref.get('HP4218', {}).get('order')
        print(f"  ─────────────────────────────────────────────────")
        print(f"  8B DS-04合計: {total_8b:,}  廠務: {int(fac) if fac else '—'}  差異: {total_8b - (int(fac) if fac else 0):+,}")
    else:
        print('  HP4218: 無記錄')

    print()
    print('=== 與廠務組織編制表 差異摘要 ===')
    from column_compare_report import run as ccr_run
    ccr_run()


if __name__ == '__main__':
    run()
