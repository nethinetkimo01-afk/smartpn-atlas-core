#!/usr/bin/env python3
"""
全量 IE 導入 — 照 27_WORKING_RULES.md Sheet名稱對照表
直接掃描 IE_XLSX_DIR 所有 xlsx，依 sheet 名稱自動分派到 ie_process 表。
不確定的 sheet 名稱記到 test_output/unknown_sheets.txt，不跳過也不猜。
導完讀 SUM.C2B 產出對比表 test_output/ie_import_comparison.xlsx
"""
import sys, os, re, sqlite3
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

IE_XLSX_DIR = r'D:\smartpn-atlas-core\data\source_files\IE_xlsx'
DB_PATH     = os.path.join(os.path.dirname(__file__), 'data', 'atlas.db')
OUT_DIR     = os.path.join(os.path.dirname(__file__), 'test_output')
os.makedirs(OUT_DIR, exist_ok=True)

# ── Sheet 名稱對照表（照 27_WORKING_RULES.md 定案）──────────────────────────
SKIP_PAT = re.compile(
    r'^(sum[.\s]|sum$|sheet\d+$|chi ti|同材共裁|ac$|stt$|bang ke|nhân|nhan luc|'
    r'summary|bộ phận|bo phan|tong hop|khai bao|%cs|%ac|tc$)',
    re.I
)

def classify_sheet(name: str):
    """Return (segment, zone) or None if unknown, 'skip' if known-skip."""
    n = name.strip()
    nl = n.lower()
    if SKIP_PAT.match(nl): return 'skip'
    if re.match(r'^ac[_\s\-\(]', nl, re.I): return 'skip'
    if re.match(r'^%', nl): return 'skip'
    # Cutting
    if re.match(r'^cutting[_\s\-\(]?emma', nl): return ('cutting','EMMA')
    if re.match(r'^cutting[_\s\-\(]?tc', nl):   return ('cutting','裁斷機')
    if re.match(r'^cutting', nl):                return ('cutting','裁斷機')
    if nl == 'atom':                             return ('cutting','ATOM')
    if re.match(r'^atom\s*[-–]\s*laser', nl):    return ('cutting','_auto')  # ATOM-LASER mixed
    if re.match(r'^自动[裁化]', nl) or re.match(r'^自动化', nl): return ('cutting','_auto')
    if '自动化' in nl and re.search(r'(màu|color)', nl, re.I): return ('cutting','_auto')
    if nl in ('tự động cắt','tat dong cat'): return ('cutting','_auto')
    if re.match(r'^laser', nl):                  return ('cutting','Laser')
    if 'yinghui' in nl or 'yin hui' in nl:       return ('cutting','YINGHUI')
    if re.match(r'^(移印|转印|chuyển in)', nl):   return ('cutting','移印')
    # Stitching
    if re.match(r'^sub[.\s_-]?stic', nl.strip()): return ('stitching','支流')  # Sub-Stiching/Sub.Stitching variants
    if re.match(r'^stitching.*sub', nl):   return ('stitching','支流')
    if re.match(r'^stitching', nl):        return ('stitching','主流')
    if re.match(r'^(电脑针车|cs$|电脑|điện toán)', nl, re.I): return ('stitching','電腦針車')
    if '折边' in n or '折邊' in n:               return ('stitching','折边')
    if '鞋舌' in n:                              return ('stitching','鞋舌組')
    # Assembly
    if re.match(r'^assembly[\s._]?1', nl) or re.match(r'^assembly 1', nl): return ('assembly','成型主區')
    if re.match(r'^assembly[\s._]?2', nl) or re.match(r'^assembly 2', nl): return ('assembly','成型主區')
    if '面照射' in n or re.match(r'^assembly.*uv', nl):    return ('assembly','成型UV')
    if '照射' in n and ('成型' in n or 'uv' in nl):        return ('assembly','成型UV')
    if re.match(r'^assembly', nl):               return ('assembly','成型主區')
    # STF
    if '打粗' in n:                              return ('stf','打粗')
    if '水洗' in n:                              return ('stf','水洗')
    if '贴大底' in n or '貼大底' in n or '贴底' in n or '貼底' in n: return ('stf','貼底')
    if '贴中底' in n or '貼中底' in n:           return ('stf','貼底')
    if '组底面照射' in n or '組底面照射' in n or re.match(r'^sum.?stock', nl): return ('stf','照射')
    if '照射' in n:                              return ('stf','照射')  # standalone 照射 → STF
    if '橡膠' in n or '橡胶' in n:              return ('stf','貼底')
    return None

def ws_to_grid(ws):
    grid = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                grid.setdefault(cell.row, {})[cell.column] = cell.value
    return grid

def tf(v):
    if v is None: return None
    try: return float(str(v).replace(',','').strip())
    except: return None

def ti(v):
    if v is None: return None
    try:
        s = str(v).strip().split('.')[0]
        n = int(s)
        return n if n > 0 else None
    except: return None

def split_vn_zh(name):
    if not name: return '', ''
    s = str(name).strip()
    for i, c in enumerate(s):
        if '一' <= c <= '鿿' or '㐀' <= c <= '䶿':
            return s[:i].strip(), s[i:].strip()
    return s, ''

def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA journal_mode=WAL')
    return conn

def resolve_header_id(conn, path):
    fn = os.path.basename(path)
    arts = re.findall(r'[A-Z]{2}\d{4,6}', fn)
    for art in arts:
        row = conn.execute(
            '''SELECT h.id, h.eolr, h.model_name FROM ob_articles a
               JOIN ob_header h ON h.id=a.header_id
               WHERE a.art=? ORDER BY h.id DESC LIMIT 1''', (art,)
        ).fetchone()
        if row:
            return row['id'], row['eolr'] or 120, row['model_name'] or '', art
    return None, 120, '', (arts[0] if arts else '')

def insert_process(conn, *, header_id, art, segment, zone, seq, name_raw,
                   standard_time=None, actual_ops=None, mat_cat=None,
                   layers=None, qty=None, cut_h=None,
                   normal_time=None, allowance_pct=None,
                   mk_std=None, ski_std=None, att_std=None, edg_std=None, ht_std=None,
                   source_sheet=''):
    viet, zh = split_vn_zh(name_raw)
    conn.execute('''
        INSERT OR IGNORE INTO ie_process
        (header_id, art, segment, zone, seq, process_name, process_name_zh,
         mat_cat, layers_per_cut, qty_per_pair, cut_per_hour,
         normal_time, allowance_pct, standard_time, actual_operators,
         post_marking_std, post_skiving_std, post_attach_std, post_edge_std, post_heat_std,
         is_locked, source_sheet, stage)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    ''', (header_id, art, segment, zone, seq,
          (viet or name_raw)[:200], zh[:200] if zh else '',
          (mat_cat or '')[:50],
          layers, qty, cut_h,
          normal_time, allowance_pct, standard_time, actual_ops,
          mk_std, ski_std, att_std, edg_std, ht_std,
          1, source_sheet[:100], 1))

def import_cutting_tc(conn, header_id, art, ws, zone='裁斷機', source_sheet=''):
    grid = ws_to_grid(ws)
    count = 0
    last_mat = ''
    for rn in sorted(grid.keys()):
        if rn < 4: continue
        d = grid[rn]
        seq = ti(d.get(3))
        if seq is None: continue
        mat_raw = str(d.get(2) or '').strip()
        if mat_raw and len(mat_raw) > 1 and not any(
            kw in mat_raw for kw in ['Tổng','Mục','IE prep','Thuyết','分類']
        ):
            last_mat = mat_raw
        name_raw = str(d.get(4) or '').strip()
        if not name_raw: continue
        insert_process(conn, header_id=header_id, art=art, segment='cutting', zone=zone,
                       seq=seq, name_raw=name_raw, mat_cat=last_mat,
                       layers=tf(d.get(5)), qty=tf(d.get(6)), cut_h=tf(d.get(7)),
                       standard_time=tf(d.get(8)), actual_ops=tf(d.get(10)),
                       mk_std=tf(d.get(11)), ski_std=tf(d.get(13)),
                       att_std=tf(d.get(15)), edg_std=tf(d.get(17)), ht_std=tf(d.get(19)),
                       source_sheet=source_sheet)
        count += 1
    return count

def import_auto_cut(conn, header_id, art, ws, source_sheet=''):
    grid = ws_to_grid(ws)
    counts = {}
    for rn in sorted(grid.keys()):
        if rn < 4: continue
        d = grid[rn]
        seq = ti(d.get(2))
        if seq is None: continue
        name_raw = str(d.get(3) or '').strip()
        if not name_raw: continue
        mach = str(d.get(12) or '').strip().upper()
        if 'ATOM' in mach:            zone = 'ATOM'
        elif 'LASER' in mach:         zone = 'Laser'
        elif 'YINGHUI' in mach:       zone = 'YINGHUI'
        elif '移印' in mach:          zone = '移印'
        elif '轉印' in mach or '转印' in mach: zone = '轉印'
        else:                         zone = mach or '裁斷機'
        insert_process(conn, header_id=header_id, art=art, segment='cutting', zone=zone,
                       seq=seq, name_raw=name_raw,
                       layers=tf(d.get(4)), normal_time=tf(d.get(6)),
                       allowance_pct=tf(d.get(7)), standard_time=tf(d.get(8)),
                       actual_ops=tf(d.get(11)), source_sheet=source_sheet)
        counts[zone] = counts.get(zone, 0) + 1
    return counts

def import_std_zone(conn, header_id, art, ws, segment, zone, source_sheet=''):
    grid = ws_to_grid(ws)
    count = 0
    for rn in sorted(grid.keys()):
        if rn < 4: continue
        d = grid[rn]
        seq = ti(d.get(2))
        if seq is None: continue
        name_raw = str(d.get(3) or '').strip()
        if not name_raw: continue
        insert_process(conn, header_id=header_id, art=art, segment=segment, zone=zone,
                       seq=seq, name_raw=name_raw,
                       normal_time=tf(d.get(4)), allowance_pct=tf(d.get(5)),
                       standard_time=tf(d.get(6)), actual_ops=tf(d.get(8)),
                       source_sheet=source_sheet)
        count += 1
    return count

_SEG_KW = {
    'cutting':   ['裁斷','cắt','cutting'],
    'stitching': ['針車','may','stitching','针车'],
    'assembly':  ['成型','assembly'],
    'stf':       ['貼底','打底','stf','stock','貼大底'],
}

def read_sum_c2b(wb):
    ws = None
    for candidate in ('SUM.C2B','Sum.C2B','sum.c2b','SUM C2B','SUMC2B'):
        if candidate in wb.sheetnames:
            ws = wb[candidate]
            break
    if ws is None: return {s: None for s in _SEG_KW}
    grid = ws_to_grid(ws)
    totals = {s: None for s in _SEG_KW}
    for rn in sorted(grid.keys()):
        d = grid[rn]
        for col, val in d.items():
            if val is None: continue
            vs = str(val).strip().lower()
            for seg, kws in _SEG_KW.items():
                if totals[seg] is not None: continue
                if any(kw.lower() in vs for kw in kws):
                    for c2, v2 in d.items():
                        fv = tf(v2)
                        if fv and fv > 0 and c2 != col:
                            totals[seg] = fv
                            break
    return totals

def main():
    import openpyxl

    conn = get_conn()
    unknown = []
    results = []

    xlsx_files = sorted([
        f for f in os.listdir(IE_XLSX_DIR)
        if f.lower().endswith('.xlsx') and not f.startswith('~$')
    ])
    print(f"Found {len(xlsx_files)} xlsx files\n")

    for fname in xlsx_files:
        fpath = os.path.join(IE_XLSX_DIR, fname)
        header_id, eolr, model_name, art = resolve_header_id(conn, fpath)
        if header_id is None:
            print(f"SKIP (no header): {fname[:70]}")
            continue

        print(f"[{header_id}] {art}  {fname[:55]}")

        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f"  ERROR opening: {e}")
            continue

        orig = read_sum_c2b(wb)

        for seg in ('cutting','stitching','assembly','stf'):
            conn.execute("DELETE FROM ie_process WHERE header_id=? AND segment=?", (header_id, seg))

        file_counts = {}
        for sname in wb.sheetnames:
            cls = classify_sheet(sname)
            if cls == 'skip': continue
            if cls is None:
                unknown.append({'file': fname, 'sheet': sname, 'header_id': header_id, 'art': art})
                print(f"  ? UNKNOWN: {sname!r}")
                continue
            seg, zone = cls
            ws = wb[sname]
            try:
                if seg == 'cutting':
                    if zone == '_auto':
                        c = import_auto_cut(conn, header_id, art, ws, sname)
                        for z, n in c.items():
                            file_counts[f'c/{z}'] = file_counts.get(f'c/{z}', 0) + n
                    else:
                        n = import_cutting_tc(conn, header_id, art, ws, zone, sname)
                        file_counts[f'c/{zone}'] = file_counts.get(f'c/{zone}', 0) + n
                elif seg == 'stitching':
                    n = import_std_zone(conn, header_id, art, ws, 'stitching', zone, sname)
                    file_counts[f's/{zone}'] = file_counts.get(f's/{zone}', 0) + n
                elif seg == 'assembly':
                    n = import_std_zone(conn, header_id, art, ws, 'assembly', zone, sname)
                    file_counts[f'a/{zone}'] = file_counts.get(f'a/{zone}', 0) + n
                elif seg == 'stf':
                    n = import_std_zone(conn, header_id, art, ws, 'stf', zone, sname)
                    file_counts[f'f/{zone}'] = file_counts.get(f'f/{zone}', 0) + n
            except Exception as e:
                print(f"  ERROR {sname!r}: {e}")
        wb.close()
        conn.commit()

        def seg_total(segment):
            r = conn.execute(
                "SELECT SUM(standard_time) FROM ie_process WHERE header_id=? AND segment=?",
                (header_id, segment)
            ).fetchone()
            return r[0]

        imported = {s: seg_total(s) for s in _SEG_KW}
        results.append({'model': model_name, 'art': art, 'orig': orig, 'imported': imported})
        print(f"  rows: {file_counts}")

    unk_path = os.path.join(OUT_DIR, 'unknown_sheets.txt')
    with open(unk_path, 'w', encoding='utf-8') as f:
        f.write(f"Unknown sheets: {len(unknown)} total\n{'='*60}\n")
        for u in unknown:
            f.write(f"[{u['header_id']}] {u['art']}  {u['file'][:55]}  sheet={u['sheet']!r}\n")
    print(f"\nUnknown sheets log: {unk_path}  ({len(unknown)} entries)")

    _write_comparison(results)
    conn.close()
    print("Done.")

def _write_comparison(results):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl not available, skipping")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '導入對比'

    headers = ['鞋型','ART',
               'SUM裁斷原始','導入後裁斷','裁斷差異',
               'SUM針車原始','導入後針車','針車差異',
               'SUM成型原始','導入後成型','成型差異',
               'SUM貼底原始','導入後貼底','貼底差異']
    hf = PatternFill('solid', fgColor='1A2744')
    ht = Font(bold=True, color='FFFFFF')
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.fill = hf; c.font = ht
        c.alignment = Alignment(horizontal='center')

    red = PatternFill('solid', fgColor='FFD0D0')
    grn = PatternFill('solid', fgColor='D0FFD0')
    segs = list(_SEG_KW.keys())
    for ri, row in enumerate(results, 2):
        ws.cell(ri, 1, row['model'])
        ws.cell(ri, 2, row['art'])
        for si, seg in enumerate(segs):
            orig = row['orig'].get(seg)
            imp  = row['imported'].get(seg)
            diff = round(imp - orig, 4) if (orig is not None and imp is not None) else None
            ws.cell(ri, 3+si*3, round(orig,4) if orig else None)
            ws.cell(ri, 4+si*3, round(imp,4)  if imp  else None)
            dc = ws.cell(ri, 5+si*3, diff)
            if diff is not None:
                dc.fill = grn if abs(diff) < 0.01 else red

    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(w+3, 25)

    out = os.path.join(OUT_DIR, 'ie_import_comparison.xlsx')
    wb.save(out)
    print(f"Comparison: {out}")

if __name__ == '__main__':
    main()
