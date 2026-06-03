#!/usr/bin/env python3
"""
DS-01 Season Plan Excel importer (CLI).
Usage:
    python import_ds01.py <path.xlsx>
    python import_ds01.py <path.xlsx> --dry-run

PK: Article ID + Product Type DESC + Calendar Month
All columns stored as raw_data JSON; PK + Quantity extracted.
"""
import sys, os, json

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import database as db

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

COL_MAP = {
    'article id':'article_id','article #':'article_id','article no':'article_id',
    'article number':'article_id','article':'article_id',
    'art':'article_id','art #':'article_id','art#':'article_id',
    'product type desc':'product_type_desc','product type description':'product_type_desc',
    'product type':'product_type_desc','type desc':'product_type_desc',
    'type description':'product_type_desc',
    'calendar month':'calendar_month','month':'calendar_month',
    'cal. month':'calendar_month','cal month':'calendar_month',
    'total':'quantity','quantity':'quantity','qty':'quantity',
    'total qty':'quantity','total quantity':'quantity',
}


def read_sp_excel(path):
    if not HAS_OPENPYXL:
        return None, 'openpyxl not installed. Run: pip install openpyxl'

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_row_idx, headers = None, []
    for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
        cells = [str(c or '').strip().lower() for c in row]
        if sum(1 for c in cells if c in COL_MAP) >= 2:
            header_row_idx = i + 1
            headers = [str(c or '').strip() for c in row]
            break

    if header_row_idx is None:
        return None, 'Cannot find header row with Article ID, Product Type, and Calendar Month columns'

    col_to_field, unmapped = {}, []
    for idx, h in enumerate(headers):
        key = h.lower().strip()
        if key in COL_MAP:
            col_to_field[idx] = COL_MAP[key]
        elif key:
            unmapped.append(h)

    pk_fields = ['article_id', 'product_type_desc', 'calendar_month']
    missing = [f for f in pk_fields if f not in col_to_field.values()]
    if missing:
        return None, f'Missing PK columns: {missing}. Headers found: {headers[:20]}'

    records = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(c is not None and str(c).strip() for c in row):
            continue
        rec, raw = {}, {}
        for idx, val in enumerate(row):
            hdr = headers[idx] if idx < len(headers) else f'col_{idx}'
            raw[hdr] = val
            if idx in col_to_field:
                rec[col_to_field[idx]] = val
        art    = str(rec.get('article_id', '') or '').strip()
        ptdesc = str(rec.get('product_type_desc', '') or '').strip()
        month  = str(rec.get('calendar_month', '') or '').strip()
        if not (art and ptdesc and month):
            continue
        rec['raw_data'] = json.dumps(raw, default=str)
        records.append(rec)

    return records, {'total': len(records), 'unmapped_cols': unmapped}


def import_file(path):
    records, info = read_sp_excel(path)
    if records is None:
        return {'ok': False, 'error': info}
    result = db.import_ds01_records(records)
    result['file_info'] = info
    return result


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: python import_ds01.py <path.xlsx> [--dry-run]')
        sys.exit(1)

    path = sys.argv[1]
    if not os.path.exists(path):
        print(f'File not found: {path}')
        sys.exit(1)

    if '--dry-run' in sys.argv:
        records, info = read_sp_excel(path)
        if records is None:
            print(f'Error: {info}')
            sys.exit(1)
        print(f'DRY RUN — {len(records)} records found')
        print(f'Unmapped columns: {info.get("unmapped_cols", [])}')
        if records:
            sample = records[0]
            print(f'Sample: ART={sample.get("article_id")} | Type={sample.get("product_type_desc")} | Month={sample.get("calendar_month")} | Qty={sample.get("quantity")}')
        sys.exit(0)

    db.init_db()
    result = import_file(path)

    if result['ok']:
        s  = result.get('stats', {})
        fi = result.get('file_info', {})
        print(f'Import complete: {os.path.basename(path)}')
        print(f'  New:       {s.get("new", 0)}')
        print(f'  Updated:   {s.get("updated", 0)}')
        print(f'  Unchanged: {s.get("unchanged", 0)}')
        print(f'  Errors:    {s.get("errors", 0)}')
        if fi.get('unmapped_cols'):
            print(f'  Unmapped columns: {fi["unmapped_cols"]}')
        for e in s.get('error_details', [])[:5]:
            print(f'  ERR: {e}')
    else:
        print(f'Import failed: {result.get("error")}')
        sys.exit(1)
