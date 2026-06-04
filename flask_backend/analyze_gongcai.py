#!/usr/bin/env python3
"""
同材共裁明細自動生成腳本 (Tongcai Gongcai Detail Generator)

Generates 同材共裁 (same-material co-cutting) detail report for a DS-04 group.

Workflow:
  1. Parse DS-04 schedule → ART + order quantity for the requested group.
  2. For each ART, find the matching IE Excel file (by ART code in filename).
  3. Parse the IE Cutting sheet:
       - Find the column whose header contains "刀" (標准刀数).
       - Rows where that column value is NOT a number → machine-cut parts
         (ATOM / GCN / LASER / DK / YINGHIU …).
       - Extract: part name (col D) + layers (col E) + pieces (col F) + machine type.
  4. Build output rows:
       STT | LEAN | 鞋型 | ART | 訂單量 | 部件名稱 | 機器類型 | 層數 | 片數 | 勾選

Usage:
    python analyze_gongcai.py <schedule.xlsx> --group 加一A組 [--eolr 120]
    python analyze_gongcai.py <schedule.xlsx> --group 加一A組 --dry-run
"""
import sys, os, re, glob
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import database as db

# ── Config ────────────────────────────────────────────────────────────────────

IE_FOLDER_DEFAULT = r'C:\Users\user\OneDrive\Desktop\IE'

_ART_RE   = re.compile(r'[A-Z]{2}\d{4,6}')
_QTY_RE   = re.compile(r'--(\d+)\s*\(')

# Known machine-type prefixes in the G column
_MACHINE_RE = re.compile(
    r'^(ATOM|GCN|LASER|DK|YINGHUI|YINGHIU|EMMA|AUTO|PAD|ELITRON|AUTO-?CLICK)',
    re.IGNORECASE
)


# ── IE file index ─────────────────────────────────────────────────────────────

def build_ie_index(ie_folder):
    """
    Scan ie_folder for *.xlsx files.
    Returns {ART_UPPER: [(filepath, eolr_or_None), ...]}
    Each file can carry multiple ARTs extracted from its filename.
    """
    index = {}
    for path in glob.glob(os.path.join(ie_folder, '*.xlsx')):
        fn = os.path.basename(path)
        if fn.startswith('~$'):
            continue
        arts = _ART_RE.findall(fn)
        eolr = None
        if '120' in fn and '双' in fn:
            eolr = 120
        elif '60' in fn and '双' in fn:
            eolr = 60
        for art in arts:
            index.setdefault(art.upper(), []).append((path, eolr))
    return index


def find_ie_file(art, eolr, ie_index):
    """Return best IE filepath for art+eolr, or None."""
    candidates = ie_index.get(art.upper(), [])
    if not candidates:
        return None
    for path, e in candidates:
        if e == eolr:
            return path
    return candidates[0][0]


# ── DS-04 schedule parsing ────────────────────────────────────────────────────

def _is_order_cell(val):
    s = str(val or '').strip()
    return bool(_ART_RE.search(s) and '--' in s)


def _parse_order_cell(val):
    """Return list of (art, qty) from one order-number cell."""
    s = str(val or '').strip()
    arts = _ART_RE.findall(s)
    m = _QTY_RE.search(s)
    qty = int(m.group(1)) if m else 0
    if not arts:
        return []
    per = qty // len(arts)
    rem = qty % len(arts)
    return [(a, per + (rem if i == 0 else 0)) for i, a in enumerate(arts)]


def extract_group_orders(wb, group_hint):
    """
    Find the sheet matching group_hint and extract {art: total_qty}.
    Returns (orders_dict, sheet_title_used).
    """
    gh = group_hint.strip().lower()
    ws = None
    for sheet in wb.worksheets:
        if gh in sheet.title.strip().lower():
            ws = sheet
            break
    if ws is None:
        ws = wb.worksheets[0]

    orders = {}
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if not _is_order_cell(cell):
                continue
            for art, qty in _parse_order_cell(cell):
                orders[art] = orders.get(art, 0) + qty

    return orders, getattr(ws, 'title', ws.title if ws else '')


# ── Cutting sheet parser ──────────────────────────────────────────────────────

def _find_knife_col(ws):
    """
    Scan first 20 rows for the column whose header contains '刀'.
    Returns 0-based column index, or None.
    """
    for row in ws.iter_rows(max_row=20, values_only=True):
        for ci, val in enumerate(row):
            if val is not None and '刀' in str(val):
                return ci
    return None


def _try_float(val):
    """Return float(val) or raise ValueError."""
    return float(str(val).strip())


def _load_cells_with_merges(ws):
    """
    Build {(row_1based, col_0based): value} for the whole sheet,
    expanding merged cell ranges so every cell in a merge carries the
    top-left cell's value.  This fixes vertically merged G-column cells
    like G24:G27 where only R24 returns a value from iter_rows.
    """
    data = {}
    for rn, row in enumerate(ws.iter_rows(values_only=True), 1):
        for ci, val in enumerate(row):
            data[(rn, ci)] = val

    for merge in ws.merged_cells.ranges:
        # merged_cells use 1-based column numbers
        top_val = data.get((merge.min_row, merge.min_col - 1))
        for row in range(merge.min_row, merge.max_row + 1):
            for col in range(merge.min_col, merge.max_col + 1):
                data[(row, col - 1)] = top_val

    return data


def parse_cutting_gongcai(wb):
    """
    Parse the Cutting sheet in wb.
    Returns list of dicts for rows where G column (刀數欄) is NOT numeric,
    i.e. machine-cut rows (ATOM / GCN / YINGHUI / LASER / …).

    Confirmed column layout (0-based, consistent across all IE files):
      B(1) = 材料类别
      C(2) = 序号
      D(3) = 部件名称    ← part name
      E(4) = 层数        ← layers
      F(5) = 片数/雙     ← pieces per pair
      G(6) = 標准刀数    ← numeric = normal cut; text = machine type
      H(7) = 正常時間    ← CT normal time (秒/雙)

    Merged cells (e.g. G24:G27 = ATOM spanning 4 rows) are expanded
    so every row in the group carries the correct machine-type value.
    """
    ws = None
    for sh in wb.sheetnames:
        sl = sh.lower()
        if 'cutting' in sl or 'pha cắt' in sl or ('cat' in sl and 'concat' not in sl):
            ws = wb[sh]
            break
    if ws is None:
        return []

    k = _find_knife_col(ws)
    if k is None:
        return []

    mat_col  = k - 5   # B
    part_col = k - 3   # D  部件名称
    lay_col  = k - 2   # E  层数
    pcs_col  = k - 1   # F  片数
    ct_col   = k + 1   # H  正常時間

    # Expand merged cells so we don't miss rows under a vertical merge
    cells = _load_cells_with_merges(ws)
    max_row = ws.max_row

    def get(rn, ci):
        return cells.get((rn, ci))

    result = []
    last_mat = ''

    for rn in range(1, max_row + 1):
        if rn < 12:
            continue

        knife_val = get(rn, k)
        if knife_val is None or str(knife_val).strip() in ('', '.'):
            continue

        s = str(knife_val).strip()

        # Track material category (carried down when empty)
        mat_cell = get(rn, mat_col)
        if mat_cell and str(mat_cell).strip() not in ('', '.'):
            last_mat = str(mat_cell).strip()

        try:
            _try_float(s)
            continue  # numeric → normal cut row
        except ValueError:
            pass

        if '刀' in s:
            continue  # header cell

        # Skip footer/note rows: valid machine types must match known prefix pattern
        if not _MACHINE_RE.match(s):
            continue

        part_raw = get(rn, part_col)
        lay_raw  = get(rn, lay_col)
        pcs_raw  = get(rn, pcs_col)
        ct_raw   = get(rn, ct_col)

        part_name = str(part_raw).strip() if part_raw is not None else ''
        machine   = s.upper()

        try:
            layers = float(lay_raw) if lay_raw is not None else None
        except (ValueError, TypeError):
            layers = None
        try:
            pieces = float(pcs_raw) if pcs_raw is not None else None
        except (ValueError, TypeError):
            pieces = None
        try:
            ct = float(ct_raw) if ct_raw is not None else None
        except (ValueError, TypeError):
            ct = None

        # Skip rows that have no meaningful data after merge expansion
        if not part_name and machine in ('', '.'):
            continue

        result.append({
            'part_name':    part_name,
            'machine_type': machine,
            'layers':       layers,
            'pieces':       pieces,
            'ct':           ct,
            'mat_cat':      last_mat,
        })

    return result


# ── DS-02 model name lookup ───────────────────────────────────────────────────

def _get_model_name(art):
    try:
        conn = db.get_conn()
        row = conn.execute(
            'SELECT model_name FROM ds02_fob WHERE art = ? LIMIT 1', (art,)
        ).fetchone()
        conn.close()
        return (row[0] or '').strip() if row else ''
    except Exception:
        return ''


# ── Main analyze ──────────────────────────────────────────────────────────────

def analyze(ds04_path, group, eolr=120, ie_folder=None):
    """
    Generate 同材共裁 detail for a DS-04 group.

    Args:
        ds04_path:  Path to the DS-04 schedule Excel file.
        group:      Group name (e.g. "加一A組") — matched against sheet titles.
        eolr:       Expected output rate (60 or 120) for IE file selection.
        ie_folder:  Folder containing IE Excel files (defaults to IE_FOLDER_DEFAULT).

    Returns dict:
        ok, group, eolr, sheet_used, total_arts, total_rows,
        missing_ie, rows
    """
    if not HAS_OPENPYXL:
        return {'ok': False, 'error': 'openpyxl not installed'}
    if not os.path.exists(ds04_path):
        return {'ok': False, 'error': f'DS-04 file not found: {ds04_path}'}

    _folder = ie_folder or IE_FOLDER_DEFAULT
    if not os.path.isdir(_folder):
        return {'ok': False, 'error': f'IE folder not found: {_folder}'}

    # ── Step 1: parse DS-04 schedule ──────────────────────────────────────
    try:
        wb04 = openpyxl.load_workbook(ds04_path, data_only=True, read_only=True)
    except Exception as e:
        return {'ok': False, 'error': f'Cannot open DS-04 file: {e}'}

    orders, sheet_used = extract_group_orders(wb04, group)
    wb04.close()

    if not orders:
        return {
            'ok': False,
            'error': f'No orders found for group "{group}"',
            'sheet_used': sheet_used,
        }

    # ── Step 2: build IE index ────────────────────────────────────────────
    ie_index = build_ie_index(_folder)

    # ── Step 3: parse each ART's IE Cutting sheet ─────────────────────────
    rows = []
    missing_ie = []
    stt = 1

    # Sort ARTs for deterministic output
    for art in sorted(orders):
        qty = orders[art]
        ie_path = find_ie_file(art, eolr, ie_index)

        if ie_path is None:
            missing_ie.append(art)
            continue

        model_name = _get_model_name(art)

        try:
            wb_ie = openpyxl.load_workbook(ie_path, data_only=True)
            parts = parse_cutting_gongcai(wb_ie)
            wb_ie.close()
        except Exception:
            missing_ie.append(art)
            continue

        for part in parts:
            rows.append({
                'stt':          stt,
                'lean':         '',
                'model_name':   model_name,
                'art':          art,
                'qty':          qty,
                'part_name':    part['part_name'],
                'machine_type': part['machine_type'],
                'layers':       part['layers'],
                'pieces':       part['pieces'],
                'ct':           part['ct'],
                'checked':      '',
            })
            stt += 1

    return {
        'ok': True,
        'group': group,
        'eolr': eolr,
        'sheet_used': sheet_used,
        'total_arts': len(orders),
        'total_rows': len(rows),
        'missing_ie': missing_ie,
        'rows': rows,
    }


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import argparse, json

    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    parser = argparse.ArgumentParser(description='同材共裁明細自動生成')
    parser.add_argument('file', help='DS-04 schedule Excel path')
    parser.add_argument('--group', required=True, help='Group name (e.g. 加一A組)')
    parser.add_argument('--eolr', type=int, default=120, help='EOLR (default 120)')
    parser.add_argument('--ie-folder', default=IE_FOLDER_DEFAULT,
                        help='IE Excel folder path')
    parser.add_argument('--dry-run', action='store_true',
                        help='Print summary table, no DB writes')
    args = parser.parse_args()

    db.init_db()
    result = analyze(args.file, args.group, args.eolr, args.ie_folder)

    if not result.get('ok'):
        print(f"ERROR: {result.get('error')}")
        sys.exit(1)

    if args.dry_run:
        print(f"Group:      {result['group']}  EOLR={result['eolr']}")
        print(f"Sheet:      {result['sheet_used']}")
        print(f"ARTs:       {result['total_arts']}  |  Rows: {result['total_rows']}")
        if result['missing_ie']:
            print(f"Missing IE ({len(result['missing_ie'])}): {', '.join(result['missing_ie'])}")
        print()
        hdr = (f"{'STT':>4}  {'ART':12}  {'QTY':>7}  "
               f"{'部件名稱(D)':30}  {'機器(G)':12}  {'Lay':>4}  {'Pcs':>4}  {'CT(H)':>8}")
        print(hdr)
        print('-' * len(hdr))
        for r in result['rows']:
            lay = str(int(r['layers'])) if r['layers'] is not None else '-'
            pcs = str(int(r['pieces'])) if r['pieces'] is not None else '-'
            ct  = f"{r['ct']:.2f}" if r['ct'] is not None else '-'
            print(f"{r['stt']:>4}  {r['art']:12}  {r['qty']:>7}  "
                  f"{r['part_name'][:30]:30}  {r['machine_type']:12}  "
                  f"{lay:>4}  {pcs:>4}  {ct:>8}")
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
