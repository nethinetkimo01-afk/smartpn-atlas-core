from flask import Flask, request, jsonify, send_from_directory, session, redirect
from flask_cors import CORS
import database as db
import os
import tempfile
import json as _json
import subprocess
import sys

try:
    from analyze_ds04 import analyze as _ds04_analyze
    HAS_DS04 = True
except ImportError:
    HAS_DS04 = False

try:
    from analyze_ds05 import analyze as _ds05_analyze
    HAS_DS05 = True
except ImportError:
    HAS_DS05 = False

try:
    from analyze_gongcai import analyze as _gongcai_analyze
    HAS_GONGCAI = True
except ImportError:
    HAS_GONGCAI = False

try:
    import openpyxl
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

# ── Excel column maps ────────────────────────────────────────────────────────

DS02_COL_MAP = {
    'article #':'art','article#':'art','article no':'art','article number':'art',
    'art':'art','art #':'art',
    'model #':'model_no','model#':'model_no','model no':'model_no','model name':'model_name',
    'silhouette number':'silhouette_no','silhouette no':'silhouette_no','sil. no':'silhouette_no',
    'factory':'factory','season':'season','category':'category',
    'lc total':'lc_total','lc_total':'lc_total',
    'lc ctb':'lc_ctb','lc_ctb':'lc_ctb','ctb':'lc_ctb',
    'cutting':'lc_cutting','lc cutting':'lc_cutting','lc_cutting':'lc_cutting',
    'stitching':'lc_stitching','lc stitching':'lc_stitching','lc_stitching':'lc_stitching',
    'stockfitting':'lc_stockfitting','lc stockfitting':'lc_stockfitting',
    'stock fitting':'lc_stockfitting','lc_stockfitting':'lc_stockfitting',
    'assembly':'lc_assembly','lc assembly':'lc_assembly','lc_assembly':'lc_assembly',
    'stage':'stage','valid from':'valid_from','valid_from':'valid_from',
}

DS01_COL_MAP = {
    'article id':'article_id','article #':'article_id','article no':'article_id',
    'article number':'article_id','article':'article_id','art':'article_id','art #':'article_id',
    'product type desc':'product_type_desc','product type description':'product_type_desc',
    'product type':'product_type_desc','type desc':'product_type_desc',
    'calendar month':'calendar_month','month':'calendar_month','cal. month':'calendar_month',
    'total':'quantity','quantity':'quantity','qty':'quantity','total qty':'quantity',
}


def _parse_excel(path, col_map, pk_field):
    if not HAS_XLSX:
        return None, 'openpyxl not installed — run: pip install openpyxl'
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_row_idx, headers = None, []
    for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
        cells = [str(c or '').strip().lower() for c in row]
        if sum(1 for c in cells if c in col_map) >= 2:
            header_row_idx = i + 1
            headers = [str(c or '').strip() for c in row]
            break
    if header_row_idx is None:
        return None, 'Cannot find header row (need ≥2 recognised column names)'
    col_to_field = {}
    unmapped = []
    for idx, h in enumerate(headers):
        key = h.lower().strip()
        if key in col_map:
            col_to_field[idx] = col_map[key]
        elif key:
            unmapped.append(h)
    if pk_field not in col_to_field.values():
        return None, f'Primary key column not found. Headers: {headers[:20]}'
    records = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(c is not None and str(c).strip() for c in row):
            continue
        rec, raw = {}, {}
        for idx, val in enumerate(row):
            hdr = headers[idx] if idx < len(headers) else f'col_{idx}'
            raw[hdr] = val
            if idx in col_to_field:
                rec[col_to_field[idx]] = val
        if not str(rec.get(pk_field, '') or '').strip():
            continue
        rec['raw_data'] = _json.dumps(raw, default=str)
        records.append(rec)
    return records, {'total': len(records), 'unmapped_cols': unmapped[:10]}

app = Flask(__name__, static_folder='..', static_url_path='')
app.secret_key = os.environ.get('ATLAS_SECRET', 'smartpn-allocation-demo-key')
CORS(app, supports_credentials=True)

# ── Startup: print 00_MUST_READ_FIRST so every Code session sees the rules ──
_must_read = os.path.join(os.path.dirname(__file__), '..', '00_HANDOFF', '00_MUST_READ_FIRST.md')
try:
    with open(_must_read, encoding='utf-8') as _f:
        print('\n' + '='*60)
        print(_f.read())
        print('='*60 + '\n')
except Exception:
    pass

# ── Allocation users (Phase 1, no password — demo identity switch) ─────────────
ALLOC_USERS = {
    'jim':     {'role': 'admin',     'unit': None,        'name': 'Jim (Admin)'},
    'tongcai': {'role': 'unit_user', 'unit': '同材共裁自動化', 'name': '同材共裁自動化'},
    'dianno':  {'role': 'unit_user', 'unit': '電腦針車折邊',  'name': '電腦針車折邊'},
    'dacu':    {'role': 'unit_user', 'unit': '打粗水洗照射',  'name': '打粗水洗照射'},
}
def _current_user():
    u = session.get('alloc_user')
    if u and u in ALLOC_USERS:
        return {'username': u, **ALLOC_USERS[u]}
    return None

# ── sys_users session auth helper ────────────────────────────────────────────

def _auth_user():
    """Return current sys_users record from session, or None."""
    uid = session.get('user_id')
    if not uid:
        return None
    return db.get_user_by_id(uid)

def _require_admin():
    u = _auth_user()
    if not u or u['role'] != 'admin':
        return jsonify({'ok': False, 'error': '需要管理員權限'}), 403
    return None

def _require_manager():
    """Allow admin or manager roles."""
    u = _auth_user()
    if not u or u['role'] not in ('admin', 'manager'):
        return jsonify({'ok': False, 'error': '需要管理員或主管權限'}), 403
    return None

def _can_edit_ie(header_id):
    """Returns (can_edit: bool, err_tuple_or_None).
    admin: always yes. manager: read-only (唯讀). data_entry: only assigned. others: no."""
    u = _auth_user()
    if not u:
        return False, (jsonify({'ok': False, 'error': '請先登入'}), 401)
    if u['role'] == 'admin':
        return True, None
    if u['role'] == 'manager':
        return False, (jsonify({'ok': False, 'error': 'manager 無法編輯 IE 工序資料'}), 403)
    if u['role'] == 'data_entry':
        assigned = db.get_assigned_header_ids(u['id'])
        if header_id in (assigned or []):
            return True, None
        return False, (jsonify({'ok': False, 'error': '此鞋型未指派給你，無法編輯'}), 403)
    return False, (jsonify({'ok': False, 'error': '您沒有編輯權限'}), 403)

# ── Global login guard ───────────────────────────────────────────────────────

_OPEN_PATHS = {'/login', '/api/login', '/api/allocation/login'}

@app.before_request
def _require_login():
    if request.endpoint == 'static':
        return
    if request.path in _OPEN_PATHS:
        return
    uid = session.get('user_id')
    alloc = session.get('alloc_user')
    if uid or (alloc and alloc in ALLOC_USERS):
        return
    if request.path.startswith('/api/'):
        return jsonify({'ok': False, 'error': '請先登入'}), 401
    return redirect('/login')

# ── Frontend ─────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return send_from_directory('..', 'ds03_ob_interface.html')

@app.route('/ie')
def ie_interface():
    return send_from_directory('..', 'ie_interface.html')

@app.route('/ie/<int:header_id>')
def ie_detail(header_id):
    return send_from_directory('..', 'ie_detail.html')

@app.route('/ie/<int:header_id>/detail')
def ie_cell_detail(header_id):
    return send_from_directory('..', 'ie_cell_detail.html')

@app.route('/ie/matrix')
def ie_matrix_page():
    return send_from_directory('..', 'ie_matrix.html')

@app.route('/ie/cutting')
def ie_cutting_page():
    return send_from_directory('..', 'ie_cutting.html')

@app.route('/allocation')
def allocation_page():
    return send_from_directory('..', 'allocation.html')

# ── IE Interface API ─────────────────────────────────────────────────────────

@app.route('/api/ie/list', methods=['GET'])
def ie_list():
    u = _auth_user()
    result = db.list_ie_records()
    if result.get('ok') and u and u['role'] == 'data_entry':
        assigned = set(db.get_assigned_header_ids(u['id']))
        for rec in result.get('records', []):
            rec['my_assigned'] = rec['id'] in assigned
    return jsonify(result)

@app.route('/api/ie/list_all', methods=['GET'])
def ie_list_all():
    # Admin/manager: returns ALL records regardless of assignment
    err = _require_manager()
    if err: return err
    return jsonify(db.list_ie_records())

@app.route('/api/ie/assignments_by_user', methods=['GET'])
def ie_assignments_by_user():
    err = _require_manager()
    if err: return err
    user_id = request.args.get('user_id')
    if not user_id:
        return jsonify({'ok': False, 'error': 'user_id required'}), 400
    ids = db.get_assigned_header_ids(int(user_id))
    return jsonify({'ok': True, 'header_ids': ids})

@app.route('/api/ie/detail/<int:header_id>')
def ie_detail_api(header_id):
    return jsonify(db.get_ie_model_detail(header_id))

@app.route('/api/ie/<int:header_id>/sheets', methods=['GET'])
def ie_sheet_names(header_id):
    return jsonify(db.get_ie_sheet_names(header_id))

@app.route('/api/ie/<int:header_id>/sheet', methods=['GET'])
def ie_sheet_grid(header_id):
    sheet_name = request.args.get('name', '')
    if not sheet_name:
        return jsonify({'ok': False, 'error': 'name param required'}), 400
    return jsonify(db.get_ie_sheet_grid(header_id, sheet_name))

@app.route('/api/ie/remove_art', methods=['POST'])
def ie_remove_art():
    err = _require_manager()
    if err: return err
    data = request.get_json(force=True)
    art = data.get('art', '').strip()
    header_id = data.get('header_id')
    if not art or not header_id:
        return jsonify({'ok': False, 'error': 'art and header_id required'}), 400
    return jsonify(db.remove_art_from_header(art, int(header_id)))


@app.route('/api/ie/<int:header_id>/delete', methods=['POST'])
def ie_delete_header(header_id):
    err = _require_manager()
    if err: return err
    return jsonify(db.delete_ie_header(header_id))

@app.route('/api/ie/create_record', methods=['POST'])
def ie_create_record():
    err = _require_manager()
    if err: return err
    data = request.get_json(force=True)
    art = (data.get('art') or '').strip()
    model_name = (data.get('model_name') or '').strip()
    eolr = data.get('eolr', 120)
    if not art or not model_name:
        return jsonify({'ok': False, 'error': 'art and model_name required'}), 400
    if int(eolr) not in (60, 120):
        return jsonify({'ok': False, 'error': 'eolr must be 60 or 120'}), 400
    return jsonify(db.create_ie_record(art, model_name, int(eolr)))

@app.route('/api/ie/export/<int:header_id>')
def ie_export(header_id):
    if not HAS_XLSX:
        return jsonify({'ok': False, 'error': 'openpyxl not installed'}), 500
    result = db.get_ie_model_detail(header_id)
    if not result['ok']:
        return jsonify(result), 404
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    hdr_font  = Font(bold=True, color='FFFFFF')
    hdr_fill  = PatternFill('solid', fgColor='1A2744')
    hdr_align = Alignment(horizontal='center', vertical='center')
    thin      = Side(style='thin', color='C0C8D8')
    hdr_bord  = Border(left=thin, right=thin, bottom=thin)

    model_name = result['model_name'] or 'model'
    all_arts   = list({a for h in result['headers'] for a in h.get('arts', [])})
    all_eolrs  = [h['eolr'] for h in result['headers']]

    TABS = [
        ('cutting',    'Cutting'),
        ('atom',       'ATOM-自動化'),
        ('stitch_a',   'Stitching 支流'),
        ('stitch_b',   'Stitching 主流'),
        ('assembly_1', 'Assembly 1'),
        ('assembly_2', 'Assembly 2'),
        ('sum_stock',  'SUM_Stock'),
    ]
    OB_COLS = ['零件(越)', '零件(中)', '材料類型', '層數', '數量/雙',
               '刀模數', 'CT(秒)', '允差(%)', 'ST(秒)',
               '工序', '畫線', '削邊', '貼合', '塗邊', '熱壓']
    OB_KEYS = ['partViet','partZh','matCat','layers','qtyPr',
               'knives','ct','allowance','st','ops','marking','skiving',
               'attaching','edgePaint','heatPress']

    for tab_key, tab_name in TABS:
        ws = wb.create_sheet(tab_name)
        # Gather rows for this tab from all headers
        combined_rows = []
        for h in result['headers']:
            for row in h.get('sheets', {}).get(tab_key, []):
                combined_rows.append(row)

        # Header row
        col_headers = ['EOLR'] + OB_COLS if len(result['headers']) > 1 else OB_COLS
        row_keys    = OB_KEYS
        for ci, ch in enumerate(col_headers, 1):
            cell = ws.cell(row=1, column=ci, value=ch)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = hdr_bord

        if combined_rows:
            for ri, row in enumerate(combined_rows, 2):
                ci = 1
                if len(result['headers']) > 1:
                    ws.cell(row=ri, column=ci, value='').border = Border(left=thin, right=thin)
                    ci += 1
                for k in OB_KEYS:
                    ws.cell(row=ri, column=ci, value=row.get(k, ''))
                    ci += 1
        else:
            ws.cell(row=2, column=1, value='（尚無明細數據）')

        ws.column_dimensions[ws.cell(1,1).column_letter].width = 12
        for ci in range(2, len(col_headers)+1):
            ws.column_dimensions[ws.cell(1,ci).column_letter].width = 14

    # Summary sheet
    ws_sum = wb.create_sheet('MP 彙總')
    sum_headers = ['EOLR', 'ART 列表', '裁斷MP', '針車MP', '成型MP', '庫存MP', '來源']
    for ci, ch in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=ci, value=ch)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = hdr_bord
    for ri, h in enumerate(result['headers'], 2):
        ep = h.get('epph', {})
        ws_sum.cell(row=ri, column=1, value=h['eolr'])
        ws_sum.cell(row=ri, column=2, value=', '.join(h.get('arts', [])))
        ws_sum.cell(row=ri, column=3, value=ep.get('cutting') or None)
        ws_sum.cell(row=ri, column=4, value=ep.get('stitching') or None)
        ws_sum.cell(row=ri, column=5, value=ep.get('assembly') or None)
        ws_sum.cell(row=ri, column=6, value=ep.get('stock') or None)
        ws_sum.cell(row=ri, column=7, value=ep.get('source', ''))
    for ci, w in enumerate([8, 30, 10, 10, 10, 10, 16], 1):
        ws_sum.column_dimensions[ws_sum.cell(1,ci).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = model_name[:30].replace('/', '-').replace('\\', '-')
    filename = f'IE_{safe_name}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/ie/cutting', methods=['GET'])
def ie_cutting_api():
    art  = request.args.get('art', '').strip() or None
    flag = request.args.get('flag', '').strip() or None
    return jsonify(db.get_ie_cutting_process(art=art, flag=flag))

@app.route('/api/ie/cutting/arts', methods=['GET'])
def ie_cutting_arts_api():
    return jsonify(db.get_ie_cutting_arts())

@app.route('/api/ie/<int:header_id>/process', methods=['GET'])
def ie_process_by_header(header_id):
    segment = request.args.get('segment', 'cutting')
    return jsonify(db.get_ie_process_by_header(header_id, segment))

@app.route('/api/ie/cell/<int:header_id>', methods=['GET'])
def ie_cell_data(header_id):
    segment = request.args.get('segment', 'cutting')
    eolr    = request.args.get('eolr', 120)
    return jsonify(db.get_ie_cell_data(header_id, segment, eolr))

@app.route('/api/ie/<int:header_id>/can_edit', methods=['GET'])
def ie_can_edit(header_id):
    ok, _ = _can_edit_ie(header_id)
    return jsonify({'can_edit': ok})

@app.route('/api/bianzhi/can_edit', methods=['GET'])
def bianzhi_can_edit():
    u = _auth_user()
    can = bool(u and u['role'] in ('admin', 'manager'))
    return jsonify({'can_edit': can})

@app.route('/api/ie/stages/<int:header_id>', methods=['GET', 'POST'])
def ie_stages(header_id):
    if request.method == 'POST':
        ok, err = _can_edit_ie(header_id)
        if not ok: return err
        d = request.get_json(force=True)
        return jsonify(db.create_ie_stage(header_id, d.get('stage_name', '新版本')))
    return jsonify(db.get_ie_stages(header_id))

@app.route('/api/ie/cell/save', methods=['POST'])
def ie_cell_save():
    d = request.get_json(force=True)
    hid = db.get_header_id_by_process(d.get('cell_id'))
    ok, err = _can_edit_ie(hid)
    if not ok: return err
    return jsonify(db.save_ie_edit(
        d.get('cell_id'), d.get('stage_id'),
        d.get('field'), d.get('value'), d.get('user', 'anonymous')
    ))

@app.route('/api/ie/cell/approve', methods=['POST'])
def ie_cell_approve():
    d = request.get_json(force=True)
    return jsonify(db.approve_ie_edit(d.get('log_id'), d.get('approver', 'system')))

@app.route('/api/ie/cell/add_row', methods=['POST'])
def ie_cell_add_row():
    d = request.get_json(force=True)
    ok, err = _can_edit_ie(d.get('header_id'))
    if not ok: return err
    return jsonify(db.add_ie_process_row(
        d.get('header_id'), d.get('segment'), d.get('zone'),
        d.get('process_name', '新工序'), d.get('standard_time'),
        d.get('stage_id'), d.get('user', 'demo'),
        d.get('part_name'), d.get('tct'),
        mat_cat=d.get('mat_cat'),
        process_name_zh=d.get('process_name_zh'),
        cut_per_hour=d.get('cut_per_hour'),
        qty_per_pair=d.get('qty_per_pair'),
        layers_per_cut=d.get('layers_per_cut'),
        actual_operators=d.get('actual_operators'),
        normal_time=d.get('normal_time'),
        allowance_pct=d.get('allowance_pct'),
    ))

@app.route('/api/ie/cell/delete_row', methods=['POST'])
def ie_cell_delete_row():
    d = request.get_json(force=True)
    hid = db.get_header_id_by_process(d.get('process_id'))
    ok, err = _can_edit_ie(hid)
    if not ok: return err
    return jsonify(db.delete_ie_process_row(
        d.get('process_id'), d.get('stage_id'), d.get('user', 'demo')
    ))

@app.route('/api/ie/cell/insert_row', methods=['POST'])
def ie_cell_insert_row():
    d = request.get_json(force=True)
    hid = db.get_header_id_by_process(d.get('after_process_id'))
    ok, err = _can_edit_ie(hid)
    if not ok: return err
    return jsonify(db.insert_ie_process_row_after(
        d.get('after_process_id'),
        d.get('process_name', '新工序'),
        d.get('stage_id'), d.get('user', 'demo'),
        part_name=d.get('part_name'), tct=d.get('tct'),
        mat_cat=d.get('mat_cat'),
        process_name_zh=d.get('process_name_zh'),
        cut_per_hour=d.get('cut_per_hour'),
        qty_per_pair=d.get('qty_per_pair'),
        layers_per_cut=d.get('layers_per_cut'),
        actual_operators=d.get('actual_operators'),
        normal_time=d.get('normal_time'),
        allowance_pct=d.get('allowance_pct'),
        standard_time=d.get('standard_time'),
    ))

@app.route('/api/ie/cell/save_group', methods=['POST'])
def ie_cell_save_group():
    d = request.get_json(force=True)
    ok, err = _can_edit_ie(d.get('header_id'))
    if not ok: return err
    return jsonify(db.save_ie_process_group(
        d.get('header_id'), d.get('segment'), d.get('zone'),
        d.get('stage_id'), d.get('process_ids', []),
        d.get('headcount'), d.get('note', '')
    ))

@app.route('/api/ie/cell/update_group', methods=['POST'])
def ie_cell_update_group():
    d = request.get_json(force=True)
    hid = db.get_header_id_by_group(d.get('group_id'))
    ok, err = _can_edit_ie(hid)
    if not ok: return err
    return jsonify(db.update_ie_process_group(d.get('group_id'), d.get('headcount')))

@app.route('/api/ie/cell/delete_group', methods=['POST'])
def ie_cell_delete_group():
    d = request.get_json(force=True)
    hid = db.get_header_id_by_group(d.get('group_id'))
    ok, err = _can_edit_ie(hid)
    if not ok: return err
    return jsonify(db.delete_ie_process_group(d.get('group_id')))

@app.route('/api/ie/<int:header_id>/groups', methods=['GET'])
def ie_get_groups(header_id):
    segment = request.args.get('segment', 'cutting')
    return jsonify(db.get_ie_process_groups(header_id, segment))

@app.route('/api/ie/<int:header_id>/sum', methods=['GET'])
def ie_sum_api(header_id):
    eolr = request.args.get('eolr', 120)
    return jsonify(db.get_ie_sum(header_id, eolr))

@app.route('/ie/<int:header_id>/sum')
def ie_sum_page(header_id):
    return send_from_directory('..', 'ie_sum.html')

@app.route('/api/ie/matrix', methods=['GET'])
def ie_matrix_api():
    return jsonify(db.get_ie_matrix())

# ── IE Zone Import ────────────────────────────────────────────────────────────

@app.route('/api/ie/import/search', methods=['GET'])
def ie_import_search():
    q = (request.args.get('q') or '').strip()
    if not q:
        return jsonify({'ok': True, 'results': []})
    exclude = request.args.get('exclude')
    exclude_id = int(exclude) if exclude and exclude.isdigit() else None
    results = db.search_ie_headers(q, exclude_id)
    return jsonify({'ok': True, 'results': results})

@app.route('/api/ie/import/zones', methods=['GET'])
def ie_import_zones():
    src = request.args.get('source_header_id')
    seg = request.args.get('segment', 'cutting')
    if not src or not src.isdigit():
        return jsonify({'ok': False, 'error': 'source_header_id required'}), 400
    zones = db.get_ie_import_zones(int(src), seg)
    return jsonify({'ok': True, 'zones': zones})

@app.route('/api/ie/import/apply', methods=['POST'])
def ie_import_apply():
    d = request.get_json(force=True)
    target_hid  = d.get('target_header_id')
    source_hid  = d.get('source_header_id')
    segment     = d.get('segment')
    zone        = d.get('zone')
    overwrite   = bool(d.get('overwrite', False))
    if not all([target_hid, source_hid, segment, zone]):
        return jsonify({'ok': False, 'error': '缺少必要參數'}), 400
    target_hid = int(target_hid)
    source_hid = int(source_hid)
    ok, err = _can_edit_ie(target_hid)
    if not ok:
        return err
    if source_hid == target_hid:
        return jsonify({'ok': False, 'error': '來源與目標為同一鞋型'}), 400
    result = db.import_ie_zone(target_hid, source_hid, segment, zone, overwrite)
    return jsonify(result)

@app.route('/api/ie/update_mp', methods=['POST'])
def ie_update_mp():
    data = request.get_json(force=True)
    header_id = data.get('header_id')
    if not header_id:
        return jsonify({'ok': False, 'error': 'header_id required'}), 400
    return jsonify(db.update_ie_mp(
        int(header_id),
        data.get('cutting'), data.get('stitching'),
        data.get('assembly'), data.get('stock')
    ))

# ── 勾選分配系統 (Allocation Phase 1) ───────────────────────────────────────────

@app.route('/api/allocation/login', methods=['POST'])
def alloc_login():
    d = request.get_json(force=True)
    u = (d.get('username') or '').strip().lower()
    if u not in ALLOC_USERS:
        return jsonify({'ok': False, 'error': 'unknown user'}), 401
    session['alloc_user'] = u
    return jsonify({'ok': True, 'username': u, **ALLOC_USERS[u]})

@app.route('/api/allocation/logout', methods=['POST'])
def alloc_logout():
    session.pop('alloc_user', None)
    return jsonify({'ok': True})

@app.route('/api/allocation/me', methods=['GET'])
def alloc_me():
    u = _current_user()
    return jsonify({'ok': True, 'user': u})

@app.route('/api/allocation/units', methods=['GET'])
def alloc_units():
    return jsonify({'ok': True, 'units': db.ALLOCATION_UNITS})

@app.route('/api/allocation/leans', methods=['GET'])
def alloc_leans():
    return jsonify(db.get_allocation_leans(request.args.get('month')))

@app.route('/api/allocation/prefill', methods=['POST'])
def alloc_prefill():
    u = _current_user()
    if not u or u['role'] != 'admin':
        return jsonify({'ok': False, 'error': 'admin only'}), 403
    d = request.get_json(force=True) or {}
    hid = d.get('header_id')
    return jsonify(db.prefill_allocation(int(hid) if hid else None, d.get('month')))

@app.route('/api/allocation/items', methods=['GET'])
def alloc_items():
    u = _current_user()
    unit = request.args.get('unit') or None
    # unit_user is scoped to its own unit regardless of requested unit
    if u and u['role'] == 'unit_user':
        unit = u['unit']
    return jsonify(db.get_allocation_items(
        month=request.args.get('month'), unit=unit,
        lean=request.args.get('lean') or None,
        art=request.args.get('art') or None,
        header_id=request.args.get('header_id') or None))

@app.route('/api/allocation/check', methods=['POST'])
def alloc_check():
    u = _current_user()
    if not u:
        return jsonify({'ok': False, 'error': 'not logged in'}), 401
    d = request.get_json(force=True)
    item_id = d.get('id')
    item = db.get_allocation_item(item_id)
    if not item:
        return jsonify({'ok': False, 'error': 'item not found'}), 404
    # Backend-enforced unit permission (not just front-end hiding)
    if u['role'] != 'admin' and item.get('target_unit') != u['unit']:
        return jsonify({'ok': False,
                        'error': f"forbidden: {u['unit']} cannot edit {item.get('target_unit')}"}), 403
    return jsonify(db.set_allocation_check(item_id, d.get('is_checked'), u['username']))

@app.route('/api/allocation/csa_mp', methods=['GET'])
def alloc_csa_mp():
    return jsonify(db.get_csa_mp(
        lean=request.args.get('lean') or None,
        art=request.args.get('art') or None,
        eolr=request.args.get('eolr', 120),
        month=request.args.get('month')))

@app.route('/api/allocation/parts', methods=['GET'])
def alloc_parts():
    u = _current_user()
    unit = request.args.get('unit') or None
    if u and u['role'] == 'unit_user':
        unit = u['unit']
    return jsonify(db.get_allocation_parts(
        month=request.args.get('month'),
        unit=unit,
        lean=request.args.get('lean') or None
    ))


@app.route('/api/allocation/export_ref', methods=['GET'])
def alloc_export_ref():
    """Download reference xlsx — format differs per unit type."""
    if not HAS_XLSX:
        return jsonify({'ok': False, 'error': 'openpyxl not installed'}), 500
    from flask import send_file
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    unit = request.args.get('unit', '')
    month = request.args.get('month')
    data = db.get_allocation_parts(month=month, unit=unit or None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (unit[:28] if unit else '參考表')

    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='1A2744')
    lean_fill = PatternFill('solid', fgColor='0F1E38')
    thin = Side(style='thin', color='C0C8D8')
    bord = Border(left=thin, right=thin, bottom=thin, top=thin)
    ctr = Alignment(horizontal='center')

    # Unit-specific column definitions
    is_tongcai = unit == '同材共裁自動化' or not unit
    is_dianno  = unit == '電腦針車折邊'
    is_dacu    = unit == '打粗水洗照射'

    if is_dacu:
        headers = ['LEAN', '鞋型', 'ART', '訂單', '流程', 'TCT(秒/雙)', '需求人力', '勾選']
        col_widths = [8, 22, 10, 8, 22, 11, 11, 8]
    elif is_dianno:
        headers = ['LEAN', '鞋型', 'ART', '訂單', '配件', 'CT標準時間', 'Output', '理論人數', '勾選']
        col_widths = [8, 22, 10, 8, 22, 12, 9, 10, 8]
    else:  # 同材共裁 or all
        headers = ['LEAN', '鞋型', 'ART', '訂單', '部件名稱', 'Zone',
                   '刀數/H', '層數', '片數/雙', '總片數', 'CT秒/雙', 'Output', '理論人數', '勾選']
        col_widths = [8, 22, 10, 8, 22, 9, 9, 7, 9, 9, 9, 9, 9, 8]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.border = bord; c.alignment = ctr

    ri = 2
    for lg in data.get('leans', []):
        lean_label = lg.get('lean') or '—'
        for mg in lg.get('models', []):
            order_qty = mg.get('order_qty') or 0
            for it in mg.get('items', []):
                ct = it.get('ct_sec')
                chk = '外移' if it.get('is_checked') else ('留CSA(主裁斷)' if it.get('is_csa_locked') else '留CSA')
                if is_dacu:
                    demand = round(order_qty * (ct or 0) / 3600 / 222, 2) if ct and order_qty else None
                    vals = [lean_label, mg.get('model_name', ''), mg.get('art', ''), order_qty,
                            it.get('part_name', ''), ct, demand, chk]
                elif is_dianno:
                    vals = [lean_label, mg.get('model_name', ''), mg.get('art', ''), order_qty,
                            it.get('part_name', ''), ct, it.get('output'), it.get('theory_mp'), chk]
                else:
                    vals = [lean_label, mg.get('model_name', ''), mg.get('art', ''), order_qty,
                            it.get('part_name', ''), it.get('zone', ''),
                            it.get('cut_per_hour'), it.get('layers'), it.get('qty_per_pair'),
                            it.get('total_pieces'), ct, it.get('output'), it.get('theory_mp'), chk]
                for ci, v in enumerate(vals, 1):
                    ws.cell(row=ri, column=ci, value=v).border = bord
                ri += 1
            # model subtotal
            n_cols = len(headers)
            sub_c = ws.cell(row=ri, column=n_cols - 1, value=mg.get('ie_mp'))
            sub_c.font = Font(bold=True)
            ws.cell(row=ri, column=n_cols, value='小計').font = Font(bold=True)
            ri += 1
        # LEAN footer
        c = ws.cell(row=ri, column=1, value=f'LEAN {lean_label} 合計')
        c.font = Font(bold=True, color='FFFFFF'); c.fill = lean_fill
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=3)
        ws.cell(row=ri, column=4,
                value=f'外移MP: {lg.get("allocated_mp")}  留CSA: {lg.get("csa_mp")}')
        ri += 1

    if ri == 2:
        ws.cell(row=2, column=1, value='（尚無資料 — 請先預填）')

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    safe_unit = unit.replace('/', '_') if unit else 'all'
    return send_file(buf, as_attachment=True,
                     download_name=f'ref_{safe_unit}_{month or "all"}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/allocation/export', methods=['GET'])
def alloc_export():
    if not HAS_XLSX:
        return jsonify({'ok': False, 'error': 'openpyxl not installed'}), 500
    unit = request.args.get('unit', '')
    if unit not in db.ALLOCATION_UNITS:
        return jsonify({'ok': False, 'error': 'unknown unit'}), 400
    month = request.args.get('month')
    data = db.get_allocation_export_rows(unit, month)
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = unit[:28]
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='1A2744')
    thin = Side(style='thin', color='C0C8D8'); bord = Border(left=thin, right=thin, bottom=thin, top=thin)
    if unit == '打粗水洗照射':
        headers = ['LEAN', 'ART', '工序', 'TCT', '需求人力(=訂單÷(3600÷TCT)÷222)']
        keys = lambda r: [r.get('lean') or '', r.get('art') or '', r.get('process') or '',
                          r.get('tct'), r.get('headcount')]
    else:
        headers = ['LEAN', 'ART', '部件', '工序', 'TCT', '理論MP']
        keys = lambda r: [r.get('lean') or '', r.get('art') or '', r.get('part_name') or '',
                          r.get('process') or '', r.get('tct'), r.get('theory_mp')]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.border = bord
        c.alignment = Alignment(horizontal='center', vertical='center')
    for ri, r in enumerate(data.get('rows', []), 2):
        for ci, v in enumerate(keys(r), 1):
            ws.cell(row=ri, column=ci, value=v)
    if not data.get('rows'):
        ws.cell(row=2, column=1, value='（尚無已勾選外移工序 — 先在 /allocation 勾選並由 admin 預填）')
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = 22 if ci == len(headers) else 14
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = {'同材共裁自動化': 'unit_同材共裁.xlsx', '電腦針車折邊': 'unit_電腦針車.xlsx',
             '打粗水洗照射': 'unit_打粗水洗.xlsx'}[unit]
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── DS-03 OB ─────────────────────────────────────────────────────────────────

@app.route('/api/ds03/save', methods=['POST'])
def save_ob():
    data = request.get_json(force=True)
    return jsonify(db.save_ob_record(data))

@app.route('/api/ds03/load', methods=['GET'])
def load_ob():
    art  = request.args.get('art', '')
    eolr = request.args.get('eolr', 60)
    run  = request.args.get('run', 1)
    if not art:
        return jsonify({'ok': False, 'error': 'art is required'}), 400
    return jsonify(db.load_ob_record(art, eolr, run))

@app.route('/api/ds03/list', methods=['GET'])
def list_ob():
    return jsonify(db.list_ob_records())

@app.route('/api/ds03/delete', methods=['DELETE'])
def delete_ob():
    art  = request.args.get('art', '')
    eolr = request.args.get('eolr', 60)
    run  = request.args.get('run', 1)
    return jsonify(db.delete_ob_record(art, eolr, run))

@app.route('/api/ds03/add_art', methods=['POST'])
def add_art():
    data = request.get_json(force=True)
    art = data.get('art', '').strip()
    header_id = data.get('header_id')
    if not art or not header_id:
        return jsonify({'ok': False, 'error': 'art and header_id required'}), 400
    return jsonify(db.add_art_to_header(art, int(header_id)))

# ── Lookup ────────────────────────────────────────────────────────────────────

@app.route('/api/lookup/all', methods=['GET'])
def get_lookup():
    return jsonify(db.get_all_lookup())

@app.route('/api/lookup/add', methods=['POST'])
def add_lookup():
    data = request.get_json(force=True)
    return jsonify(db.add_lookup_entry(data.get('viet',''), data.get('zh','')))

# ── DS-02 cross-table helper ─────────────────────────────────────────────────

@app.route('/api/ds02/epph', methods=['GET'])
def get_epph():
    art = request.args.get('art', '')
    if not art:
        return jsonify({'ok': False, 'error': 'art is required'}), 400
    return jsonify(db.get_epph_by_art(art))

# ── Import Admin ──────────────────────────────────────────────────────────────

@app.route('/admin')
def admin_page():
    return send_from_directory('..', 'import_admin.html')

@app.route('/api/stats', methods=['GET'])
def get_stats():
    return jsonify(db.get_db_stats())

@app.route('/api/ds02/upload', methods=['POST'])
def upload_ds02():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'ok': False, 'error': 'Empty filename'}), 400
    suffix = '.xlsx' if f.filename.lower().endswith('.xlsx') else '.xls'
    tmp = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
    tmp_path = tmp.name
    tmp.close()
    try:
        f.save(tmp_path)
        records, info = _parse_excel(tmp_path, DS02_COL_MAP, 'art')
        if records is None:
            return jsonify({'ok': False, 'error': info})
        result = db.import_ds02_records(records)
        result['file_info'] = info
        return jsonify(result)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

@app.route('/api/ds01/upload', methods=['POST'])
def upload_ds01():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'ok': False, 'error': 'Empty filename'}), 400
    suffix = '.xlsx' if f.filename.lower().endswith('.xlsx') else '.xls'
    tmp = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
    tmp_path = tmp.name
    tmp.close()
    try:
        f.save(tmp_path)
        records, info = _parse_excel(tmp_path, DS01_COL_MAP, 'article_id')
        if records is None:
            return jsonify({'ok': False, 'error': info})
        result = db.import_ds01_records(records)
        result['file_info'] = info
        return jsonify(result)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

@app.route('/api/ds02/list', methods=['GET'])
def list_ds02():
    limit  = int(request.args.get('limit', 200))
    offset = int(request.args.get('offset', 0))
    return jsonify(db.list_ds02_records(limit, offset))

@app.route('/api/ds01/list', methods=['GET'])
def list_ds01():
    limit  = int(request.args.get('limit', 200))
    offset = int(request.args.get('offset', 0))
    return jsonify(db.list_ds01_records(limit, offset))

# ── DS-04 Schedule Analyzer ───────────────────────────────────────────────────

@app.route('/api/ds04/analyze', methods=['GET'])
def ds04_analyze():
    if not HAS_DS04:
        return jsonify({'ok': False, 'error': 'analyze_ds04 module not available'}), 500
    file_path = request.args.get('file', '').strip()
    dept      = request.args.get('dept', '').strip()
    group     = request.args.get('group', '').strip()
    try:
        eolr = int(request.args.get('eolr', 120))
    except ValueError:
        eolr = 120
    if not file_path:
        return jsonify({'ok': False, 'error': 'file parameter required'}), 400
    result = _ds04_analyze(file_path, dept, group, eolr)
    return jsonify(result)

# ── DS-05 大底課進度表 Analyzer ───────────────────────────────────────────────

@app.route('/api/ds05/analyze', methods=['GET'])
def ds05_analyze():
    if not HAS_DS05:
        return jsonify({'ok': False, 'error': 'analyze_ds05 module not available'}), 500
    file_path    = request.args.get('file', '').strip()
    group_filter = request.args.get('group', '').strip()
    if not file_path:
        return jsonify({'ok': False, 'error': 'file parameter required'}), 400
    result = _ds05_analyze(file_path, group_filter)
    return jsonify(result)

# ── 同材共裁 Analyzer ─────────────────────────────────────────────────────────

@app.route('/api/gongcai/analyze', methods=['GET'])
def gongcai_analyze():
    if not HAS_GONGCAI:
        return jsonify({'ok': False, 'error': 'analyze_gongcai module not available'}), 500
    file_path  = request.args.get('file', '').strip()
    group      = request.args.get('group', '').strip()
    ie_folder  = request.args.get('ie_folder', '').strip() or None
    try:
        eolr = int(request.args.get('eolr', 120))
    except ValueError:
        eolr = 120
    if not file_path:
        return jsonify({'ok': False, 'error': 'file parameter required'}), 400
    if not group:
        return jsonify({'ok': False, 'error': 'group parameter required'}), 400
    result = _gongcai_analyze(file_path, group, eolr, ie_folder)
    return jsonify(result)

# ── Result correction import ─────────────────────────────────────────────────

@app.route('/api/result/import_corrections', methods=['POST'])
def import_corrections():
    """
    Accept a corrected comparison_table.xlsx and write Jim's manual MP values
    to ob_epph (source='manual_correction').

    Expects multipart/form-data with file='comparison_table.xlsx'.
    Reads rows where 狀態 == 'MISMATCH' or 'MISSING_IE' and 結果表Cut/Stitch/Asm
    columns have values — those are Jim's corrected values.
    """
    if not HAS_XLSX:
        return jsonify({'ok': False, 'error': 'openpyxl not available'}), 500

    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'file field required'}), 400

    f = request.files['file']
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        f.save(tmp.name)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        ws = wb.active
        headers = [str(c.value or '').strip() for c in next(ws.iter_rows(max_row=1))]

        # Column indices (0-based after converting to list)
        def col(name):
            try: return headers.index(name)
            except ValueError: return None

        c_art    = col('ART')
        c_lean   = col('LEAN')
        c_qty    = col('訂單')
        c_status = col('狀態')
        c_rc     = col('結果表Cut')
        c_rs     = col('結果表Stitch')
        c_ra     = col('結果表Asm')

        if c_art is None:
            return jsonify({'ok': False, 'error': 'ART column not found'}), 400

        updated = []
        skipped = []

        conn = db.get_conn()
        ts = db.now_iso()

        for row in ws.iter_rows(min_row=2, values_only=True):
            art    = str(row[c_art] or '').strip() if c_art is not None else ''
            status = str(row[c_status] or '').strip() if c_status is not None else ''
            if not art or status == 'OK':
                continue

            def _fv(idx):
                if idx is None: return None
                v = row[idx]
                if v is None or str(v).strip() in ('', '-', 'None'): return None
                try: return float(v)
                except: return None

            ref_cut = _fv(c_rc)
            ref_s   = _fv(c_rs)
            ref_asm = _fv(c_ra)

            if ref_cut is None and ref_s is None and ref_asm is None:
                skipped.append(art)
                continue

            # Find or create ob_header for this ART via ob_articles (EOLR=120 default)
            h_row = conn.execute(
                '''SELECT a.header_id AS id FROM ob_articles a
                   WHERE a.art=? ORDER BY a.id DESC LIMIT 1''', (art,)
            ).fetchone()

            if h_row:
                header_id = h_row[0]
                e_row = conn.execute(
                    'SELECT id FROM ob_epph WHERE header_id=?', (header_id,)
                ).fetchone()
                if e_row:
                    fields = []
                    vals = []
                    if ref_cut is not None: fields.append('cutting=?');   vals.append(ref_cut)
                    if ref_s   is not None: fields.append('stitching=?'); vals.append(ref_s)
                    if ref_asm is not None: fields.append('assembly=?');  vals.append(ref_asm)
                    fields.append("source='manual_correction'")
                    conn.execute(
                        f'UPDATE ob_epph SET {",".join(fields)} WHERE header_id=?',
                        vals + [header_id]
                    )
                else:
                    conn.execute(
                        '''INSERT INTO ob_epph (header_id, cutting, stitching, assembly, stock, source)
                           VALUES (?,?,?,?,0,'manual_correction')''',
                        (header_id, ref_cut or 0, ref_s or 0, ref_asm or 0)
                    )
            else:
                cur = conn.execute(
                    '''INSERT INTO ob_header (model_name, season, material, category, eolr, run, created_at, updated_at)
                       VALUES ('', '', '', '', 120, 1, ?, ?)''', (ts, ts)
                )
                header_id = cur.lastrowid
                conn.execute(
                    'INSERT INTO ob_articles (header_id, art) VALUES (?,?)', (header_id, art)
                )
                conn.execute(
                    '''INSERT INTO ob_epph (header_id, cutting, stitching, assembly, stock, source)
                       VALUES (?,?,?,?,0,'manual_correction')''',
                    (header_id, ref_cut or 0, ref_s or 0, ref_asm or 0)
                )

            updated.append({'art': art, 'cut': ref_cut, 'stitch': ref_s, 'asm': ref_asm})

        conn.commit()
        conn.close()

        return jsonify({
            'ok': True,
            'updated': len(updated),
            'skipped': len(skipped),
            'rows': updated[:20],
        })

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        try: os.unlink(tmp_path)
        except: pass


@app.route('/ds04')
def ds04_page():
    return send_from_directory('..', 'ds04.html')

@app.route('/eolr-settings')
def eolr_settings_page():
    return send_from_directory('..', 'eolr_settings.html')

@app.route('/bianche')
def bianche_page():
    return send_from_directory('..', 'bianche.html')

# ── 廠務組織編制表 (bianzhi) ──────────────────────────────────────────────────

@app.route('/api/bianzhi/summary', methods=['GET'])
def bianzhi_summary():
    month = request.args.get('month', '2026-06')
    return jsonify(db.get_bianzhi_summary(month))

@app.route('/api/bianzhi/detail', methods=['GET'])
def bianzhi_detail():
    month = request.args.get('month', '2026-06')
    return jsonify(db.get_bianzhi_detail(month))

@app.route('/api/bianzhi/lean_bianzhi', methods=['POST'])
def bianzhi_lean_bianzhi():
    err = _require_manager()
    if err: return err
    d = request.get_json(force=True)
    return jsonify(db.set_bianche_model_manual(
        d.get('lean'), d.get('model_name'), d.get('month'),
        headcount=d.get('bianzhi')))

@app.route('/api/bianzhi/unit_manual', methods=['POST'])
def bianzhi_unit_manual():
    err = _require_manager()
    if err: return err
    d = request.get_json(force=True)
    return jsonify(db.set_bianzhi_unit_manual(d.get('month'), d.get('unit'), d.get('field'), d.get('value')))

@app.route('/api/bianzhi/monthly_manual', methods=['POST'])
def bianzhi_monthly_manual_post():
    err = _require_manager()
    if err: return err
    d = request.get_json(force=True)
    return jsonify(db.set_bianzhi_monthly_manual(d.get('month'), d.get('key'), d.get('value')))

# ── DS-04 API ─────────────────────────────────────────────────────────────────

@app.route('/api/ds04/order', methods=['POST'])
def ds04_add_order():
    data = request.get_json(force=True)
    return jsonify(db.ds04_add_order(data))

@app.route('/api/ds04/order/<int:order_id>', methods=['PUT'])
def ds04_update_order(order_id):
    data = request.get_json(force=True)
    return jsonify(db.ds04_update_order(order_id, data))

@app.route('/api/ds04/order/<int:order_id>', methods=['DELETE'])
def ds04_delete_order(order_id):
    month = request.args.get('month', '')
    return jsonify(db.ds04_delete_order(order_id, month))

@app.route('/api/ds04/lock', methods=['GET'])
def get_ds04_lock():
    month = request.args.get('month', '2026-06')
    return jsonify(db.ds04_get_lock_status(month))

@app.route('/api/ds04/lock', methods=['POST'])
def lock_ds04():
    data = request.get_json(force=True)
    action = data.get('action', 'lock')
    month  = data.get('month', '2026-06')
    if action == 'unlock':
        return jsonify(db.ds04_unlock_month(month))
    return jsonify(db.ds04_lock_month(month, data.get('locked_by', '')))

@app.route('/api/eolr-settings', methods=['GET'])
def get_eolr_settings():
    month = request.args.get('month', '2026-06')
    return jsonify(db.get_eolr_settings(month))

@app.route('/api/eolr-settings', methods=['POST'])
def set_eolr_setting():
    data = request.get_json(force=True)
    return jsonify(db.set_eolr_setting(
        data.get('lean'), data.get('month', '2026-06'),
        data.get('eolr', 120), data.get('updated_by', '')
    ))

@app.route('/api/bianche', methods=['GET'])
def get_bianche():
    month = request.args.get('month', '2026-06')
    return jsonify(db.get_bianche_data(month))

@app.route('/api/bianche/manual', methods=['POST'])
def set_bianche_manual():
    data = request.get_json(force=True)
    return jsonify(db.set_bianche_manual(
        data.get('lean'), data.get('month', '2026-06'),
        data.get('manager_mp', 0), data.get('headcount', 0),
        data.get('updated_by', '')
    ))

@app.route('/api/ds04/orders', methods=['GET'])
def ds04_orders():
    dept     = request.args.get('dept', '')
    lean     = request.args.get('lean', '')
    outsource = request.args.get('outsource', '')
    return jsonify(db.ds04_get_orders(
        dept=dept or None,
        lean=lean or None,
        outsource=outsource or None
    ))

@app.route('/api/ds04/filters', methods=['GET'])
def ds04_filters():
    return jsonify(db.ds04_get_filters())

@app.route('/api/ds04/export', methods=['GET'])
def ds04_export():
    if not HAS_XLSX:
        return jsonify({'ok': False, 'error': 'openpyxl not installed'}), 500
    dept     = request.args.get('dept', '')
    lean     = request.args.get('lean', '')
    outsource = request.args.get('outsource', '')
    result = db.ds04_get_orders(
        dept=dept or None, lean=lean or None, outsource=outsource or None
    )
    if not result['ok']:
        return jsonify(result), 500
    import io as _io
    from openpyxl.styles import Font, PatternFill
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = 'DS04進度表'
    headers = ['部門', 'LEAN', '鞋型名稱', 'ART', '訂單號', '數量', '交期', '外包鞋面']
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='DDEEFF')
    for r in result['rows']:
        ws2.append([
            r['dept'], r['lean'], r['model_name'], r['art'],
            r['order_no'], r['qty'], r['delivery_date'],
            'Y' if r['is_outsource_upper'] else ''
        ])
    buf = _io.BytesIO()
    wb2.save(buf)
    buf.seek(0)
    from flask import send_file
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='ds04_進度表.xlsx')

# ── User management (/admin/users) ───────────────────────────────────────────

@app.route('/admin/users')
def admin_users_page():
    err = _require_manager()
    if err: return err
    return send_from_directory('..', 'admin_users.html')

@app.route('/api/users', methods=['GET'])
def api_users_list():
    err = _require_manager()
    if err: return err
    return jsonify({'ok': True, 'users': db.list_users()})

@app.route('/api/users', methods=['POST'])
def api_users_create():
    err = _require_manager()
    if err: return err
    me = _auth_user()
    d = request.get_json(force=True) or {}
    role = d.get('role', 'read_only')
    if me.get('role') != 'admin' and role == 'admin':
        return jsonify({'ok': False, 'error': '無權建立管理員帳號'}), 403
    return jsonify(db.create_user(
        d.get('username',''), d.get('display_name',''),
        role, d.get('password',''),
        d.get('active',1)
    ))

@app.route('/api/users/<int:uid>', methods=['PUT'])
def api_users_update(uid):
    err = _require_manager()
    if err: return err
    me = _auth_user()
    d = request.get_json(force=True) or {}
    if me.get('role') != 'admin':
        if d.get('role') == 'admin':
            return jsonify({'ok': False, 'error': '無權設定管理員角色'}), 403
        target = db.get_user_by_id(uid)
        if target and target.get('role') == 'admin':
            return jsonify({'ok': False, 'error': '無權修改管理員帳號'}), 403
    return jsonify(db.update_user(
        uid,
        display_name=d.get('display_name'),
        role=d.get('role'),
        password=d.get('password') or None,
        active=d.get('active')
    ))

@app.route('/api/users/<int:uid>', methods=['DELETE'])
def api_users_delete(uid):
    err = _require_manager()
    if err: return err
    me = _auth_user()
    if me.get('role') != 'admin':
        target = db.get_user_by_id(uid)
        if target and target.get('role') == 'admin':
            return jsonify({'ok': False, 'error': '無權刪除管理員帳號'}), 403
    return jsonify(db.delete_user(uid))

# ── Test-output file downloads ───────────────────────────────────────────────

@app.route('/download/<path:filename>')
def download_test_output(filename):
    from flask import send_from_directory
    safe = os.path.basename(filename)
    return send_from_directory(
        os.path.join(os.path.dirname(__file__), 'test_output'),
        safe, as_attachment=True
    )

# ── Bianche CSA (model-level) + OCS/RB/QC + Alloc lock ──────────────────────

@app.route('/api/bianche/csa', methods=['GET'])
def bianche_csa():
    return jsonify(db.get_bianche_csa_data(request.args.get('month', '2026-06')))

@app.route('/api/bianche/model_manual', methods=['POST'])
def bianche_model_manual():
    d = request.get_json(force=True) or {}
    return jsonify(db.set_bianche_model_manual(
        d.get('lean'), d.get('model_name'), d.get('month', '2026-06'),
        d.get('manager_mp'), d.get('headcount'), d.get('updated_by', '')
    ))

@app.route('/api/bianche/dept/<dept>', methods=['GET'])
def bianche_dept(dept):
    return jsonify(db.get_bianche_dept(dept, request.args.get('month', '2026-06')))

@app.route('/api/bianche/dept_hc', methods=['POST'])
def bianche_dept_hc():
    d = request.get_json(force=True) or {}
    return jsonify(db.set_bianche_dept_hc(
        d.get('dept'), d.get('group'), d.get('month', '2026-06'),
        d.get('headcount'), d.get('shoe_detail', ''), d.get('updated_by', '')
    ))

@app.route('/api/bianche/lean_hc', methods=['GET'])
def bianche_lean_hc_get():
    return jsonify(db.get_bianche_lean_hcs(request.args.get('month', '2026-06')))

@app.route('/api/bianche/lean_hc', methods=['POST'])
def bianche_lean_hc_post():
    d = request.get_json(force=True) or {}
    return jsonify(db.set_bianche_lean_hc(
        d.get('lean'), d.get('month', '2026-06'), d.get('headcount')
    ))

@app.route('/api/bianche/export', methods=['GET'])
def bianche_export_xlsx():
    if not HAS_XLSX:
        return jsonify({'ok': False, 'error': 'openpyxl not installed'}), 500
    month = request.args.get('month', '2026-06')
    data = db.get_bianche_export_data(month)
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from flask import send_file

    wb = openpyxl.Workbook()
    wb.security  # make sure security works
    thin = Side(style='thin', color='C5D0E0')
    bord = Border(left=thin, right=thin, bottom=thin, top=thin)
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    grey_fill = PatternFill('solid', fgColor='E8EEF8')
    ctr = Alignment(horizontal='center', vertical='center')
    lock_prot = Protection(locked=True)
    edit_prot = Protection(locked=False)

    def _hdr(ws, row, cols):
        for ci, h in enumerate(cols, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = hdr_font; c.fill = hdr_fill; c.border = bord; c.alignment = ctr; c.protection = lock_prot

    def _val(ws, row, col, v, editable=False, fmt=None):
        c = ws.cell(row=row, column=col, value=v)
        c.border = bord; c.protection = edit_prot if editable else lock_prot
        if editable: c.fill = PatternFill('solid', fgColor='FFFDE7')
        if fmt: c.number_format = fmt

    # === CSA Sheet ===
    LEAN_ORDER_LIST = [
        '1A','1B','1C','2A','2B','2C','3A','3B','3C','4A','4B','4C',
        '5A','5B','5C','6A','6B','6C','7A','7B','7C','8A','8B','8C',
        '9A','9B','9C','10A','10B','10C','11A1','11A2','11B1','11B2','11C'
    ]
    def _lean_key(l): i = LEAN_ORDER_LIST.index(l) if l in LEAN_ORDER_LIST else 999; return i

    ws = wb.active; ws.title = 'CSA'
    _hdr(ws, 1, ['LEAN','鞋型名稱','ART','訂單','裁斷MP','針車MP','成型MP','協理給','LEAN編制'])
    ri = 2
    csa_data = data.get('csa', {})
    rows = csa_data.get('rows', [])
    lean_hc = csa_data.get('lean_hc', {})
    by_lean = {}
    for r in rows:
        by_lean.setdefault(r['lean'], []).append(r)
    for lean in sorted(by_lean.keys(), key=_lean_key):
        lrows = by_lean[lean]
        lean_headcount = lean_hc.get(lean, 0)
        for i, r in enumerate(lrows):
            s = lambda v: round(float(v), 4) if v is not None else None
            hc_val = lean_headcount if i == 0 else None
            hc_ed  = i == 0
            for ci, (v, ed) in enumerate([
                (r['lean'], False), (r['model_name'], False), (r['arts'], False), (r['qty'], False),
                (s(r['cutting_mp']), False), (s(r['stitching_mp']), False), (s(r['assembly_mp']), False),
                (r['manager_mp'] or 0, True), (hc_val, hc_ed)
            ], 1):
                _val(ws, ri, ci, v, ed)
            ri += 1
    ws.protection.sheet = True; ws.protection.password = 'atlas2026'

    # === OCS / RB / QC sheets ===
    for dept_key, dept_data in [('OCS', data.get('ocs',{})), ('RB', data.get('rb',{})), ('QC', data.get('qc',{}))]:
        ws2 = wb.create_sheet(dept_key)
        _hdr(ws2, 1, ['部門','組別','鞋型明細','上月編制','本月編制'])
        ri = 2
        for sec in dept_data.get('sections', []):
            for g in sec.get('groups', []):
                _val(ws2, ri, 1, dept_key, False)
                _val(ws2, ri, 2, g['group'], False)
                _val(ws2, ri, 3, g.get('shoe_detail',''), True)
                _val(ws2, ri, 4, g.get('last_month_hc'), False)
                _val(ws2, ri, 5, g.get('this_month_hc', 0), True)
                ri += 1
        ws2.protection.sheet = True; ws2.protection.password = 'atlas2026'

    for ws_s in wb.worksheets:
        ws_s.column_dimensions['A'].width = 14
        ws_s.column_dimensions['B'].width = 22
        ws_s.column_dimensions['C'].width = 16

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'廠務編制表_{month}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/bianche/import_manual', methods=['POST'])
def bianche_import_manual():
    """Import xlsx: only read 協理給/編制/本月編制 columns."""
    if not HAS_XLSX:
        return jsonify({'ok': False, 'error': 'openpyxl not installed'}), 500
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'file required'}), 400
    month = request.form.get('month', '2026-06')
    f = request.files['file']
    import io, openpyxl, tempfile, os as _os
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    f.save(tmp.name); tmp.close()
    updated_csa = updated_dept = 0
    try:
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
        # CSA sheet: cols A=LEAN B=model C=ART D=qty E=裁斷 F=針 G=成型 H=協理給 I=LEAN編制
        if 'CSA' in wb.sheetnames:
            ws = wb['CSA']
            seen_lean_hc = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                lean = str(row[0] or '').strip()
                model = str(row[1] or '').strip()
                if not lean or not model:
                    continue
                mgr = row[7]
                lean_hc_val = row[8]
                if mgr is not None:
                    db.set_bianche_model_manual(lean, model, month, mgr, None, 'import')
                    updated_csa += 1
                if lean_hc_val is not None and lean not in seen_lean_hc:
                    db.set_bianche_lean_hc(lean, month, lean_hc_val)
                    seen_lean_hc.add(lean)
        # OCS/RB/QC: cols A=dept B=group C=shoe_detail D=last_hc E=this_hc
        for dept_key in ['OCS','RB','QC']:
            if dept_key in wb.sheetnames:
                ws = wb[dept_key]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    dept = str(row[0] or '').strip()
                    grp  = str(row[1] or '').strip()
                    shoe = str(row[2] or '').strip()
                    hc   = row[4]
                    if dept and grp and hc is not None:
                        db.set_bianche_dept_hc(dept, grp, month, hc, shoe, 'import')
                        updated_dept += 1
        return jsonify({'ok': True, 'updated_csa': updated_csa, 'updated_dept': updated_dept})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        try: _os.unlink(tmp.name)
        except: pass

@app.route('/api/allocation/lock', methods=['GET'])
def alloc_lock_get():
    return jsonify(db.alloc_get_lock(request.args.get('month', '2026-06')))

@app.route('/api/allocation/lock', methods=['POST'])
def alloc_lock_post():
    u = _current_user()
    if not u or u['role'] != 'admin':
        return jsonify({'ok': False, 'error': 'admin only'}), 403
    d = request.get_json(force=True) or {}
    month = d.get('month', '2026-06')
    if d.get('action') == 'unlock':
        return jsonify(db.alloc_unlock_month(month))
    return jsonify(db.alloc_lock_month(month, u.get('username', '')))


@app.route('/api/allocation/fix_defaults', methods=['POST'])
def alloc_fix_defaults():
    """Admin: set all non-裁斷機 items to is_checked=1 for a month."""
    u = _current_user()
    if not u or u['role'] != 'admin':
        return jsonify({'ok': False, 'error': 'admin only'}), 403
    d = request.get_json(force=True) or {}
    month = d.get('month', '2026-06')
    return jsonify(db.alloc_fix_default_checked(month))

# ── Login / Logout / Me ──────────────────────────────────────────────────────

@app.route('/login')
def login_page():
    return send_from_directory('..', 'login.html')

@app.route('/api/login', methods=['POST'])
def api_login():
    d = request.get_json(force=True) or {}
    username = (d.get('username') or '').strip().lower()
    password = d.get('password') or ''
    user = db.verify_login(username, password)
    if not user:
        return jsonify({'ok': False, 'error': '帳號或密碼錯誤'}), 401
    session['user_id'] = user['id']
    # Also keep alloc_user for backward compatibility if unit user
    UNIT_MAP = {'tongcai': '同材共裁自動化', 'dianno': '電腦針車折邊', 'dacu': '打粗水洗照射'}
    if username in UNIT_MAP:
        session['alloc_user'] = username
    return jsonify({'ok': True, 'user': {
        'id': user['id'], 'username': user['username'],
        'display_name': user['display_name'], 'role': user['role']
    }})

@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'ok': True})

@app.route('/api/me', methods=['GET'])
def api_me():
    u = _auth_user()
    if not u:
        return jsonify({'ok': False, 'user': None})
    return jsonify({'ok': True, 'user': u})

# ── IE header EOLR / season update ──────────────────────────────────────────

@app.route('/api/ie/<int:header_id>/update_eolr', methods=['POST'])
def ie_update_eolr(header_id):
    err = _require_manager()
    if err: return err
    d = request.get_json(force=True) or {}
    return jsonify(db.update_ie_header_eolr(header_id, d.get('eolr', 120)))

@app.route('/api/ie/<int:header_id>/update_season', methods=['POST'])
def ie_update_season(header_id):
    err = _require_manager()
    if err: return err
    d = request.get_json(force=True) or {}
    return jsonify(db.update_ie_header_season(header_id, d.get('season', '')))

@app.route('/api/ie/<int:header_id>/update_meta', methods=['POST'])
def ie_update_meta(header_id):
    err = _require_manager()
    if err: return err
    d = request.get_json(force=True) or {}
    return jsonify(db.update_ie_header_meta(
        header_id,
        season=d.get('season'),
        material=d.get('material'),
    ))

@app.route('/api/ie/add_art', methods=['POST'])
def ie_add_art():
    err = _require_manager()
    if err: return err
    d = request.get_json(force=True) or {}
    return jsonify(db.add_art_to_header(d.get('art', ''), d.get('header_id')))

# ── IE Assignments ────────────────────────────────────────────────────────────

@app.route('/api/ie/<int:header_id>/assignments', methods=['GET'])
def ie_get_assignments(header_id):
    return jsonify({'ok': True, 'assignments': db.get_ie_assignments(header_id=header_id)})

@app.route('/api/ie/<int:header_id>/assign', methods=['POST'])
def ie_assign(header_id):
    err = _require_manager()
    if err: return err
    d = request.get_json(force=True) or {}
    return jsonify(db.set_ie_assignment(header_id, d.get('user_id')))

@app.route('/api/ie/<int:header_id>/unassign', methods=['POST'])
def ie_unassign(header_id):
    err = _require_manager()
    if err: return err
    d = request.get_json(force=True) or {}
    return jsonify(db.remove_ie_assignment(header_id, d.get('user_id')))

# ── IE Review Workflow ────────────────────────────────────────────────────────

@app.route('/api/ie/review/submit', methods=['POST'])
def ie_review_submit():
    u = _auth_user()
    if not u:
        return jsonify({'ok': False, 'error': '請先登入'}), 401
    d = request.get_json(force=True) or {}
    return jsonify(db.submit_ie_review(
        d.get('header_id'), d.get('stage_id'), u['username']
    ))

@app.route('/api/ie/review/list', methods=['GET'])
def ie_review_list():
    u = _auth_user()
    if not u:
        return jsonify({'ok': False, 'error': '請先登入'}), 401
    status = request.args.get('status')
    header_id = request.args.get('header_id')
    reviews = db.get_reviews(
        header_id=int(header_id) if header_id else None,
        status=status or None
    )
    return jsonify({'ok': True, 'reviews': reviews})

@app.route('/api/ie/review/<int:review_id>/approve', methods=['POST'])
def ie_review_approve(review_id):
    err = _require_manager()
    if err: return err
    u = _auth_user()
    return jsonify(db.approve_review(review_id, u['username']))

@app.route('/api/ie/review/<int:review_id>/reject', methods=['POST'])
def ie_review_reject(review_id):
    err = _require_manager()
    if err: return err
    u = _auth_user()
    d = request.get_json(force=True) or {}
    return jsonify(db.reject_review(review_id, u['username'], d.get('reason', '')))

# ── IE Stage Approval ─────────────────────────────────────────────────────────

@app.route('/api/ie/stages/<int:header_id>/<int:stage_id>/approve', methods=['POST'])
def ie_stage_approve(header_id, stage_id):
    err = _require_manager()
    if err: return err
    return jsonify(db.set_stage_approved(stage_id, header_id))

# ── Health check ─────────────────────────────────────────────────────────────

@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'ok': True, 'version': '1.6'})

# ── System update ─────────────────────────────────────────────────────────────

def _repo_root():
    return os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))

@app.route('/api/system/version_status', methods=['GET'])
def system_version_status():
    err = _require_manager()
    if err: return err
    root = _repo_root()
    try:
        subprocess.run(['git', 'fetch', 'origin', 'main'], cwd=root, timeout=15,
                       capture_output=True)
        local = subprocess.run(['git', 'rev-parse', '--short', 'HEAD'],
                               cwd=root, capture_output=True, text=True).stdout.strip()
        remote = subprocess.run(['git', 'rev-parse', '--short', 'origin/main'],
                                cwd=root, capture_output=True, text=True).stdout.strip()
        return jsonify({'ok': True, 'up_to_date': local == remote,
                        'local_commit': local, 'remote_commit': remote})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/system/update', methods=['POST'])
def system_update():
    err = _require_manager()
    if err: return err
    root = _repo_root()
    try:
        dirty = subprocess.run(['git', 'status', '--porcelain'], cwd=root,
                               capture_output=True, text=True).stdout.strip()
        if dirty:
            return jsonify({'ok': False, 'error': '本地有改動，請聯絡管理員'}), 409
        pull = subprocess.run(['git', 'pull', 'origin', 'main'], cwd=root,
                              capture_output=True, text=True, timeout=60)
        if pull.returncode != 0:
            return jsonify({'ok': False, 'error': pull.stderr.strip() or pull.stdout.strip()}), 500
        new_commit = subprocess.run(['git', 'rev-parse', '--short', 'HEAD'],
                                    cwd=root, capture_output=True, text=True).stdout.strip()
        import threading
        def _restart():
            import time; time.sleep(1); os._exit(0)
        threading.Thread(target=_restart, daemon=True).start()
        return jsonify({'ok': True, 'new_commit': new_commit})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    db.init_db()
    print('Atlas Data System starting on http://0.0.0.0:5000')
    app.run(host='0.0.0.0', port=5000, debug=False)
