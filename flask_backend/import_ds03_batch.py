#!/usr/bin/env python3
"""
DS-03 OB Batch Importer.
Processes a folder of historical OB Excel files:
  1. Extracts Vietnamese→Chinese part name pairs → adds to lookup_viet_zh table
  2. Extracts header info (ART, Season, EOLR) and imports skeleton OB records

Usage:
    python import_ds03_batch.py <folder>
    python import_ds03_batch.py <folder> --dry-run
    python import_ds03_batch.py <folder> --lookup-only
    python import_ds03_batch.py <folder> --xlsx-only      # skip .xls (use when .xlsx counterpart exists)
    python import_ds03_batch.py <folder> --fresh          # clear ob_header before importing
    python import_ds03_batch.py <folder> --dry-run --lookup-only

ART/EOLR source priority:
  1. Filename: first [A-Z]{2}\\d{4,6} token → ART; '120双'→eolr=120, '60双'→eolr=60
  2. Content (SUM.C2B / SUM sheet): fallback when filename has no ART token
"""
import sys, os, glob, re as _re

# Force UTF-8 output on Windows consoles
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import database as db

# ── Filename ART/EOLR extraction ─────────────────────────────────────────────

_FN_ART_RE = _re.compile(r'[A-Z]{2}\d{4,6}')

def fn_art(filename):
    """Return first ART code found in filename, or ''."""
    m = _FN_ART_RE.search(os.path.basename(filename))
    return m.group(0) if m else ''

def fn_eolr(filename):
    """Return EOLR inferred from '120双'/'60双' prefix in filename, or None."""
    base = os.path.basename(filename)
    if '120' in base and '双' in base:
        return 120
    if '60' in base and '双' in base:
        return 60
    return None

# Files to skip: known exact-duplicate xlsx pairs (Recovered versions / stray copies).
# Keep the cleaner filename, skip these.
_SKIP_FILENAMES = {
    '120双_FW26_GHOST SPRINT W_HQ3330..xlsx',        # duplicate of HQ3330.xlsx (stray dot)
    '120双 LA TRAINER OG-KH9682-83-84- vải lưới+da lộn+da giả (Recovered).xlsx',  # Recovered dup
    '60双_SS25 TOKYO MJ_KI5323_da thật (da cá)_thêm 1 ng mài thô mg - Copy.xlsx',  # Copy dup
}

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Language detection ────────────────────────────────────────────────────────

VIET_CHARS = set(
    'ăâđêôơưắặấầẩẫẻẽẹếềểễệịỉĩọỏõốồổỗộớờởỡợụủũứừửữựỳỵỷỹ'
    'àáãèéìíòóùúýắặấầẩẫẻẽẹếềểễệịỉĩọỏõốồổỗộớờởỡợụủũứừửữựỳỵỷỹ'
)

def has_viet(text):
    return any(c in VIET_CHARS for c in str(text or '').lower())

def has_chinese(text):
    return any('一' <= c <= '鿿' for c in str(text or ''))

# ── Sheet-to-department mapping ───────────────────────────────────────────────

DEPT_PATTERNS = {
    'cutting':     ['cutting', 'cắt', 'cat '],
    'atom':        ['atom', 'tự động', 'automat'],
    'tongcai':     ['tongcai', '同材', 'tong cai', 'dong'],
    'diannao':     ['diannao', '电脑', 'dian nao', 'computer', 'needle'],
    'stitch-zhi':  ['支流', 'zhi liu', 'tributary'],
    'stitch-zhu':  ['主流', 'zhu liu', 'main flow'],
    'assembly1':   ['assembly 1', 'lắp ráp 1', 'assembly1', 'ráp 1'],
    'assembly2':   ['assembly 2', 'lắp ráp 2', 'assembly2', 'ráp 2'],
    'dacui':       ['打粗', 'da cui', 'rough'],
    'xishi':       ['水洗', 'xi shi', 'wash'],
    'tiedadi':     ['贴大底', 'tie da di', 'sole attach'],
    'zhaoshe':     ['照射', 'zhao she', 'irradiat'],
    'chenxing':    ['成型', 'chen xing', 'mold'],
}

def match_dept(sheet_name):
    name = sheet_name.lower().strip()
    # Check for Stitching generically (both zhi and zhu)
    if 'stitch' in name or 'may' in name or '针车' in name or '縫' in name:
        if any(x in name for x in ['支', 'tributary', 'sub', '1', 'zhi']):
            return 'stitch-zhi'
        return 'stitch-zhu'
    for key, patterns in DEPT_PATTERNS.items():
        if any(p.lower() in name for p in patterns):
            return key
    return None


def extract_header(wb):
    """Extract ART, EOLR, season, model from workbook using file-format-aware logic."""
    header = {}

    # ── Strategy 1: SUM.C2B sheet ──────────────────────────────────────────
    # Prefer shortest matching sheet name to avoid variant sheets like
    # "SUM.C2B 67-68" being chosen over the canonical "SUM.C2B".
    _sumcb_name = None
    _best = float('inf')
    for _s in wb.sheetnames:
        _k = _s.lower().replace(' ', '')
        if 'sum' in _k and 'c2b' in _k and len(_s) < _best:
            _sumcb_name = _s
            _best = len(_s)
    for sh_name in ([_sumcb_name] if _sumcb_name else []):
        if True:
            ws = wb[sh_name]
            for row in ws.iter_rows(max_row=25, values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    s = str(cell).strip()
                    sl = s.lower()
                    # ART: "Mã số Article: KH9662" or "Article: JI1282"
                    if 'art' not in header and ('article' in sl or 'mã số' in sl):
                        m = _re.search(r':\s*([A-Z]{2}\d{4,6})', s)
                        if m:
                            header['art'] = m.group(1)
                    # Season: "Mùa sản xuất Season : FW25"
                    if 'season' not in header and ('season' in sl or 'mùa' in sl or 'mua' in sl):
                        m = _re.search(r'((?:FW|SS|SP)\d{2})', s, _re.IGNORECASE)
                        if m:
                            header['season'] = m.group(1).upper()
                    # EOLR: "120双/H" embedded in target-output cell
                    if 'eolr' not in header:
                        m = _re.search(r'(\d{2,3})双/H', s)
                        if m:
                            header['eolr'] = int(m.group(1))
            break

    # ── Strategy 2: SUM sheet (older format with column headers) ───────────
    if 'art' not in header:
        for sh_name in wb.sheetnames:
            if sh_name.strip().upper() == 'SUM':
                ws = wb[sh_name]
                art_col = eolr_col = None
                for i, row in enumerate(ws.iter_rows(max_row=6, values_only=True)):
                    cells = [str(c).strip() if c is not None else '' for c in row]
                    cl = [c.lower() for c in cells]
                    if 'art' in cl and art_col is None:
                        art_col = cl.index('art')
                    if 'eolr' in cl and eolr_col is None:
                        eolr_col = cl.index('eolr')
                    if art_col is not None and i > 1:
                        # value row below headers
                        if art_col < len(cells) and cells[art_col]:
                            v = cells[art_col]
                            if _re.match(r'[A-Z]{2}\d{4,6}', v):
                                header['art'] = v
                        if eolr_col is not None and eolr_col < len(cells) and cells[eolr_col]:
                            try:
                                header['eolr'] = int(float(cells[eolr_col]))
                            except ValueError:
                                pass
                        if header.get('art'):
                            break
                break

    # ── Strategy 3: Model name from first Cutting sheet ────────────────────
    for sh_name in wb.sheetnames:
        if 'cutting' in sh_name.lower() or 'pha cắt' in sh_name.lower():
            ws = wb[sh_name]
            for row in ws.iter_rows(max_row=10, values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    s = str(cell).strip()
                    if 'model name' in s.lower() and ':' in s:
                        m = _re.search(r'Model Name\s*:\s*(.+)', s, _re.IGNORECASE)
                        if m and 'model' not in header:
                            header['model'] = m.group(1).strip()[:50]
                    if '鞋型名称' in s and ':' in s:
                        m = _re.search(r'鞋型名称[：:]\s*(.+)', s)
                        if m and 'model' not in header:
                            header['model'] = m.group(1).strip()[:50]
            break

    return header


def extract_mp_from_sumcb(wb):
    """
    Extract Cutting / Stitching / Assembly / Stockfitting MP (No. of Operators)
    from the SUM_C2B sheet.

    Strategy:
      Pass 1 – scan rows 1-15 for section header cells:
               "Cutting Section" / "Stitching Section" / "Assembly Section" /
               "Stockfitting Section" → record the column index of each.
               The "No.of Operators" value is one column to the right.
      Pass 2 – find the data row (row ~16) where col 1 contains "双/H";
               read the numeric value at each section's operator column.
      Fallback – if Stockfitting was not in the top header, scan the bottom
               summary table (right-side subtable) for "Stockfitting" row.

    Returns {'cutting': float|None, 'stitching': float|None,
             'assembly': float|None, 'stock': float|None}
    """
    # Prefer the shortest SUM_C2B sheet (exact "SUM.C2B" / "SUM C2B" wins over
    # variant sheets like "SUM.C2B 67-68" or "SUM.C2B (2 màu chỉ)").
    ws = None
    best_len = float('inf')
    for sh_name in wb.sheetnames:
        sh_key = sh_name.lower().replace(' ', '')
        if 'sum' in sh_key and 'c2b' in sh_key:
            if len(sh_name) < best_len:
                ws = wb[sh_name]
                best_len = len(sh_name)
    if ws is None:
        return {}

    mp = {'cutting': None, 'stitching': None, 'assembly': None, 'stock': None}

    # Pass 1: locate section header columns
    section_ops_col = {}   # dept_key → col_idx of "No.of Operators"
    for row in ws.iter_rows(max_row=15, values_only=True):
        for col_idx, val in enumerate(row):
            if val is None:
                continue
            s = str(val).strip()
            sl = s.lower()
            if 'cutting' in sl and ('section' in sl or '裁断' in s):
                section_ops_col.setdefault('cutting', col_idx + 1)
            if 'stitching' in sl and ('section' in sl or '针车' in s):
                section_ops_col.setdefault('stitching', col_idx + 1)
            if 'assembly' in sl and ('section' in sl or '成型' in s):
                section_ops_col.setdefault('assembly', col_idx + 1)
            if 'stockfitting' in sl and ('section' in sl or '大底' in s):
                section_ops_col.setdefault('stock', col_idx + 1)

    if not section_ops_col:
        return mp

    # Pass 2: find data row (col 1 contains "双/H")
    for row in ws.iter_rows(min_row=14, max_row=22, values_only=True):
        col1 = str(row[1] if len(row) > 1 and row[1] is not None else '').strip()
        if '双/H' in col1 or '双/h' in col1.lower():
            for dept, col_idx in section_ops_col.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    try:
                        mp[dept] = float(row[col_idx])
                    except (ValueError, TypeError):
                        pass
            break

    # Fallback: if Stockfitting not in top header, search bottom summary table.
    # The RIGHT subtable (starting around col 9) has actual MP values:
    #   "Stockfitting | <MP value> | PPH"
    if mp.get('stock') is None:
        found_dept_header = False
        for row in ws.iter_rows(min_row=20, max_row=40, values_only=True):
            cells = [str(c).strip() if c is not None else '' for c in row]
            for ci, cell in enumerate(cells):
                # Detect "部门" header to know we're in a subtable
                if '部门' in cell or 'department' in cell.lower():
                    found_dept_header = True
                if found_dept_header and 'stockfitting' in cell.lower():
                    # Look for numeric value in next 1-2 cols
                    for offset in (1, 2):
                        if ci + offset < len(cells) and cells[ci + offset]:
                            try:
                                v = float(cells[ci + offset])
                                if v > 0:
                                    mp['stock'] = v
                                    break
                            except (ValueError, TypeError):
                                pass
                    if mp.get('stock') is not None:
                        break
            if mp.get('stock') is not None:
                break

    return mp


def extract_viet_zh(ws):
    """Return dict of viet_lower → zh from adjacent Viet+Chinese cell pairs."""
    pairs = {}
    for row in ws.iter_rows(values_only=True):
        vals = [str(v).strip() if v is not None else '' for v in row]
        for i in range(len(vals) - 1):
            v1, v2 = vals[i], vals[i + 1]
            if len(v1) < 2 or len(v2) < 2:
                continue
            if has_viet(v1) and has_chinese(v2) and not has_chinese(v1):
                pairs[v1.lower()] = v2
            elif has_chinese(v1) and has_viet(v2) and not has_chinese(v2):
                pairs[v2.lower()] = v1
    return pairs


def process_file(path, dry_run=False, lookup_only=False):
    if not HAS_OPENPYXL:
        return {'ok': False, 'error': 'openpyxl not installed'}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return {'ok': False, 'error': f'Cannot open: {e}'}

    result = {
        'file': os.path.basename(path),
        'lookup_pairs': {},
        'header': {},
        'depts': [],
        'ok': True,
    }

    # Collect Viet-Chinese pairs from all sheets
    for ws in wb.worksheets:
        result['lookup_pairs'].update(extract_viet_zh(ws))

    # Header from workbook (multi-sheet aware)
    content_hdr = extract_header(wb)

    # ART priority: filename first, then content fallback
    art_from_file = fn_art(path)
    art = art_from_file or str(content_hdr.get('art', '')).strip()

    # EOLR priority: filename prefix (120双/60双), then content, then default 60
    eolr_from_file = fn_eolr(path)
    if eolr_from_file is not None:
        eolr = eolr_from_file
    else:
        try:
            eolr = int(float(str(content_hdr.get('eolr', 60)).strip()))
        except (ValueError, TypeError):
            eolr = 60

    hdr = dict(content_hdr)
    hdr['art']  = art
    hdr['eolr'] = eolr
    result['header'] = hdr

    # Identify department sheets
    for ws in wb.worksheets:
        dept = match_dept(ws.title)
        if dept:
            result['depts'].append({'sheet': ws.title, 'dept': dept})

    # Extract MP values from SUM_C2B
    mp_vals = extract_mp_from_sumcb(wb)
    result['mp'] = mp_vals

    if not dry_run:
        # Save lookup pairs
        if result['lookup_pairs']:
            conn = db.get_conn()
            ts = db.now_iso()
            try:
                for viet, zh in result['lookup_pairs'].items():
                    conn.execute(
                        'INSERT INTO lookup_viet_zh (viet, zh, created_at) VALUES (?,?,?) '
                        'ON CONFLICT(viet) DO UPDATE SET zh=excluded.zh',
                        (viet, zh, ts))
                conn.commit()
            finally:
                conn.close()

        # Import OB skeleton if ART resolved and not lookup-only
        if not lookup_only and art:
            ob_data = {
                'header': {
                    'art': art, 'season': hdr.get('season', ''),
                    'model': hdr.get('model', ''), 'material': hdr.get('material', ''),
                    'category': hdr.get('category', ''), 'eolr': eolr, 'run': 1
                },
                'epph': {
                    'cutting':    mp_vals.get('cutting')    or 0,
                    'stitching':  mp_vals.get('stitching')  or 0,
                    'assembly':   mp_vals.get('assembly')   or 0,
                    'stock':      mp_vals.get('stock')      or 0,
                },
                'sheets': {}
            }
            result['ob_save'] = db.save_ob_record(ob_data)

    return result


def batch_process(folder, dry_run=False, lookup_only=False, xlsx_only=False, fresh=False):
    files = glob.glob(os.path.join(folder, '**', '*.xlsx'), recursive=True)
    if not xlsx_only:
        files += glob.glob(os.path.join(folder, '**', '*.xls'), recursive=True)

    # Strip temp/lock files and known duplicates
    files = [
        f for f in files
        if not os.path.basename(f).startswith('~$')
        and os.path.basename(f) not in _SKIP_FILENAMES
    ]
    files.sort()

    if not files:
        print(f'No Excel files found in: {folder}')
        return

    print(f'Found {len(files)} Excel file(s)  (xlsx_only={xlsx_only})')
    if dry_run:
        print('DRY RUN — no data written to DB\n')

    # Fresh mode: wipe existing OB data before re-importing
    if fresh and not dry_run:
        conn = db.get_conn()
        try:
            conn.execute('DELETE FROM ob_rows')
            conn.execute('DELETE FROM ob_epph')
            conn.execute('DELETE FROM ob_header')
            conn.commit()
            print('FRESH: cleared ob_header, ob_epph, ob_rows\n')
        finally:
            conn.close()

    all_pairs = {}
    ok_count = 0
    err_count = 0
    new_count = 0
    upd_count = 0

    for path in files:
        rel = os.path.relpath(path, folder)
        print(f'Processing: {rel}')
        r = process_file(path, dry_run=dry_run, lookup_only=lookup_only)

        if not r['ok']:
            print(f'  ERROR: {r["error"]}')
            err_count += 1
            continue

        ok_count += 1
        pairs = r.get('lookup_pairs', {})
        all_pairs.update(pairs)

        if pairs:
            print(f'  Lookup pairs: {len(pairs)}')
            for v, z in list(pairs.items())[:3]:
                print(f'    {v} → {z}')
            if len(pairs) > 3:
                print(f'    ... +{len(pairs) - 3} more')

        h = r.get('header', {})
        mp = r.get('mp', {})
        if h:
            art_disp = h.get('art', '?')
            mp_str = (f'  C={mp.get("cutting","?")} S={mp.get("stitching","?")}'
                      f' A={mp.get("assembly","?")} St={mp.get("stock","?")}')
            print(f'  Header: ART={art_disp} | Season={h.get("season","?")} | EOLR={h.get("eolr","?")} | MP:{mp_str}')

        depts = r.get('depts', [])
        if depts:
            print(f'  Dept sheets: {", ".join(d["dept"] for d in depts)}')

        ob = r.get('ob_save', {})
        if ob:
            if ob.get('ok'):
                if ob.get('new'):
                    new_count += 1
                else:
                    upd_count += 1
            else:
                print(f'  OB save warning: {ob.get("error")}')

    print(f'\n{"=" * 50}')
    print(f'Done: {ok_count} files OK, {err_count} errors')
    if not lookup_only and not dry_run:
        print(f'OB records: {new_count} new, {upd_count} updated')
    print(f'Total unique lookup pairs: {len(all_pairs)}')
    if dry_run:
        print('(DRY RUN — nothing saved)')


if __name__ == '__main__':
    if len(sys.argv) < 2 or sys.argv[1] in ('-h', '--help'):
        print(__doc__)
        sys.exit(0)

    folder = sys.argv[1]
    if not os.path.isdir(folder):
        print(f'Not a directory: {folder}')
        sys.exit(1)

    dry_run     = '--dry-run'     in sys.argv
    lookup_only = '--lookup-only' in sys.argv
    xlsx_only   = '--xlsx-only'   in sys.argv
    fresh       = '--fresh'       in sys.argv

    if not dry_run:
        db.init_db()

    batch_process(folder, dry_run=dry_run, lookup_only=lookup_only,
                  xlsx_only=xlsx_only, fresh=fresh)
