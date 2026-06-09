#!/usr/bin/env python3
"""
結果表自動生成 + 比對腳本 (Result Table Generator & Comparator)

輸入：
  - DS-04 進度表 Excel 路徑
  - DS-05 大底課進度表 Excel 路徑（OCS 大底課）
  - 廠務組織編制表 Excel 路徑（比對目標）
  - 月份（用於選擇編制表 sheet，格式 "6.2026"）

輸出：
  1. CSA 結果表：LEAN | 鞋型(Model+ART) | 訂單 | 裁斷MP | 針車MP | 成型MP | 協理給 | 合計
     - 來源：DS-04 訂單 + DS-02 Model Name + DS-03 ob_epph MP
     - 所有加1~加12，所有組別
  2. OCS 大底課：T群組 | AD代碼 | 鞋型 | 訂單
     - 來源：DS-05 大底課進度表
  3. 與廠務組織編制表 sheet 逐行比對，輸出差異清單

Usage:
    python analyze_result_table.py
        --schedule <DS04.xlsx>
        --bianche <廠務組織編制表.xlsx>
        --month 6.2026
        [--ds05 <大底課進度表.xlsx>]
        [--eolr 120]
        [--output flask_backend/test_output/result_full.txt]
"""
import sys, os, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import database as db

_ART_RE  = re.compile(r'[A-Z]{2}\d{4,6}')
_QTY_RE  = re.compile(r'--(\d+)\s*\(')


# ── DS-04 schedule: extract all groups from all sheets ───────────────────────

def _parse_order_cell(val):
    s = str(val or '').strip()
    arts = _ART_RE.findall(s)
    m = _QTY_RE.search(s)
    qty = int(m.group(1)) if m else 0
    if not arts:
        return []
    per = qty // len(arts)
    rem = qty % len(arts)
    return [(a, per + (rem if i == 0 else 0)) for i, a in enumerate(arts)]


def _is_order_cell(val):
    s = str(val or '').strip()
    return bool(_ART_RE.search(s) and '--' in s)


def extract_all_groups(wb):
    """
    Scan all sheets in a DS-04 schedule workbook.
    Returns list of {'sheet': str, 'group': str, 'orders': {art: qty}}.
    """
    results = []
    for ws in wb.worksheets:
        # Skip non-production sheets
        title = ws.title.strip()
        if not re.search(r'\d+\s*部', title) and title not in ('1部 ', '2部 ', '3部'):
            if not any(c.isdigit() for c in title):
                continue

        # Collect all order cells → orders by ART
        orders = {}
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if not _is_order_cell(cell):
                    continue
                for art, qty in _parse_order_cell(cell):
                    orders[art] = orders.get(art, 0) + qty

        if orders:
            results.append({
                'sheet': title,
                'orders': orders,
            })
    return results


# ── DS-03 ob_epph MP lookup ───────────────────────────────────────────────────

def get_mp(art, eolr=120):
    """Return (cutting, stitching, assembly) from ob_epph for ART+EOLR."""
    try:
        conn = db.get_conn()
        row = conn.execute(
            '''SELECT e.cutting, e.stitching, e.assembly
               FROM ob_epph e
               JOIN ob_header h ON h.id = e.header_id
               JOIN ob_articles a ON a.header_id = h.id
               WHERE a.art = ? AND h.eolr = ?
               ORDER BY h.id DESC LIMIT 1''',
            (art, eolr)
        ).fetchone()
        conn.close()
        if row and any(v for v in row):
            return row[0] or 0, row[1] or 0, row[2] or 0
        return None
    except Exception:
        return None


# ── DS-02 model name lookup ───────────────────────────────────────────────────

def get_model_name(art):
    try:
        conn = db.get_conn()
        row = conn.execute(
            'SELECT model_name FROM ds02_fob WHERE art = ? LIMIT 1', (art,)
        ).fetchone()
        conn.close()
        return (row[0] or '').strip() if row else ''
    except Exception:
        return ''


# ── 廠務組織編制表 reader ─────────────────────────────────────────────────────

def read_bianche_csa(wb, sheet_name):
    """
    Read CSA rows from a 廠務組織編制表 sheet.
    Returns list of {lean, model, order, cut, stitch, asm, assist, total, headcount}.
    """
    ws = None
    for sh in wb.sheetnames:
        if sh.strip() == sheet_name.strip():
            ws = wb[sh]
            break
    if ws is None:
        # Try partial match
        for sh in wb.sheetnames:
            if sheet_name in sh:
                ws = wb[sh]
                break
    if ws is None:
        return []

    rows = []
    current_lean = ''

    for rn, row in enumerate(ws.iter_rows(values_only=True), 1):
        col_a = str(row[0] or '').strip() if row else ''
        col_b = str(row[1] or '').strip() if len(row) > 1 else ''

        # LEAN group header (1A, 1B, 2A, etc.)
        if re.match(r'^\d+[A-Z]$', col_a):
            current_lean = col_a

        # Data row: col A empty/dot, col B has model name
        if (not col_a or col_a == '.') and col_b and len(col_b) > 3:
            def _val(idx):
                v = row[idx] if len(row) > idx else None
                if v is None or str(v).strip() in ('.', ''):
                    return None
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return None

            order  = _val(6)   # col G
            cut    = _val(7)   # col H
            stitch = _val(8)   # col I
            asm    = _val(9)   # col J
            assist = _val(10)  # col K
            total  = _val(11)  # col L

            if current_lean and (cut is not None or order is not None):
                rows.append({
                    'lean':     current_lean,
                    'model':    col_b[:60],
                    'order':    order,
                    'cut':      cut,
                    'stitch':   stitch,
                    'asm':      asm,
                    'assist':   assist or 0,
                    'total':    total,
                })

    return rows


# ── CSA result table generator ────────────────────────────────────────────────

def build_csa_rows(schedule_groups, eolr=120):
    """
    For each (sheet, orders), join DS-02 model name + DS-03 ob_epph MP.
    Returns list of result rows.
    """
    rows = []
    for group_info in schedule_groups:
        sheet = group_info['sheet']
        orders = group_info['orders']

        for art in sorted(orders):
            qty = orders[art]
            model_name = get_model_name(art)
            mp = get_mp(art, eolr)

            cut = round(mp[0], 1) if mp else None
            stitch = round(mp[1], 1) if mp else None
            asm = round(mp[2], 1) if mp else None
            total = round(cut + stitch + asm, 1) if mp else None

            rows.append({
                'sheet':     sheet,
                'art':       art,
                'model':     model_name,
                'qty':       qty,
                'cut':       cut,
                'stitch':    stitch,
                'asm':       asm,
                'assist':    0,
                'total':     total,
                'mp_ok':     mp is not None,
            })

    return rows


# ── OCS 大底課 ────────────────────────────────────────────────────────────────

def build_ocs_rows(ds05_path):
    """Parse DS-05 大底課進度表 and return T-group rows."""
    if not ds05_path or not os.path.exists(ds05_path):
        return [], f'DS-05 file not found: {ds05_path}'
    try:
        from analyze_ds05 import analyze as ds05_analyze
        result = ds05_analyze(ds05_path)
        if not result.get('ok'):
            return [], result.get('error', 'DS-05 error')
        rows = []
        for g in result.get('groups', []):
            for m in g.get('models', []):
                rows.append({
                    'group':    g['group_name'],
                    'ad_code':  m['ad_code'],
                    'model':    m['model_name'],
                    'qty':      m['total_qty'],
                })
        return rows, None
    except Exception as e:
        return [], str(e)


# ── Comparison ────────────────────────────────────────────────────────────────

def compare_csa(our_rows, ref_rows):
    """
    Compare our calculated rows against reference 廠務組織編制表 rows.
    Returns list of diff records.
    """
    diffs = []

    # Build reference lookup: model_fragment → ref_row (best-effort)
    ref_lookup = {}
    for r in ref_rows:
        key = r['lean'] + '|' + r['model'][:20].upper()
        ref_lookup[key] = r

    for r in our_rows:
        # Try to match by ART in model string
        matched = None
        for k, ref in ref_lookup.items():
            if r['art'] in ref['model'] or (r['model'] and r['model'][:15] in ref['model']):
                matched = ref
                break

        if matched:
            cut_diff   = (r['cut'] or 0) - (matched['cut'] or 0)
            stitch_diff = (r['stitch'] or 0) - (matched['stitch'] or 0)
            asm_diff   = (r['asm'] or 0) - (matched['asm'] or 0)
            if abs(cut_diff) > 0.5 or abs(stitch_diff) > 0.5 or abs(asm_diff) > 0.5:
                diffs.append({
                    'art': r['art'],
                    'model': r['model'][:30],
                    'lean': matched['lean'],
                    'our_cut': r['cut'], 'ref_cut': matched['cut'], 'diff_cut': round(cut_diff, 1),
                    'our_stitch': r['stitch'], 'ref_stitch': matched['stitch'], 'diff_stitch': round(stitch_diff, 1),
                    'our_asm': r['asm'], 'ref_asm': matched['asm'], 'diff_asm': round(asm_diff, 1),
                })
        elif r['mp_ok']:
            diffs.append({
                'art': r['art'],
                'model': r['model'][:30],
                'lean': '?',
                'our_cut': r['cut'], 'ref_cut': None, 'diff_cut': None,
                'our_stitch': r['stitch'], 'ref_stitch': None, 'diff_stitch': None,
                'our_asm': r['asm'], 'ref_asm': None, 'diff_asm': None,
            })

    return diffs


# ── Main ──────────────────────────────────────────────────────────────────────

def analyze(schedule_path, bianche_path, month_sheet, ds05_path=None, eolr=120):
    """
    Full result table generation and comparison.
    Returns dict with csa_rows, ocs_rows, diffs, errors.
    """
    if not HAS_OPENPYXL:
        return {'ok': False, 'error': 'openpyxl not installed'}

    errors = []

    # ── DS-04 schedule ─────────────────────────────────────────────────────
    if not os.path.exists(schedule_path):
        return {'ok': False, 'error': f'Schedule file not found: {schedule_path}'}
    try:
        wb04 = openpyxl.load_workbook(schedule_path, data_only=True, read_only=True)
        schedule_groups = extract_all_groups(wb04)
        wb04.close()
    except Exception as e:
        return {'ok': False, 'error': f'Cannot open schedule: {e}'}

    # ── CSA rows ───────────────────────────────────────────────────────────
    csa_rows = build_csa_rows(schedule_groups, eolr)

    # ── OCS 大底課 ─────────────────────────────────────────────────────────
    ocs_rows = []
    if ds05_path:
        ocs_rows, ocs_err = build_ocs_rows(ds05_path)
        if ocs_err:
            errors.append(f'OCS: {ocs_err}')

    # ── Reference comparison ───────────────────────────────────────────────
    ref_csa = []
    diffs = []
    if bianche_path and os.path.exists(bianche_path):
        try:
            wb_ref = openpyxl.load_workbook(bianche_path, data_only=True)
            ref_csa = read_bianche_csa(wb_ref, month_sheet)
            wb_ref.close()
            diffs = compare_csa(csa_rows, ref_csa)
        except Exception as e:
            errors.append(f'Reference read error: {e}')
    else:
        errors.append(f'Reference file not found: {bianche_path}')

    return {
        'ok': True,
        'eolr': eolr,
        'schedule_sheets': len(schedule_groups),
        'total_arts': sum(len(g['orders']) for g in schedule_groups),
        'csa_rows': len(csa_rows),
        'ocs_rows': len(ocs_rows),
        'ref_rows': len(ref_csa),
        'diff_count': len(diffs),
        'csa': csa_rows,
        'ocs': ocs_rows,
        'ref': ref_csa,
        'diffs': diffs,
        'errors': errors,
    }


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import argparse

    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')

    parser = argparse.ArgumentParser(description='結果表自動生成 + 比對')
    parser.add_argument('--schedule',  required=True, help='DS-04 進度表 Excel')
    parser.add_argument('--bianche',   required=True, help='廠務組織編制表 Excel')
    parser.add_argument('--month',     default='6.2026', help='Sheet name (e.g. 6.2026)')
    parser.add_argument('--ds05',      default='', help='DS-05 大底課進度表 Excel')
    parser.add_argument('--eolr',      type=int, default=120)
    parser.add_argument('--output',    default='flask_backend/test_output/result_full.txt')
    args = parser.parse_args()

    db.init_db()

    result = analyze(
        args.schedule,
        args.bianche,
        args.month,
        args.ds05 or None,
        args.eolr,
    )

    os.makedirs(os.path.dirname(args.output), exist_ok=True)

    lines = []
    lines.append(f'=== 結果表生成報告 ===')
    lines.append(f'進度表:   {args.schedule}')
    lines.append(f'編制表:   {args.bianche}  Sheet: {args.month}')
    lines.append(f'EOLR:     {args.eolr}')
    lines.append(f'DS-04 sheets: {result["schedule_sheets"]}  ARTs: {result["total_arts"]}')
    lines.append(f'CSA rows: {result["csa_rows"]}')
    lines.append(f'OCS rows: {result["ocs_rows"]}')
    lines.append(f'Ref rows: {result["ref_rows"]}')
    lines.append(f'差異數:   {result["diff_count"]}')
    if result['errors']:
        lines.append('')
        lines.append('=== Errors ===')
        for e in result['errors']:
            lines.append(f'  {e}')

    # ── CSA table ──────────────────────────────────────────────────────
    lines.append('')
    lines.append('=== CSA 結果表（DS-03 ob_epph MP） ===')
    hdr = f"{'Sheet':8}  {'ART':12}  {'Model':30}  {'Qty':>7}  {'Cut':>6}  {'Stitch':>6}  {'Asm':>6}  {'Total':>6}  {'MP?':4}"
    lines.append(hdr)
    lines.append('-' * len(hdr))
    for r in result['csa']:
        cut    = f"{r['cut']:.1f}"    if r['cut']    is not None else '-'
        stitch = f"{r['stitch']:.1f}" if r['stitch'] is not None else '-'
        asm    = f"{r['asm']:.1f}"    if r['asm']    is not None else '-'
        total  = f"{r['total']:.1f}"  if r['total']  is not None else '-'
        mp_ok  = '✓' if r['mp_ok'] else '✗'
        lines.append(
            f"{r['sheet'][:8]:8}  {r['art']:12}  {r['model'][:30]:30}  "
            f"{r['qty']:>7}  {cut:>6}  {stitch:>6}  {asm:>6}  {total:>6}  {mp_ok:4}"
        )

    # ── OCS 大底課 ─────────────────────────────────────────────────────
    if result['ocs']:
        lines.append('')
        lines.append('=== OCS 大底課（DS-05） ===')
        for r in result['ocs']:
            lines.append(f"  T群組={r['group']:12}  AD={r['ad_code']:12}  Qty={r['qty']:>7}  {r['model'][:30]}")

    # ── Comparison diffs ───────────────────────────────────────────────
    lines.append('')
    lines.append('=== 差異清單（vs 廠務組織編制表） ===')
    if not result['diffs']:
        lines.append('  (no significant differences found or no reference data)')
    else:
        diff_hdr = f"{'LEAN':4}  {'ART':12}  {'Model':28}  {'OurCut':>7}  {'RefCut':>7}  {'ΔCut':>6}  {'OurS':>6}  {'RefS':>6}  {'ΔS':>5}  {'OurA':>6}  {'RefA':>6}  {'ΔA':>5}"
        lines.append(diff_hdr)
        lines.append('-' * len(diff_hdr))
        for d in result['diffs']:
            oc = f"{d['our_cut']:.1f}"  if d['our_cut']  is not None else '-'
            rc = f"{d['ref_cut']:.1f}"  if d['ref_cut']  is not None else '(new)'
            dc = f"{d['diff_cut']:+.1f}" if d['diff_cut'] is not None else '-'
            os_ = f"{d['our_stitch']:.1f}" if d['our_stitch'] is not None else '-'
            rs = f"{d['ref_stitch']:.1f}" if d['ref_stitch'] is not None else '(new)'
            ds = f"{d['diff_stitch']:+.1f}" if d['diff_stitch'] is not None else '-'
            oa = f"{d['our_asm']:.1f}"  if d['our_asm']  is not None else '-'
            ra = f"{d['ref_asm']:.1f}"  if d['ref_asm']  is not None else '(new)'
            da = f"{d['diff_asm']:+.1f}" if d['diff_asm'] is not None else '-'
            lines.append(
                f"{d['lean']:4}  {d['art']:12}  {d['model']:28}  "
                f"{oc:>7}  {rc:>7}  {dc:>6}  "
                f"{os_:>6}  {rs:>6}  {ds:>5}  "
                f"{oa:>6}  {ra:>6}  {da:>5}"
            )

    with open(args.output, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f'Output written to: {args.output}')
    print(f'CSA rows: {result["csa_rows"]}  OCS rows: {result["ocs_rows"]}  Diffs: {result["diff_count"]}')
    if result['errors']:
        for e in result['errors']:
            print(f'  WARNING: {e}')
