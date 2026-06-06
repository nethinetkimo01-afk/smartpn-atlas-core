#!/usr/bin/env python3
"""
Build result_table_v1.xlsx
- CSA: all 加1~加12 groups, ART/Model/Order from DS-04, MP blank (red)
- OCS: T-groups from DS-05 if available
- Fixed header rows from 廠務組織編制表 6.2026 sheet
"""
import sys, os, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter
import database as db

SCHEDULE  = r'C:\Users\user\OneDrive\Desktop\Biên chế\Jun\2026年6月份正式进度表 5 30.xlsx'
BIANCHE   = r'C:\Users\user\OneDrive\Desktop\Biên chế\Jun\2026年6月份廠務组织編制 20260524.xlsx'
DS05_FILE = r'C:\Users\user\OneDrive\Desktop\Biên chế\Jun\2026年6月份正式贴底进度表进度表 5.16. (ĐẾ).xlsx'
MONTH_SH  = '6.2026'
EOLR      = 120
OUTPUT    = r'D:\smartpn-atlas-core\flask_backend\test_output\result_table_v1.xlsx'

_ART_RE = re.compile(r'[A-Z]{2}\d{4,6}')
_QTY_RE = re.compile(r'--(\d+)\s*\(')

# LEAN group title parsing: 加一A组 → 1A, 加十一A1组 → 11A1
_CN_NUM = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5,
           '六': 6, '七': 7, '八': 8, '九': 9}
_LEAN_TITLE_RE = re.compile(r'^加(十[一二]?|[一二三四五六七八九])([A-Z])(\d*)组$')

def _parse_lean_title(s):
    """'加十一A1组' → '11A1', '加一A组' → '1A', else None."""
    m = _LEAN_TITLE_RE.match(str(s or '').strip())
    if not m:
        return None
    cn, letter, digit = m.group(1), m.group(2), m.group(3)
    if cn == '十':
        n = 10
    elif cn.startswith('十'):
        n = 10 + _CN_NUM.get(cn[1:], 0)
    else:
        n = _CN_NUM.get(cn, 0)
    return f'{n}{letter}{digit}'

# ── Styles ────────────────────────────────────────────────────────────────────
RED_FILL   = PatternFill("solid", fgColor="FFCCCC")
GRAY_FILL  = PatternFill("solid", fgColor="D9D9D9")
BLUE_FILL  = PatternFill("solid", fgColor="BDD7EE")
LEAN_FILL  = PatternFill("solid", fgColor="FFF2CC")
HDR_FILL   = PatternFill("solid", fgColor="4472C4")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
BOLD       = Font(bold=True, size=10)
NORM       = Font(size=10)
RED_FONT   = Font(color="FF0000", size=10, italic=True)
CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT       = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT      = Alignment(horizontal="right", vertical="center")
THIN       = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _is_order_cell(val):
    s = str(val or '').strip()
    return bool(_ART_RE.search(s) and '--' in s)


def _parse_order_cell(val):
    s = str(val or '').strip()
    all_arts = _ART_RE.findall(s)
    # Sum ALL numbers between '--' and '(' to support multi-lot cells:
    #   MF2606KH8402-01-02--56-36(6/13)  →  56+36=92
    #   MF2604KJ8322-03--900(5/29)       →  900
    m = re.search(r'--(.+?)\(', s)
    qty = sum(int(x) for x in re.findall(r'\d+', m.group(1))) if m else 0
    # Only real (non-MF) ARTs receive the qty; MF codes are manufacturing order
    # numbers embedded in the cell, not separate product ARTs.
    arts = [a for a in all_arts if not a.startswith('MF')]
    if not arts:
        return []
    per = qty // len(arts)
    rem = qty % len(arts)
    return [(a, per + (rem if i == 0 else 0)) for i, a in enumerate(arts)]


def get_model(art):
    try:
        conn = db.get_conn()
        row = conn.execute('SELECT model_name FROM ds02_fob WHERE art=? LIMIT 1', (art,)).fetchone()
        conn.close()
        return (row[0] or '').strip() if row else ''
    except Exception:
        return ''


# Section keywords for DS-04 sub-section filtering
_SECTION_CHENGXING = frozenset({'成型进度', '成型進度'})
_SECTION_SKIP      = frozenset({'外包鞋面', '针车进度', '针車進度'})
_ALL_SECTIONS      = _SECTION_CHENGXING | _SECTION_SKIP
# MF order number pattern for deduplication: MF2606JQ0597-01
_MF_ORDER_RE = re.compile(r'(MF\d{4}[A-Z]{2}\d{4,6}-\d+)')


# ── DS-04: extract all sheets with their orders ───────────────────────────────
def load_schedule():
    """Returns [{sheet, rows:[{lean,art,qty}], orders:{art:qty}, art_lean:{art:lean}}]

    Section-aware parsing rules (2026-06-06):
    - LEAN groups with a 成型进度 sub-section: ONLY take orders from that section.
      Main body (before section header) and 针车进度 are skipped.
    - 外包鞋面 and 针车进度 sections are always skipped.
    - MF order numbers (MFyyyyARTCODE-lot) are deduplicated within each LEAN to
      prevent double-counting from repeated column sets in the same sheet.

    Each sheet is independent — no cross-sheet aggregation.
    'orders' and 'art_lean' are kept for backwards-compat.
    """
    # Use data_only=True without read_only so we can iterate rows_data twice
    wb = openpyxl.load_workbook(SCHEDULE, data_only=True)
    results = []

    for ws in wb.worksheets:
        title = ws.title.strip()
        rows_data = list(ws.iter_rows(values_only=True))

        # Pass 1: find which LEAN groups have a 成型进度 sub-section
        lean_has_cx = set()
        tmp_lean = ''
        for row in rows_data:
            col1 = str(row[0] or '').strip() if row else ''
            col2 = str(row[1] or '').strip() if len(row) > 1 else ''
            lp = _parse_lean_title(col1)
            if lp:
                tmp_lean = lp
            if col2 in _SECTION_CHENGXING and tmp_lean:
                lean_has_cx.add(tmp_lean)

        # Pass 2: collect orders with section filtering + MF-order deduplication
        lean_art_qty = {}   # (lean, art) -> qty
        seen_mf = set()     # (lean, mf_order_num) — e.g. ('7B', 'MF2606JQ0597-01')
        current_lean = ''
        current_section = ''

        for row in rows_data:
            col1 = str(row[0] or '').strip() if row else ''
            col2 = str(row[1] or '').strip() if len(row) > 1 else ''

            lp = _parse_lean_title(col1)
            if lp:
                current_lean = lp
                current_section = ''
                continue

            if col2 in _ALL_SECTIONS:
                current_section = col2
                continue

            # Skip unwanted sections
            if current_section in _SECTION_SKIP:
                continue
            # If this LEAN has 成型进度, skip its main body (not in any section yet)
            if current_lean in lean_has_cx and current_section not in _SECTION_CHENGXING:
                continue

            for cell in row:
                if not _is_order_cell(cell):
                    continue
                cell_str = str(cell or '').strip()

                # Deduplicate by full MF order number to avoid repeated column sets
                mf_m = _MF_ORDER_RE.match(cell_str)
                if mf_m:
                    mf_key = (current_lean, mf_m.group(1))
                    if mf_key in seen_mf:
                        continue
                    seen_mf.add(mf_key)

                for art, qty in _parse_order_cell(cell):
                    key = (current_lean, art)
                    lean_art_qty[key] = lean_art_qty.get(key, 0) + qty

        if not lean_art_qty:
            continue

        rows = [{'lean': lean, 'art': art, 'qty': qty}
                for (lean, art), qty in lean_art_qty.items()]

        # Backwards-compat: per-art totals and first-seen lean
        orders = {}
        art_lean = {}
        for r in rows:
            orders[r['art']] = orders.get(r['art'], 0) + r['qty']
            if r['art'] not in art_lean:
                art_lean[r['art']] = r['lean']

        results.append({'sheet': title, 'rows': rows, 'orders': orders, 'art_lean': art_lean})
    wb.close()
    return results


# ── DS-05: load T-groups ──────────────────────────────────────────────────────
def load_ocs():
    if not os.path.exists(DS05_FILE):
        return [], f'DS-05 not found: {DS05_FILE}'
    try:
        from analyze_ds05 import analyze as ds05_analyze
        result = ds05_analyze(DS05_FILE)
        if not result.get('ok'):
            return [], result.get('error', 'DS-05 error')
        rows = []
        for g in result.get('groups', []):
            for m in g.get('models', []):
                arts = sorted({o['art'] for o in m.get('orders', [])})
                rows.append({
                    'group':      g['group_name'],
                    'headcount':  g.get('headcount', {}),
                    'ad_code':    m['ad_code'],
                    'model_name': m['model_name'],
                    'arts':       arts,
                    'total_qty':  m['total_qty'],
                })
        return rows, None
    except Exception as e:
        return [], str(e)


# ── 廠務組織編制表: read LEAN group structure ──────────────────────────────────
def load_bianche_structure():
    """
    Returns list of rows from 廠務組織編制表 6.2026 sheet.
    Each row: {lean, model, arts_in_model, order, cut, stitch, asm, assist, total}
    """
    wb = openpyxl.load_workbook(BIANCHE, data_only=True)
    ws = None
    for sh in wb.sheetnames:
        if sh.strip() == MONTH_SH:
            ws = wb[sh]
            break
    if ws is None:
        wb.close()
        return []

    rows = []
    current_lean = ''
    for row in ws.iter_rows(values_only=True):
        col_a = str(row[0] or '').strip() if row else ''
        col_b = str(row[1] or '').strip() if len(row) > 1 else ''
        if re.match(r'^\d+[A-Z]\d*$', col_a):
            current_lean = col_a
        if (not col_a or col_a == '.') and col_b and len(col_b) > 3:
            def _v(idx):
                v = row[idx] if len(row) > idx else None
                if v is None or str(v).strip() in ('.', ''):
                    return None
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return None
            arts_found = _ART_RE.findall(col_b)
            rows.append({
                'lean':   current_lean,
                'model':  col_b,
                'arts':   arts_found,
                'order':  _v(6),
                'cut':    _v(7),
                'stitch': _v(8),
                'asm':    _v(9),
                'assist': _v(10),
                'total':  _v(11),
            })
    wb.close()
    return rows


# ── 廠務組織編制表: read OCS fixed unit sections ──────────────────────────────
OCS_SECTIONS = ['組底配套', '自動化', '電腦針車', '印刷', '設備工程']

def load_ocs_fixed_units():
    """
    Scan 廠務組織編制表 6.2026 for OCS fixed-unit sections.
    Returns dict: {section_name: [{unit, headcount}]}
    Section header row: col1 = section name (may also have first sub-unit in col2).
    Sub-unit rows: col1 empty, col2 = unit name, headcount in col13 or col11.
    設備工程 appears AFTER RB/QC sections — must scan full sheet.
    """
    wb = openpyxl.load_workbook(BIANCHE, data_only=True)
    ws = None
    for sh in wb.sheetnames:
        if sh.strip() == MONTH_SH:
            ws = wb[sh]
            break
    if ws is None:
        wb.close()
        return {}

    def _headcount(row):
        for idx in [12, 10, 14]:  # col13, col11, col15
            v = row[idx] if len(row) > idx else None
            if v is not None and str(v).strip() not in ('', '.', '編制'):
                try:
                    return int(float(v))
                except (ValueError, TypeError):
                    pass
        return None

    result = {s: [] for s in OCS_SECTIONS}
    current_section = None

    for row in ws.iter_rows(values_only=True):
        col1 = str(row[0] or '').strip() if row else ''
        col2 = str(row[1] or '').strip() if len(row) > 1 else ''

        if col1 in OCS_SECTIONS:
            current_section = col1
            if col2 and col2 not in ('編制', ''):
                result[current_section].append({'unit': col2, 'headcount': _headcount(row)})
        elif current_section and not col1 and col2 and col2 not in ('編制', ''):
            result[current_section].append({'unit': col2, 'headcount': _headcount(row)})
        elif col1 and col1 not in OCS_SECTIONS:
            current_section = None  # exit OCS section on any other non-empty col1

    wb.close()
    return result


# ── 廠務組織編制表: read RB and QC sections ──────────────────────────────────────
RB_QC_SECTIONS = ['RB', 'QC']

def load_rb_qc_units():
    """
    Scan 廠務組織編制表 6.2026 for RB and QC sections.
    Returns dict: {'RB': [{unit, last_month, this_month}], 'QC': [...]}
    col13 (idx 12) = 上月人數 reference
    col15 (idx 14) = 本月人數 (editable; defaults to col13 if None)
    """
    wb = openpyxl.load_workbook(BIANCHE, data_only=True)
    ws = None
    for sh in wb.sheetnames:
        if sh.strip() == MONTH_SH:
            ws = wb[sh]
            break
    if ws is None:
        wb.close()
        return {s: [] for s in RB_QC_SECTIONS}

    def _val(row, idx):
        v = row[idx] if len(row) > idx else None
        if v is None or str(v).strip() in ('', '.', '編制'): return None
        try: return int(float(v))
        except: return None

    result = {s: [] for s in RB_QC_SECTIONS}
    current_section = None
    STOP_AFTER = {'設備工程'}

    for row in ws.iter_rows(values_only=True):
        col1 = str(row[0] or '').strip() if row else ''
        col2 = str(row[1] or '').strip() if len(row) > 1 else ''

        if col1 in RB_QC_SECTIONS:
            current_section = col1
            if col2 and col2 not in ('編制', ''):
                last_m = _val(row, 12)
                this_m = _val(row, 14) if _val(row, 14) is not None else last_m
                result[current_section].append({
                    'unit': col2, 'last_month': last_m, 'this_month': this_m
                })
            continue

        if col1 in STOP_AFTER:
            current_section = None
            continue

        if current_section:
            if not col1 and col2 and col2 not in ('編制', ''):
                last_m = _val(row, 12)
                this_m = _val(row, 14) if _val(row, 14) is not None else last_m
                result[current_section].append({
                    'unit': col2, 'last_month': last_m, 'this_month': this_m
                })
            elif col1 and col1 not in RB_QC_SECTIONS:
                current_section = None

    wb.close()
    return result


# ── Build RB or QC tab ────────────────────────────────────────────────────────
def build_rb_qc_tab(wb_obj, section_name, units):
    """section_name: 'RB' or 'QC'. Columns: 單位 / 上月人數(readonly) / 本月人數(editable)."""
    ws = wb_obj.create_sheet(section_name)

    col_widths = [42, 12, 12]
    headers    = ['單位', '上月人數', '本月人數']
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20
    ws.merge_cells('A1:C1')
    c = ws['A1']
    c.value = f'{section_name} — 廠務組織編制表 {MONTH_SH}（本月人數可修改）'
    c.font = Font(bold=True, size=12)
    c.alignment = CENTER

    row = 2
    ws.row_dimensions[row].height = 28
    for ci, h in enumerate(headers, 1):
        wc(ws, row, ci, h, font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=THIN_BORDER)
    ws.freeze_panes = 'A3'

    GRAY_LOCK = PatternFill("solid", fgColor="F2F2F2")
    row = 3

    if not units:
        ws.merge_cells('A3:C3')
        wc(ws, row, 1, f'（{section_name} 無資料）', font=NORM, align=CENTER)
        return ws

    for u in units:
        ws.row_dimensions[row].height = 14
        wc(ws, row, 1, u['unit'],       font=NORM,  align=LEFT,  border=THIN_BORDER)
        lm = u['last_month']
        tm = u['this_month']
        # 上月人數：灰底唯讀視覺提示
        wc(ws, row, 2, lm, font=NORM, fill=GRAY_LOCK, align=RIGHT, border=THIN_BORDER,
           num_fmt='#,##0')
        # 本月人數：白底，可修改
        if tm is not None:
            wc(ws, row, 3, tm, font=NORM, align=RIGHT, border=THIN_BORDER, num_fmt='#,##0')
        else:
            wc(ws, row, 3, None, fill=RED_FILL, border=THIN_BORDER)
        row += 1

    # Total row
    ws.row_dimensions[row].height = 14
    last_total = sum(u['last_month'] for u in units if u['last_month'] is not None)
    this_total = sum(u['this_month'] for u in units if u['this_month'] is not None)
    wc(ws, row, 1, '合計', font=BOLD, fill=GRAY_FILL, align=RIGHT, border=THIN_BORDER)
    wc(ws, row, 2, last_total, font=BOLD, fill=GRAY_FILL, align=RIGHT, border=THIN_BORDER, num_fmt='#,##0')
    wc(ws, row, 3, this_total, font=BOLD, fill=GRAY_FILL, align=RIGHT, border=THIN_BORDER, num_fmt='#,##0')

    # Note row
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    wc(ws, row, 1, '＊ 上月人數唯讀（灰底）。本月人數可直接修改。',
       font=Font(size=9, italic=True, color="666666"), align=LEFT)

    return ws


# ── Build one OCS fixed-unit tab ──────────────────────────────────────────────
def build_ocs_fixed_tab(wb, section_name, units):
    ws = wb.create_sheet(f'OCS_{section_name}')

    col_widths = [40, 10]
    headers    = ['單位', '人數']
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20
    ws.merge_cells('A1:B1')
    c = ws['A1']
    c.value = f'OCS {section_name} — 廠務組織編制表 {MONTH_SH}'
    c.font = Font(bold=True, size=12)
    c.alignment = CENTER

    row = 2
    ws.row_dimensions[row].height = 28
    for ci, h in enumerate(headers, 1):
        wc(ws, row, ci, h, font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=THIN_BORDER)
    ws.freeze_panes = 'A3'
    row = 3

    if not units:
        ws.merge_cells('A3:B3')
        wc(ws, row, 1, f'（{section_name} 無資料）', font=NORM, align=CENTER)
        return ws

    for u in units:
        ws.row_dimensions[row].height = 14
        wc(ws, row, 1, u['unit'],      font=NORM, align=LEFT,  border=THIN_BORDER)
        hc = u['headcount']
        if hc is not None:
            wc(ws, row, 2, hc, font=NORM, align=RIGHT, border=THIN_BORDER, num_fmt='#,##0')
        else:
            wc(ws, row, 2, None, fill=RED_FILL, border=THIN_BORDER)
        row += 1

    # Total row
    ws.row_dimensions[row].height = 14
    total = sum(u['headcount'] for u in units if u['headcount'] is not None)
    wc(ws, row, 1, '合計', font=BOLD, fill=GRAY_FILL, align=RIGHT, border=THIN_BORDER)
    wc(ws, row, 2, total,  font=BOLD, fill=GRAY_FILL, align=RIGHT, border=THIN_BORDER, num_fmt='#,##0')

    return ws


# ── Helpers for writing styled cells ─────────────────────────────────────────
def wc(ws, row, col, value, font=None, fill=None, align=None, border=None, num_fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font = font
    if fill:   cell.fill = fill
    if align:  cell.alignment = align
    if border: cell.border = border
    if num_fmt: cell.number_format = num_fmt
    return cell


# ── Build CSA sheet ───────────────────────────────────────────────────────────
def build_csa(wb, schedule_groups, bianche_rows):
    ws = wb.create_sheet('CSA')

    # Column widths
    col_widths = [6, 18, 10, 8, 7, 7, 7, 6, 6, 4]
    headers    = ['LEAN', '鞋型 (Model+ART)', 'ART', '訂單', '裁斷MP', '針車MP', '成型MP', '協理給', '合計', 'MP?']
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # LEAN comes from DS-04 group titles, not from bianche (see load_schedule)

    # Title row
    ws.row_dimensions[1].height = 20
    ws.merge_cells('A1:J1')
    c = ws['A1']
    c.value = f'CSA 結果表 — 廠務組織編制表 {MONTH_SH}  (MP欄位空白=待確認)'
    c.font = Font(bold=True, size=12)
    c.alignment = CENTER

    # Header row
    row = 2
    ws.row_dimensions[row].height = 30
    for ci, h in enumerate(headers, 1):
        wc(ws, row, ci, h, font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=THIN_BORDER)
    ws.freeze_panes = 'A3'

    row = 3

    for group_info in schedule_groups:
        sheet = group_info['sheet']
        # Use rows (per lean+art) instead of the aggregated orders dict so that
        # the same ART in different LEAN groups appears as separate CSA rows.
        sheet_rows = group_info.get('rows', [])

        # Section header
        ws.row_dimensions[row].height = 16
        ws.merge_cells(f'A{row}:J{row}')
        c = wc(ws, row, 1, f'── {sheet} ──',
               font=Font(bold=True, size=10, color="1F3864"),
               fill=GRAY_FILL, align=CENTER, border=THIN_BORDER)
        row += 1

        for r in sorted(sheet_rows, key=lambda x: (x['lean'], x['art'])):
            art   = r['art']
            qty   = r['qty']
            lean  = r['lean']
            model = get_model(art)

            ws.row_dimensions[row].height = 14
            wc(ws, row, 1,  lean,              font=BOLD, fill=LEAN_FILL if lean else None, align=CENTER, border=THIN_BORDER)
            wc(ws, row, 2,  model[:40] if model else art, font=NORM, align=LEFT, border=THIN_BORDER)
            wc(ws, row, 3,  art,               font=NORM, align=CENTER, border=THIN_BORDER)
            wc(ws, row, 4,  qty,               font=NORM, align=RIGHT,  border=THIN_BORDER, num_fmt='#,##0')
            # MP columns blank, red fill = pending
            for ci in [5, 6, 7]:
                wc(ws, row, ci, None, fill=RED_FILL, border=THIN_BORDER)
            wc(ws, row, 8,  None, border=THIN_BORDER)  # 協理給
            wc(ws, row, 9,  None, border=THIN_BORDER)  # 合計
            wc(ws, row, 10, '?', font=Font(color="FF0000", size=9), align=CENTER, border=THIN_BORDER)
            row += 1

    # Legend
    row += 1
    ws.merge_cells(f'A{row}:J{row}')
    wc(ws, row, 1, '＊ 紅色欄位 = MP 邏輯待確認（Jim 定義後更新）',
       font=RED_FONT, align=LEFT)

    return ws


# ── Build OCS sheet ───────────────────────────────────────────────────────────
def build_ocs(wb, ocs_rows, ocs_err):
    ws = wb.create_sheet('OCS大底課')

    col_widths = [12, 16, 12, 8, 18, 8]
    headers    = ['T群組', '鞋型 (Model)', 'AD代碼', 'ARTs', '訂單', '人數']
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = 'OCS 大底課 — DS-05 T群組'
    c.font = Font(bold=True, size=12)
    c.alignment = CENTER

    row = 2
    ws.row_dimensions[row].height = 30
    for ci, h in enumerate(headers, 1):
        wc(ws, row, ci, h, font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=THIN_BORDER)
    ws.freeze_panes = 'A3'
    row = 3

    if ocs_err:
        ws.merge_cells(f'A{row}:F{row}')
        wc(ws, row, 1, f'DS-05 載入失敗: {ocs_err}', font=Font(color="FF0000"), align=LEFT)
        return ws

    if not ocs_rows:
        ws.merge_cells(f'A{row}:F{row}')
        wc(ws, row, 1, '（無 DS-05 資料）', font=NORM, align=CENTER)
        return ws

    current_group = None
    for r in ocs_rows:
        if r['group'] != current_group:
            current_group = r['group']
            ws.row_dimensions[row].height = 16
            ws.merge_cells(f'A{row}:F{row}')
            hc_str = '  '.join(f"{m}:{v}人" for m, v in r['headcount'].items())
            wc(ws, row, 1, f'{current_group}   {hc_str}',
               font=Font(bold=True, size=10, color="1F3864"),
               fill=GRAY_FILL, align=LEFT, border=THIN_BORDER)
            row += 1

        arts_str = ', '.join(r['arts']) if r['arts'] else ''
        ws.row_dimensions[row].height = 14
        wc(ws, row, 1, r['group'],      font=NORM, align=CENTER, border=THIN_BORDER)
        wc(ws, row, 2, r['model_name'][:40], font=NORM, align=LEFT, border=THIN_BORDER)
        wc(ws, row, 3, r['ad_code'],    font=NORM, align=CENTER, border=THIN_BORDER)
        wc(ws, row, 4, arts_str[:40],   font=NORM, align=LEFT, border=THIN_BORDER)
        wc(ws, row, 5, r['total_qty'],  font=NORM, align=RIGHT, border=THIN_BORDER, num_fmt='#,##0')
        wc(ws, row, 6, None,            fill=RED_FILL, border=THIN_BORDER)
        row += 1

    return ws


# ── Build Ref (廠務編制表) sheet ───────────────────────────────────────────────
def build_ref(wb, bianche_rows):
    ws = wb.create_sheet('廠務編制表_Ref')

    col_widths = [6, 45, 8, 7, 7, 7, 6, 6]
    headers    = ['LEAN', '鞋型', '訂單', '裁斷', '針車', '成型', '協理', '合計']
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20
    ws.merge_cells('A1:H1')
    c = ws['A1']
    c.value = f'廠務組織編制表 {MONTH_SH} 原始資料（唯讀參考）'
    c.font = Font(bold=True, size=12)
    c.alignment = CENTER

    row = 2
    ws.row_dimensions[row].height = 28
    for ci, h in enumerate(headers, 1):
        wc(ws, row, ci, h, font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=THIN_BORDER)
    ws.freeze_panes = 'A3'
    row = 3

    current_lean = ''
    for r in bianche_rows:
        if r['lean'] != current_lean:
            current_lean = r['lean']
        ws.row_dimensions[row].height = 14
        wc(ws, row, 1, r['lean'],                      font=BOLD, fill=LEAN_FILL, align=CENTER, border=THIN_BORDER)
        wc(ws, row, 2, r['model'][:60],                font=NORM, align=LEFT,    border=THIN_BORDER)
        wc(ws, row, 3, r['order'],                     font=NORM, align=RIGHT,   border=THIN_BORDER, num_fmt='#,##0')
        wc(ws, row, 4, r['cut'],                       font=NORM, align=RIGHT,   border=THIN_BORDER)
        wc(ws, row, 5, r['stitch'],                    font=NORM, align=RIGHT,   border=THIN_BORDER)
        wc(ws, row, 6, r['asm'],                       font=NORM, align=RIGHT,   border=THIN_BORDER)
        wc(ws, row, 7, r['assist'],                    font=NORM, align=RIGHT,   border=THIN_BORDER)
        wc(ws, row, 8, r['total'],                     font=NORM, align=RIGHT,   border=THIN_BORDER)
        row += 1

    return ws


# ── Build 差異 sheet ──────────────────────────────────────────────────────────
def build_diff_sheet(wb, schedule_groups, bianche_rows):
    """Non-MP differences: ARTs in schedule not in ref, ref not in schedule."""
    ws = wb.create_sheet('非MP差異')

    # Build lookup sets
    sched_arts = set()
    for g in schedule_groups:
        sched_arts.update(g['orders'].keys())

    ref_arts = set()
    art_to_ref = {}
    for r in bianche_rows:
        for art in r['arts']:
            ref_arts.add(art)
            art_to_ref[art] = r

    in_sched_not_ref = sorted(sched_arts - ref_arts)
    in_ref_not_sched = sorted(ref_arts - sched_arts)

    headers = ['類型', 'ART', '鞋型', '備註']
    col_widths = [18, 12, 35, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20
    ws.merge_cells('A1:D1')
    c = ws['A1']
    c.value = f'非MP欄位差異（ART/鞋型/訂單）'
    c.font = Font(bold=True, size=12)
    c.alignment = CENTER

    row = 2
    for ci, h in enumerate(headers, 1):
        wc(ws, row, ci, h, font=HDR_FONT, fill=HDR_FILL, align=CENTER, border=THIN_BORDER)
    ws.freeze_panes = 'A3'
    row = 3

    # ARTs in schedule but not in ref
    if in_sched_not_ref:
        ws.merge_cells(f'A{row}:D{row}')
        wc(ws, row, 1, f'▼ 進度表有 但 廠務編制表無  ({len(in_sched_not_ref)} 個 ART)',
           font=Font(bold=True, color="C00000"), fill=GRAY_FILL, align=LEFT, border=THIN_BORDER)
        row += 1
        for art in in_sched_not_ref:
            model = get_model(art)
            ws.row_dimensions[row].height = 14
            wc(ws, row, 1, '進度表有/編制表無', font=NORM, fill=PatternFill("solid", fgColor="FFE0E0"), align=LEFT, border=THIN_BORDER)
            wc(ws, row, 2, art,   font=NORM, align=CENTER, border=THIN_BORDER)
            wc(ws, row, 3, model, font=NORM, align=LEFT,   border=THIN_BORDER)
            wc(ws, row, 4, '請確認是否需加入編制表', font=Font(color="FF6600", italic=True, size=9), align=LEFT, border=THIN_BORDER)
            row += 1

    row += 1

    # ARTs in ref but not in schedule
    if in_ref_not_sched:
        ws.merge_cells(f'A{row}:D{row}')
        wc(ws, row, 1, f'▼ 廠務編制表有 但 進度表無  ({len(in_ref_not_sched)} 個 ART)',
           font=Font(bold=True, color="0070C0"), fill=GRAY_FILL, align=LEFT, border=THIN_BORDER)
        row += 1
        for art in in_ref_not_sched:
            ref = art_to_ref.get(art, {})
            lean = ref.get('lean', '')
            model = ref.get('model', '')[:40]
            ws.row_dimensions[row].height = 14
            wc(ws, row, 1, '編制表有/進度表無', font=NORM, fill=PatternFill("solid", fgColor="E0E8FF"), align=LEFT, border=THIN_BORDER)
            wc(ws, row, 2, art,   font=NORM, align=CENTER, border=THIN_BORDER)
            wc(ws, row, 3, f'[{lean}] {model}', font=NORM, align=LEFT, border=THIN_BORDER)
            wc(ws, row, 4, '確認是否已停產或改ART', font=Font(color="0070C0", italic=True, size=9), align=LEFT, border=THIN_BORDER)
            row += 1

    return ws, len(in_sched_not_ref), len(in_ref_not_sched)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    db.init_db()
    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)

    print('Loading DS-04 schedule...')
    schedule_groups = load_schedule()
    total_arts = sum(len(g['orders']) for g in schedule_groups)
    print(f'  {len(schedule_groups)} sheets, {total_arts} ARTs')

    print('Loading DS-05 OCS...')
    ocs_rows, ocs_err = load_ocs()
    if ocs_err:
        print(f'  WARNING: {ocs_err}')
    else:
        print(f'  {len(ocs_rows)} T-group model rows')

    print('Loading 廠務組織編制表...')
    bianche_rows = load_bianche_structure()
    print(f'  {len(bianche_rows)} ref rows')

    print('Loading OCS fixed units...')
    ocs_fixed = load_ocs_fixed_units()
    for sec, units in ocs_fixed.items():
        print(f'  {sec}: {len(units)} units')

    print('Building Excel...')
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    build_csa(wb, schedule_groups, bianche_rows)
    build_ocs(wb, ocs_rows, ocs_err)
    for sec in OCS_SECTIONS:
        build_ocs_fixed_tab(wb, sec, ocs_fixed.get(sec, []))
    build_ref(wb, bianche_rows)
    ws_diff, n_sched_only, n_ref_only = build_diff_sheet(wb, schedule_groups, bianche_rows)

    wb.save(OUTPUT)
    print(f'\nSaved: {OUTPUT}')
    print(f'Sheets: {[s.title for s in wb.worksheets]}')
    print(f'\n非MP差異摘要:')
    print(f'  進度表有/編制表無: {n_sched_only} 個ART')
    print(f'  編制表有/進度表無: {n_ref_only} 個ART')


if __name__ == '__main__':
    main()
