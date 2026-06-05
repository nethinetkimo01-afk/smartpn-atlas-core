#!/usr/bin/env python3
"""
專門輸出 1A 完整結果表，格式與廠務組織編制表 6.2026 sheet 一致，並列出差異
"""
import sys, os, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
import database as db

BIANCHE = r'C:\Users\user\OneDrive\Desktop\Biên chế\Jun\2026年6月份廠務组织編制 20260524.xlsx'
SCHEDULE = r'C:\Users\user\OneDrive\Desktop\Biên chế\Jun\2026年6月份正式进度表 5 30.xlsx'
MONTH_SHEET = '6.2026'
EOLR = 120
TARGET_LEAN = '1A'

_ART_RE = re.compile(r'[A-Z]{2}\d{4,6}')
_QTY_RE = re.compile(r'--(\d+)\s*\(')


def get_mp(art):
    try:
        conn = db.get_conn()
        row = conn.execute(
            '''SELECT e.cutting, e.stitching, e.assembly
               FROM ob_epph e JOIN ob_header h ON h.id = e.header_id
               WHERE h.art = ? AND h.eolr = ?
               ORDER BY h.id DESC LIMIT 1''',
            (art, EOLR)
        ).fetchone()
        conn.close()
        if row and any(v for v in row):
            return row[0] or 0, row[1] or 0, row[2] or 0
        return None
    except Exception:
        return None


def get_model(art):
    try:
        conn = db.get_conn()
        row = conn.execute('SELECT model_name FROM ds02_fob WHERE art = ? LIMIT 1', (art,)).fetchone()
        conn.close()
        return (row[0] or '').strip() if row else ''
    except Exception:
        return ''


def parse_order_cell(val):
    s = str(val or '').strip()
    arts = _ART_RE.findall(s)
    m = _QTY_RE.search(s)
    qty = int(m.group(1)) if m else 0
    if not arts:
        return []
    per = qty // len(arts)
    rem = qty % len(arts)
    return [(a, per + (rem if i == 0 else 0)) for i, a in enumerate(arts)]


def is_order_cell(val):
    s = str(val or '').strip()
    return bool(_ART_RE.search(s) and '--' in s)


def load_schedule_orders():
    """Load all orders from DS-04 schedule (all sheets)."""
    wb = openpyxl.load_workbook(SCHEDULE, data_only=True, read_only=True)
    orders = {}  # art -> qty
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if not is_order_cell(cell):
                    continue
                for art, qty in parse_order_cell(cell):
                    orders[art] = orders.get(art, 0) + qty
    wb.close()
    return orders


def load_bianche_1a():
    """Read 1A rows from 廠務組織編制表 6.2026 sheet."""
    wb = openpyxl.load_workbook(BIANCHE, data_only=True)
    ws = None
    for sh in wb.sheetnames:
        if sh.strip() == MONTH_SHEET:
            ws = wb[sh]
            break
    if ws is None:
        print(f'ERROR: sheet {MONTH_SHEET!r} not found. Available: {wb.sheetnames}')
        return []

    rows = []
    current_lean = ''
    in_1a = False

    for rn, row in enumerate(ws.iter_rows(values_only=True), 1):
        col_a = str(row[0] or '').strip() if row else ''
        col_b = str(row[1] or '').strip() if len(row) > 1 else ''

        if re.match(r'^\d+[A-Z]$', col_a):
            current_lean = col_a
            in_1a = (current_lean == TARGET_LEAN)

        if not in_1a:
            continue

        if (not col_a or col_a == '.') and col_b and len(col_b) > 3:
            def _val(idx):
                v = row[idx] if len(row) > idx else None
                if v is None or str(v).strip() in ('.', ''):
                    return None
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return None

            order  = _val(6)
            cut    = _val(7)
            stitch = _val(8)
            asm    = _val(9)
            assist = _val(10) or 0
            total  = _val(11)

            arts_in_model = _ART_RE.findall(col_b)

            rows.append({
                'model':   col_b,
                'arts':    arts_in_model,
                'order':   order,
                'cut':     cut,
                'stitch':  stitch,
                'asm':     asm,
                'assist':  assist,
                'total':   total,
            })

    wb.close()
    return rows


def main():
    db.init_db()

    print('Loading 廠務組織編制表 1A ...')
    ref_rows = load_bianche_1a()
    print(f'  Found {len(ref_rows)} 1A reference rows\n')

    print('Loading DS-04 schedule orders ...')
    all_orders = load_schedule_orders()
    print(f'  Total ARTs in schedule: {len(all_orders)}\n')

    # ── 1. Print 1A reference table (廠務組織編制表 format) ───────────────────
    sep = '─' * 110
    print(f'{"=" * 110}')
    print(f'  廠務組織編制表 6.2026 ── LEAN 1A  (完整輸出)')
    print(f'{"=" * 110}')
    hdr = f'  {"型號 / ART":<40}  {"訂單":>7}  {"裁斷":>6}  {"針車":>6}  {"成型":>6}  {"協理":>5}  {"合計":>6}'
    print(hdr)
    print(f'  {sep}')
    for r in ref_rows:
        order  = f"{int(r['order']):>7}"   if r['order']  is not None else f"{'–':>7}"
        cut    = f"{r['cut']:.1f}"         if r['cut']    is not None else '–'
        stitch = f"{r['stitch']:.1f}"      if r['stitch'] is not None else '–'
        asm    = f"{r['asm']:.1f}"         if r['asm']    is not None else '–'
        assist = f"{r['assist']:.1f}"      if r['assist']             else '–'
        total  = f"{r['total']:.1f}"       if r['total']  is not None else '–'
        print(f"  {r['model'][:40]:<40}  {order}  {cut:>6}  {stitch:>6}  {asm:>6}  {assist:>5}  {total:>6}")

    # ── 2. Print our computed MP for each 1A ART ─────────────────────────────
    # Collect all ARTs mentioned in ref_rows
    ref_arts = []
    for r in ref_rows:
        ref_arts.extend(r['arts'])
    # Also include any ART in schedule that belongs to 1A (heuristic: check all schedule ARTs)
    # We use ref_rows as authoritative list for 1A
    all_1a_arts = list(dict.fromkeys(ref_arts))  # deduplicated, ordered

    print()
    print(f'{"=" * 110}')
    print(f'  DB 計算值（ob_epph EOLR={EOLR}）── 1A ARTs')
    print(f'{"=" * 110}')
    hdr2 = f'  {"ART":<12}  {"Model":<36}  {"訂單":>7}  {"裁斷":>6}  {"針車":>6}  {"成型":>6}  {"合計":>6}  {"MP"}'
    print(hdr2)
    print(f'  {sep}')
    our_map = {}  # art -> our row
    for art in all_1a_arts:
        qty   = all_orders.get(art, 0)
        model = get_model(art)
        mp    = get_mp(art)
        cut    = round(mp[0], 1) if mp else None
        stitch = round(mp[1], 1) if mp else None
        asm    = round(mp[2], 1) if mp else None
        total  = round(cut + stitch + asm, 1) if mp else None
        mp_ok  = '✓' if mp else '✗'
        our_map[art] = {'art': art, 'model': model, 'qty': qty,
                        'cut': cut, 'stitch': stitch, 'asm': asm, 'total': total, 'mp_ok': mp_ok}

        q = f"{qty:>7}" if qty else f"{'(無)':>7}"
        c = f"{cut:.1f}"    if cut    is not None else '–'
        s = f"{stitch:.1f}" if stitch is not None else '–'
        a = f"{asm:.1f}"    if asm    is not None else '–'
        t = f"{total:.1f}"  if total  is not None else '–'
        print(f"  {art:<12}  {model[:36]:<36}  {q}  {c:>6}  {s:>6}  {a:>6}  {t:>6}  {mp_ok}")

    # ── 3. Diff table ─────────────────────────────────────────────────────────
    print()
    print(f'{"=" * 110}')
    print(f'  差異清單  (Our DB – 廠務編制表 Ref)')
    print(f'{"=" * 110}')

    diff_hdr = f'  {"ART":<12}  {"型號":<30}  {"OurCut":>7}  {"RefCut":>7}  {"ΔCut":>6}  {"OurS":>6}  {"RefS":>6}  {"ΔS":>5}  {"OurA":>6}  {"RefA":>6}  {"ΔA":>5}'
    print(diff_hdr)
    print(f'  {sep}')

    diffs = []
    for r in ref_rows:
        for art in r['arts']:
            our = our_map.get(art)
            if not our:
                continue
            ref_cut    = r['cut']
            ref_stitch = r['stitch']
            ref_asm    = r['asm']
            our_cut    = our['cut']
            our_stitch = our['stitch']
            our_asm    = our['asm']

            d_cut    = (our_cut    or 0) - (ref_cut    or 0)
            d_stitch = (our_stitch or 0) - (ref_stitch or 0)
            d_asm    = (our_asm    or 0) - (ref_asm    or 0)

            if abs(d_cut) > 0.05 or abs(d_stitch) > 0.05 or abs(d_asm) > 0.05:
                diffs.append({
                    'art': art, 'model': r['model'][:30],
                    'our_cut': our_cut, 'ref_cut': ref_cut, 'd_cut': d_cut,
                    'our_stitch': our_stitch, 'ref_stitch': ref_stitch, 'd_stitch': d_stitch,
                    'our_asm': our_asm, 'ref_asm': ref_asm, 'd_asm': d_asm,
                })

    if not diffs:
        print('  (無差異)')
    else:
        for d in diffs:
            oc = f"{d['our_cut']:.1f}"    if d['our_cut']    is not None else '–'
            rc = f"{d['ref_cut']:.1f}"    if d['ref_cut']    is not None else '(new)'
            dc = f"{d['d_cut']:+.1f}"
            os_ = f"{d['our_stitch']:.1f}" if d['our_stitch'] is not None else '–'
            rs = f"{d['ref_stitch']:.1f}"  if d['ref_stitch'] is not None else '(new)'
            ds = f"{d['d_stitch']:+.1f}"
            oa = f"{d['our_asm']:.1f}"    if d['our_asm']    is not None else '–'
            ra = f"{d['ref_asm']:.1f}"    if d['ref_asm']    is not None else '(new)'
            da = f"{d['d_asm']:+.1f}"
            print(f"  {d['art']:<12}  {d['model']:<30}  {oc:>7}  {rc:>7}  {dc:>6}  {os_:>6}  {rs:>6}  {ds:>5}  {oa:>6}  {ra:>6}  {da:>5}")

    print()
    print(f'  合計差異: {len(diffs)} 筆 / {len(all_1a_arts)} 個 ART')

    # ── 4. ARTs in 1A ref but no MP in DB ────────────────────────────────────
    no_mp = [art for art in all_1a_arts if our_map.get(art, {}).get('mp_ok') == '✗']
    if no_mp:
        print()
        print(f'  ！DB 無 MP 資料 ({len(no_mp)} 個 ART)：')
        print('  ' + ', '.join(no_mp))


if __name__ == '__main__':
    main()
