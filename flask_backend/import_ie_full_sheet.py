#!/usr/bin/env python3
"""
Import all cells (value + formula) from every sheet of one IE xlsx into ie_sheet_data.

Usage:
    python import_ie_full_sheet.py <path_to_xlsx> [header_id]

If header_id is omitted, the script looks up the first ART found in the filename.
"""
import sys, os, re

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import database as db
import openpyxl

_ART_RE = re.compile(r'[A-Z]{2}\d{4,6}')


def resolve_header_id(conn, path):
    fn = os.path.basename(path)
    arts = _ART_RE.findall(fn)
    if not arts:
        return None, []
    for art in arts:
        row = conn.execute(
            '''SELECT h.id, h.eolr FROM ob_articles a
               JOIN ob_header h ON h.id=a.header_id
               WHERE a.art=? ORDER BY h.id DESC LIMIT 1''', (art,)
        ).fetchone()
        if row:
            return row['id'], arts
    return None, arts


def import_file(xlsx_path, header_id=None, verbose=True):
    db.init_db()

    conn = db.get_conn()
    # Run migration for ie_sheet_data if missing
    tables = [r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
    if 'ie_sheet_data' not in tables:
        conn.execute('''CREATE TABLE IF NOT EXISTS ie_sheet_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            header_id INTEGER NOT NULL,
            sheet_name TEXT NOT NULL,
            row INTEGER NOT NULL,
            col INTEGER NOT NULL,
            value TEXT,
            formula TEXT
        )''')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_ie_sheet_data_hdr ON ie_sheet_data(header_id, sheet_name)')
        conn.commit()

    if header_id is None:
        header_id, arts = resolve_header_id(conn, xlsx_path)
        if header_id is None:
            conn.close()
            print(f'ERROR: could not find header_id for {os.path.basename(xlsx_path)}')
            print(f'  ARTs in filename: {arts}')
            return {'ok': False, 'error': 'header_id not found'}
        if verbose:
            print(f'Resolved header_id={header_id} (ARTs: {arts})')
    conn.close()

    if verbose:
        print(f'File: {os.path.basename(xlsx_path)}')

    # Pass 1: formulas (data_only=False)
    if verbose:
        print('Pass 1: reading formulas ...')
    wb_f = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=False)
    sheets = list(wb_f.sheetnames)
    if verbose:
        print(f'  Sheets ({len(sheets)}): {sheets}')

    formula_map = {}  # (sheet, row, col) -> formula string
    for sn in sheets:
        ws = wb_f[sn]
        for row in ws.iter_rows():
            for cell in row:
                if not hasattr(cell, 'row') or not hasattr(cell, 'column'):
                    continue
                v = cell.value
                if v is not None and isinstance(v, str) and v.startswith('='):
                    formula_map[(sn, cell.row, cell.column)] = v
    wb_f.close()
    if verbose:
        print(f'  Formulas found: {len(formula_map)}')

    # Pass 2: computed values (data_only=True)
    if verbose:
        print('Pass 2: reading computed values ...')
    wb_v = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    sheet_rows = []
    total_cells = 0
    for sn in sheets:
        ws = wb_v[sn]
        sheet_count = 0
        for row in ws.iter_rows():
            for cell in row:
                if not hasattr(cell, 'row') or not hasattr(cell, 'column'):
                    continue
                v = cell.value
                formula = formula_map.get((sn, cell.row, cell.column))
                # Skip cells that are completely empty (no value, no formula)
                if v is None and formula is None:
                    continue
                val_str = None if v is None else str(v)
                sheet_rows.append((sn, cell.row, cell.column, val_str, formula))
                sheet_count += 1
        if verbose:
            print(f'  {sn!r}: {sheet_count} non-empty cells')
        total_cells += sheet_count
    wb_v.close()

    if verbose:
        print(f'Total non-empty cells: {total_cells}')
        print(f'Writing to DB (header_id={header_id}) ...')

    result = db.save_ie_sheet_data(header_id, sheet_rows)
    if result['ok']:
        if verbose:
            print(f'Done: {result["rows_written"]} rows written for {len(sheets)} sheets.')
    else:
        print(f'ERROR: {result["error"]}')
    return result


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    path = sys.argv[1]
    hid = int(sys.argv[2]) if len(sys.argv) > 2 else None
    r = import_file(path, hid)
    sys.exit(0 if r.get('ok') else 1)
