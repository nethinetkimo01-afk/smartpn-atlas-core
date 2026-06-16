#!/usr/bin/env python3
"""
IE 導入對比表 — 人員數字版
從 SUM.C2B 讀取各段人員配置（裁斷/針車/成型/貼底），
對比 DB 中 Σ(standard_time)/30 的理論人數，輸出 xlsx。
PPH=120 → 每雙節拍 = 3600/120 = 30 秒 → 理論人數 = ΣST / 30
"""
import sys, os, re, sqlite3
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

IE_XLSX_DIR = r'D:\smartpn-atlas-core\data\source_files\IE_xlsx'
DB_PATH     = os.path.join(os.path.dirname(__file__), 'data', 'atlas.db')
OUT_DIR     = os.path.join(os.path.dirname(__file__), 'test_output')
os.makedirs(OUT_DIR, exist_ok=True)

TAKT = 30.0   # 3600 / 120 PPH

def tf(v):
    if v is None: return None
    try: return float(str(v).replace(',','').strip())
    except: return None

def _is_headcount(v):
    """Number looks like headcount (0.5–500), not PPH or TCT."""
    f = tf(v)
    return f is not None and 0.4 < f < 600

def read_sum_c2b_mp(wb):
    """Return dict {cutting, stitching, assembly, stf} headcount from SUM.C2B main data row."""
    sumc = None
    for cand in ('SUM.C2B','Sum.C2B','sum.c2b','SUM C2B','SUMC2B'):
        if cand in wb.sheetnames:
            sumc = cand; break
    if sumc is None:
        return None

    ws = wb[sumc]
    result = {'cutting': None, 'stitching': None, 'assembly': None, 'stf': None}

    # Strategy: look for the data row where col[1] contains '双/H' (production rate)
    # that row has: [2]=cutting_MP, [5]=stitching_MP, [8]=assembly_MP, [11]=stf_MP
    for row in ws.iter_rows(max_row=60, values_only=True):
        r = list(row)
        # Pad to at least 13 elements
        while len(r) < 13:
            r.append(None)
        label = str(r[1] or '').strip()
        # Row must have a production rate in col B (e.g. '120双/H', '960双/8H')
        if not re.search(r'\d+双.*[/／]', label):
            continue
        # Col C (idx 2) must look like headcount
        if not _is_headcount(r[2]):
            continue
        # Looks like the right row — read all 4 headcounts
        result['cutting']   = tf(r[2])
        result['stitching'] = tf(r[5])
        result['assembly']  = tf(r[8])
        result['stf']       = tf(r[11])
        return result

    # Fallback: look for a row where cols 2,5,8,11 all look like headcount
    ws2 = wb[sumc]
    rows_seen = []
    for row in ws2.iter_rows(max_row=60, values_only=True):
        r = list(row)
        while len(r) < 13:
            r.append(None)
        if all(_is_headcount(r[i]) for i in (2, 5, 8, 11)):
            rows_seen.append(r)
    if rows_seen:
        r = rows_seen[0]
        result['cutting']   = tf(r[2])
        result['stitching'] = tf(r[5])
        result['assembly']  = tf(r[8])
        result['stf']       = tf(r[11])
        return result

    return result  # all None if not found

def get_db_theory(conn, header_id):
    """Return Σ(standard_time)/TAKT per segment from ie_process."""
    rows = conn.execute(
        'SELECT segment, SUM(standard_time) FROM ie_process WHERE header_id=? GROUP BY segment',
        (header_id,)
    ).fetchall()
    d = {'cutting': None, 'stitching': None, 'assembly': None, 'stf': None}
    for seg, s in rows:
        if seg in d and s is not None:
            d[seg] = round(s / TAKT, 2)
    return d

def resolve_header(conn, fpath):
    """Match xlsx filename to ob_header, return (header_id, art, model_name)."""
    fname = os.path.basename(fpath)
    arts = re.findall(r'[A-Z]{2}\d{4,5}', fname)
    for art in arts:
        row = conn.execute(
            'SELECT h.id, a.art, h.model_name FROM ob_header h JOIN ob_articles a ON a.header_id=h.id WHERE a.art=?',
            (art,)
        ).fetchone()
        if row:
            return row[0], row[1], row[2]
    return None, None, None

def main():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    conn = sqlite3.connect(DB_PATH)
    results = []

    xlsx_files = sorted([
        f for f in os.listdir(IE_XLSX_DIR)
        if f.lower().endswith('.xlsx') and not f.startswith('~$')
    ])
    print(f'Found {len(xlsx_files)} xlsx files')

    for fname in xlsx_files:
        fpath = os.path.join(IE_XLSX_DIR, fname)
        header_id, art, model_name = resolve_header(conn, fpath)
        if header_id is None:
            print(f'  SKIP (no header): {fname[:60]}')
            continue

        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f'  ERROR opening [{header_id}] {fname[:50]}: {e}')
            continue

        orig = read_sum_c2b_mp(wb)
        wb.close()

        if orig is None:
            orig = {'cutting': None, 'stitching': None, 'assembly': None, 'stf': None}
            print(f'  [{header_id}] {art}  — no SUM.C2B')
        else:
            found = [k for k, v in orig.items() if v is not None]
            if not found:
                print(f'  [{header_id}] {art}  — SUM.C2B found but no data row')
            else:
                print(f'  [{header_id}] {art}  c={orig["cutting"]} s={orig["stitching"]} a={orig["assembly"]} f={orig["stf"]}')

        theory = get_db_theory(conn, header_id)
        results.append({
            'art': art,
            'model': model_name or '',
            'orig': orig,
            'theory': theory,
        })

    conn.close()
    _write_xlsx(results)

def _write_xlsx(results):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'IE人員對比'

    hdr_fill = PatternFill('solid', fgColor='1A2744')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    sub_fill = PatternFill('solid', fgColor='2E4070')
    sub_font = Font(bold=True, color='FFFFFF', size=9)
    red_fill = PatternFill('solid', fgColor='FFB3B3')
    grn_fill = PatternFill('solid', fgColor='C8F0C8')
    ctr = Alignment(horizontal='center', vertical='center')

    # Row 1: section headers (merged)
    SEGS = [
        ('裁斷 Cutting', 3, 5),
        ('針車 Stitching', 6, 8),
        ('成型 Assembly', 9, 11),
        ('貼底 Stockfitting', 12, 14),
    ]
    ws.cell(1, 1, '鞋型').fill = hdr_fill
    ws.cell(1, 1).font = hdr_font
    ws.cell(1, 2, 'ART').fill = hdr_fill
    ws.cell(1, 2).font = hdr_font
    for label, sc, ec in SEGS:
        c = ws.cell(1, sc, label)
        c.fill = sub_fill; c.font = sub_font; c.alignment = ctr
        ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)

    # Row 2: column headers
    cols = ['鞋型', 'ART',
            'SUM裁斷原始', '導入後裁斷', '差異',
            'SUM針車原始', '導入後針車', '差異',
            'SUM成型原始', '導入後成型', '差異',
            'SUM貼底原始', '導入後貼底', '差異']
    for ci, h in enumerate(cols, 1):
        c = ws.cell(2, ci, h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = ctr

    # Data rows
    segs_order = ('cutting', 'stitching', 'assembly', 'stf')
    for ri, row in enumerate(results, 3):
        ws.cell(ri, 1, row['model'][:50] if row['model'] else '')
        ws.cell(ri, 2, row['art'])
        for si, seg in enumerate(segs_order):
            orig_v  = row['orig'].get(seg)
            thry_v  = row['theory'].get(seg)
            diff    = round(thry_v - orig_v, 2) if (orig_v is not None and thry_v is not None) else None
            base_col = 3 + si * 3
            ws.cell(ri, base_col,     round(orig_v, 2) if orig_v is not None else None)
            ws.cell(ri, base_col + 1, thry_v)
            dc = ws.cell(ri, base_col + 2, diff)
            if diff is not None:
                if abs(diff) < 0.01:
                    dc.fill = grn_fill
                else:
                    dc.fill = red_fill

    # Column widths
    from openpyxl.utils import get_column_letter
    widths = [30, 10, 12, 12, 8, 12, 12, 8, 12, 12, 8, 12, 12, 8]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'C3'
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30

    out = os.path.join(OUT_DIR, 'ie_import_comparison.xlsx')
    wb.save(out)
    print(f'\nSaved: {out}  ({len(results)} rows)')

    # Print summary stats
    total_red = 0
    for row in results:
        for seg in ('cutting', 'stitching', 'assembly', 'stf'):
            o = row['orig'].get(seg)
            t = row['theory'].get(seg)
            if o is not None and t is not None and abs(t - o) >= 0.01:
                total_red += 1
    print(f'Red cells (差異≠0): {total_red}')

if __name__ == '__main__':
    main()
