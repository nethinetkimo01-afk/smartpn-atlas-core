#!/usr/bin/env python3
"""Inspect 廠務組織編制表 structure to find RB and QC section headers."""
import sys, os, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl

BIANCHE = r'C:\Users\user\AppData\Local\Temp\bianche_tmp.xlsx'
MONTH_SH = '6.2026'

wb = openpyxl.load_workbook(BIANCHE, data_only=True, read_only=True)
ws = next((wb[s] for s in wb.sheetnames if s.strip() == MONTH_SH), None)
print(f'Sheet: {MONTH_SH}')
print()

# Print all non-empty col1 values with row number, to find section headers
KNOWN_LEANS = re.compile(r'^\d+[A-Z]\d*$')
KNOWN_OCS   = {'組底配套', '自動化', '電腦針車', '印刷', '設備工程'}

for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    col1 = str(row[0] or '').strip()
    col2 = str(row[1] or '').strip() if len(row) > 1 else ''
    col13 = row[12] if len(row) > 12 else None
    col11 = row[10] if len(row) > 10 else None
    if not col1:
        continue
    if KNOWN_LEANS.match(col1):
        continue  # skip LEAN header rows (too many)
    # Print candidate section headers
    print(f'R{i:4d}  col1={col1!r:30s}  col2={col2!r:30s}  col11={str(col11)!r:6s}  col13={str(col13)!r}')

wb.close()
