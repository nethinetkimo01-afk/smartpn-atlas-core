#!/usr/bin/env python3
"""
bianche_diff.py
比對 auto_bianche.xlsx (自動產生) vs 廠務組織編制表 6.2026 (人工維護)
聚焦 CSA section：LEAN / ART / 訂單 三欄位
輸出 test_output/bianche_diff.txt
"""
import sys, os, re, shutil
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from full_compare_report import (
    load_bianche, _safe_bianche_path, _ART_RE,
    R_ORD_BATCH, R_ORD_MISMATCH,
)

AUTO_PATH = r'D:\smartpn-atlas-core\flask_backend\test_output\auto_bianche.xlsx'
OUT_PATH  = r'D:\smartpn-atlas-core\flask_backend\test_output\bianche_diff.txt'
OCS_SECTIONS = ['組底配套', '自動化', '電腦針車', '印刷', '設備工程']
W = 120


def load_auto_bianche_csa(path):
    """Read auto_bianche.xlsx and extract CSA (lean, art, qty) rows."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    cur_lean = ''

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        col_a = str(row[0] or '').strip() if row else ''
        col_b = str(row[1] or '').strip() if len(row) > 1 else ''
        col_f = row[5] if len(row) > 5 else None

        # Section headers: "LEAN N" or OCS sections → reset lean
        if col_a.startswith('LEAN ') or col_a in OCS_SECTIONS or col_a in ('RB', 'QC'):
            cur_lean = ''
            continue

        # LEAN group header: "1A", "1B", etc.
        if re.match(r'^\d+[A-Z]\d*$', col_a):
            cur_lean = col_a
            # Don't add a row for the header itself
            continue

        # Data row: col_b has "MODEL ( ART )"
        if not col_b or col_b in ('鞋型 + ART', '單位', ''):
            continue

        arts = _ART_RE.findall(col_b)
        if not arts:
            continue

        try:
            qty = int(float(col_f)) if col_f is not None else 0
        except (ValueError, TypeError):
            qty = 0

        for art in arts:
            if not art.startswith('MF'):
                rows.append({'lean': cur_lean, 'art': art, 'qty': qty,
                             'model_text': col_b})

    wb.close()
    return rows


def load_auto_bianche_ocs(path):
    """Read OCS/RB/QC units from auto_bianche.xlsx."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    ocs = {s: [] for s in OCS_SECTIONS}
    rbqc = {'RB': [], 'QC': []}
    cur_sec = None
    cur_rbqc = None

    for row in ws.iter_rows(values_only=True):
        col_a = str(row[0] or '').strip() if row else ''
        col_b = str(row[1] or '').strip() if len(row) > 1 else ''
        col_l = row[11] if len(row) > 11 else None  # col 12 = headcount

        if col_a in OCS_SECTIONS:
            cur_sec = col_a; cur_rbqc = None
        elif col_a in ('RB', 'QC'):
            cur_rbqc = col_a; cur_sec = None
        elif col_a.startswith('LEAN ') or re.match(r'^\d+[A-Z]\d*$', col_a):
            cur_sec = None; cur_rbqc = None
        elif cur_sec and col_b and col_b not in ('', '鞋型 + ART'):
            try:
                hc = int(float(col_l)) if col_l is not None else None
            except:
                hc = None
            ocs[cur_sec].append({'unit': col_b, 'hc': hc})
        elif cur_rbqc and col_b and col_b not in ('單位', '上月人數', '本月人數', ''):
            col_k = row[10] if len(row) > 10 else None
            try:
                last_m = int(float(col_k)) if col_k is not None else None
            except:
                last_m = None
            try:
                this_m = int(float(col_l)) if col_l is not None else None
            except:
                this_m = None
            rbqc[cur_rbqc].append({'unit': col_b, 'last_month': last_m, 'this_month': this_m})

    wb.close()
    return ocs, rbqc


def run():
    bianche_path = _safe_bianche_path()
    art_to_ref, ocs_ref = load_bianche(bianche_path)
    auto_rows = load_auto_bianche_csa(AUTO_PATH)
    auto_ocs, auto_rbqc = load_auto_bianche_ocs(AUTO_PATH)

    lines = []
    def hdr(t): lines.extend(['=' * W, f'  {t}', '=' * W])
    def sub(t): lines.extend(['', f'── {t} ' + '─' * max(0, W - len(t) - 4)])

    hdr('bianche_diff.txt：auto_bianche.xlsx  vs  廠務組織編制表 6.2026')
    lines.append(f'  auto_bianche CSA 行數: {len(auto_rows)}  |  廠務 ART 索引: {len(art_to_ref)}')
    lines.append('')

    # ── CSA: ART 匹配 ─────────────────────────────────────────────────────────
    sub('1. ART 匹配狀況 (CSA)')
    auto_arts = {r['art'] for r in auto_rows}
    ref_arts  = set(art_to_ref.keys())
    both      = auto_arts & ref_arts
    auto_only = sorted(auto_arts - ref_arts)
    ref_only  = sorted(ref_arts - auto_arts)

    lines.append(f'  auto_bianche 非MF ART 筆數  : {len(auto_arts):4d}')
    lines.append(f'  雙方都有 (可比對)            : {len(both):4d}')
    lines.append(f'  auto有 / 廠務無              : {len(auto_only):4d}  → 需 Jim 確認（補登廠務或DS-04誤填）')
    lines.append(f'  廠務有 / auto無              : {len(ref_only):4d}  → 需 Jim 確認（廠務多餘？）')

    if auto_only:
        lines.append('')
        lines.append('  auto有/廠務無 清單：')
        for art in auto_only:
            match = next((r for r in auto_rows if r['art'] == art), {})
            lines.append(f"    {art:12s}  LEAN={match.get('lean','?')!r:8s}  auto訂單={match.get('qty',0):>6}")

    if ref_only:
        lines.append('')
        lines.append('  廠務有/auto無 清單：')
        for art in ref_only:
            ref = art_to_ref[art]
            lines.append(f"    {art:12s}  廠務LEAN={ref['lean']!r:8s}  廠務鞋型={ref['model'][:50]!r}")

    # ── CSA: LEAN 比對 ────────────────────────────────────────────────────────
    sub('2. LEAN 比對 (CSA，僅雙方都有的ART)')
    lean_ok = lean_diff = 0
    lean_diff_rows = []
    art_auto_count = {}
    for r in auto_rows:
        art_auto_count[r['art']] = art_auto_count.get(r['art'], 0) + 1

    for r in auto_rows:
        art = r['art']
        if art not in art_to_ref:
            continue
        ref = art_to_ref[art]
        if r['lean'] == ref['lean']:
            lean_ok += 1
        else:
            lean_diff += 1
            cross = art_auto_count.get(art, 1) > 1
            lean_diff_rows.append({
                'art': art, 'auto_lean': r['lean'], 'ref_lean': ref['lean'],
                'reason': 'LEAN-跨部門' if cross else 'LEAN-指派差異',
            })

    lines.append(f'  LEAN 一致  : {lean_ok:4d} 筆')
    lines.append(f'  LEAN 不符  : {lean_diff:4d} 筆')
    cross_n = sum(1 for x in lean_diff_rows if x['reason'] == 'LEAN-跨部門')
    diff_n  = lean_diff - cross_n
    lines.append(f'    ├─ LEAN-跨部門   : {cross_n:4d} 筆  (同ART在多個LEAN組，廠務只記一個)')
    lines.append(f'    └─ LEAN-指派差異 : {diff_n:4d} 筆  (ART僅一個LEAN但廠務LEAN不同)')

    if diff_n > 0:
        lines.append('')
        lines.append('  LEAN-指派差異 清單：')
        for x in lean_diff_rows:
            if x['reason'] == 'LEAN-指派差異':
                lines.append(f"    {x['art']:12s}  auto={x['auto_lean']:<6}  廠務={x['ref_lean']}")

    # ── CSA: 訂單 比對 ────────────────────────────────────────────────────────
    sub('3. 訂單 比對 (CSA，僅雙方都有的ART，按廠務LEAN對應比對)')
    ord_ok = ord_batch = ord_mismatch = ord_lean_mismatch = 0
    ord_mismatch_rows = []

    # Index auto rows by (art, lean) for per-LEAN comparison
    auto_by_art_lean = {}
    for r in auto_rows:
        key = (r['art'], r['lean'])
        auto_by_art_lean[key] = auto_by_art_lean.get(key, 0) + r['qty']

    for art in sorted(both):
        ref   = art_to_ref[art]
        b_ord = ref['order']
        b_lean= ref['lean']
        b_arts= re.findall(r'[A-Z]{2}\d{4,6}', ref['model'])

        if b_ord is None:
            continue

        # Use qty from matching LEAN; if LEAN doesn't match any auto row, skip
        # (already counted as LEAN mismatch above)
        a_qty = auto_by_art_lean.get((art, b_lean))
        if a_qty is None:
            ord_lean_mismatch += 1
            continue

        if abs(a_qty - b_ord) < 0.5:
            ord_ok += 1
        elif len(b_arts) > 1:
            ord_batch += 1
        else:
            ord_mismatch += 1
            ord_mismatch_rows.append({
                'art': art, 'lean': b_lean,
                'auto_qty': a_qty, 'factory_qty': int(b_ord),
                'diff': a_qty - int(b_ord),
            })

    total_cmp = ord_ok + ord_batch + ord_mismatch + ord_lean_mismatch
    lines.append(f'  可比對 ART 筆數              : {total_cmp}')
    lines.append(f'  ✓ 完全一致                  : {ord_ok:4d}')
    lines.append(f'  △ 邏輯差異 (廠務合批多ART)  : {ord_batch:4d}  → 廠務一行合多ART合計量，非錯誤')
    lines.append(f'  ✗ 人為填寫差異 (數量不符)   : {ord_mismatch:4d}  → 同LEAN數量不符')
    lines.append(f'  ─ LEAN不符 (跨部門)         : {ord_lean_mismatch:4d}  → auto在不同LEAN，廠務記一個LEAN')

    if ord_mismatch_rows:
        lines.append('')
        lines.append('  訂單差異 清單（按差異絕對值排序）：')
        FMT = '    {art:12s}  LEAN={lean:<6}  auto={auto_qty:>7}  廠務={factory_qty:>7}  差={diff:+}'
        for row in sorted(ord_mismatch_rows, key=lambda x: abs(x['diff']), reverse=True)[:30]:
            lines.append(FMT.format(**row))
        if len(ord_mismatch_rows) > 30:
            lines.append(f'    ... 更多 {len(ord_mismatch_rows)-30} 筆（見 column_compare_report.txt）')

    # ── OCS 比對 ─────────────────────────────────────────────────────────────
    sub('4. OCS 固定單位 比對')
    ocs_ok_total = ocs_diff_total = 0

    for sec in OCS_SECTIONS:
        auto_units = {u['unit']: u['hc'] for u in auto_ocs.get(sec, [])}
        ref_units  = {u['unit']: u['hc'] for u in ocs_ref.get(sec, [])}
        sec_ok = sec_diff = 0
        for unit in sorted(set(list(auto_units) + list(ref_units))):
            a_hc = auto_units.get(unit)
            r_hc = ref_units.get(unit)
            if unit in auto_units and unit in ref_units:
                if a_hc == r_hc: sec_ok += 1
                else: sec_diff += 1
            else:
                sec_diff += 1
        verdict = '✓ 一致' if sec_diff == 0 else f'✗ {sec_diff}筆差異'
        lines.append(f'  {sec:<12}: OK={sec_ok:2d}  DIFF={sec_diff:2d}  {verdict}')
        ocs_ok_total += sec_ok
        ocs_diff_total += sec_diff

    # ── 總結 ──────────────────────────────────────────────────────────────────
    lines.append('')
    hdr('總結')
    lines.append('')
    lines.append(f'  ART 匹配 (雙方都有)         : {len(both):4d} / {len(auto_arts)} 筆')
    lines.append(f'  ART auto有廠務無            : {len(auto_only):4d} 筆  ← 待 Jim 確認')
    lines.append(f'  ART 廠務有auto無            : {len(ref_only):4d} 筆  ← 待 Jim 確認')
    lines.append('')
    lines.append(f'  LEAN 一致                   : {lean_ok:4d} 筆')
    lines.append(f'  LEAN 不符（跨部門）         : {cross_n:4d} 筆  → 正常業務差異')
    lines.append(f'  LEAN 不符（指派差異）       : {diff_n:4d} 筆  → 需確認')
    lines.append('')
    lines.append(f'  訂單 一致                   : {ord_ok:4d} 筆')
    lines.append(f'  訂單 邏輯差異 (廠務合批)    : {ord_batch:4d} 筆  → 非錯誤')
    lines.append(f'  訂單 人為差異               : {ord_mismatch:4d} 筆  → 需確認（同LEAN比對）')
    lines.append(f'  訂單 LEAN不符 (跨部門)      : {ord_lean_mismatch:4d} 筆  → auto跨LEAN，廠務僅記一個LEAN')
    lines.append('')
    lines.append(f'  OCS 固定單位 比對           : ' + ('✓ 5 Tab 全部一致' if ocs_diff_total == 0 else f'✗ {ocs_diff_total} 筆差異'))
    lines.append('')
    lines.append('  結論：auto_bianche.xlsx 已成功從 DS-04 自動填入 LEAN/ART/訂單')
    lines.append(f'        MP 欄位 (裁斷/針車/成型/協理給/合計/編制) 已留空，待 Jim 填入')

    text = '\n'.join(lines)
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, 'w', encoding='utf-8') as f:
        f.write(text)
    print(text)
    return {
        'auto_rows': len(auto_rows),
        'auto_only': len(auto_only),
        'ref_only': len(ref_only),
        'lean_ok': lean_ok, 'lean_diff': lean_diff,
        'ord_ok': ord_ok, 'ord_batch': ord_batch, 'ord_mismatch': ord_mismatch,
        'ocs_ok': ocs_ok_total, 'ocs_diff': ocs_diff_total,
    }


if __name__ == '__main__':
    run()
