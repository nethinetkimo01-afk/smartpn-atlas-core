#!/usr/bin/env python3
"""
MP 假說驗證：
  假說：廠務 MP = ob_epph.section × (actual_EOLR / ob_header.eolr)

  從 1A / 2B / 3A LEAN 各抽取 ART，對照廠務編制表，輸出誤差分析。
  不修改任何資料。
"""
import sys, os, re
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
import database as db

# ─── CONFIG ───────────────────────────────────────────────────────────────────
BIANCHE = r'C:\Users\user\OneDrive\Desktop\Biên chế\Jun\2026年6月份廠務组织編制 20260524.xlsx'
MONTH_SHEET = '6.2026'

# Actual EOLR per LEAN group (provided by Jim)
LEAN_ACTUAL_EOLR = {
    '1A': 75, '1B': 145,
    '2A': 120, '2B': 60,
    '3A': 150,
    '8B': 120, '9A': 140,
}

TARGET_LEANS = ['1A', '2B', '3A']
SAMPLE_PER_LEAN = 4   # aim for ~4 per group → ~12 rows total

OUTPUT = os.path.join(os.path.dirname(__file__), 'test_output', 'mp_hypothesis_test.txt')
_ART_RE = re.compile(r'[A-Z]{2}\d{4,6}')

# ─── DB helpers ───────────────────────────────────────────────────────────────

def get_epph_for_art(art):
    """
    Return list of (header_id, eolr, cutting, stitching, assembly) for an ART.
    """
    conn = db.get_conn()
    rows = conn.execute(
        '''SELECT h.id, h.eolr, e.cutting, e.stitching, e.assembly
           FROM ob_epph e
           JOIN ob_header h ON h.id = e.header_id
           JOIN ob_articles a ON a.header_id = h.id
           WHERE a.art = ?
           ORDER BY h.eolr DESC''',
        (art,)
    ).fetchall()
    conn.close()
    return rows


def best_epph(art, target_eolr=None):
    """
    Pick the best ob_epph row for an ART.
    Prefer exact eolr match, then any available.
    Returns (db_eolr, cutting, stitching, assembly) or None.
    """
    rows = get_epph_for_art(art)
    if not rows:
        return None
    if target_eolr:
        exact = [r for r in rows if r[1] == target_eolr]
        if exact:
            r = exact[0]
            return (r[1], r[2], r[3], r[4])
    r = rows[0]
    return (r[1], r[2], r[3], r[4])

# ─── Read BIANCHE ──────────────────────────────────────────────────────────────

def load_bianche_lean(target_leans):
    """
    Read BIANCHE sheet, return dict: lean -> [row_dict, ...]
    row_dict: model, arts, cut, stitch, asm, total
    """
    if not os.path.exists(BIANCHE):
        print(f'ERROR: BIANCHE file not found: {BIANCHE}')
        return {}

    wb = openpyxl.load_workbook(BIANCHE, data_only=True)
    ws = None
    for sh in wb.sheetnames:
        if sh.strip() == MONTH_SHEET:
            ws = wb[sh]
            break
    if ws is None:
        # Try first sheet
        ws = wb.active
        print(f'  WARNING: sheet {MONTH_SHEET!r} not found, using {ws.title!r}')

    result = {lean: [] for lean in target_leans}
    current_lean = ''

    for row in ws.iter_rows(values_only=True):
        col_a = str(row[0] or '').strip() if row else ''
        col_b = str(row[1] or '').strip() if len(row) > 1 else ''

        if re.match(r'^\d+[A-Z]$', col_a):
            current_lean = col_a

        if current_lean not in target_leans:
            continue

        # Row with model data: col_a empty (or '.'), col_b has model name
        if col_b and len(col_b) > 2 and (not col_a or col_a == '.'):
            def _v(idx):
                v = row[idx] if len(row) > idx else None
                if v is None or str(v).strip() in ('.', ''):
                    return None
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return None

            arts_found = _ART_RE.findall(col_b)
            if not arts_found:
                continue

            result[current_lean].append({
                'lean': current_lean,
                'model': col_b,
                'arts':  arts_found,
                'ref_cut':    _v(7),
                'ref_stitch': _v(8),
                'ref_asm':    _v(9),
                'ref_total':  _v(11),
            })

    wb.close()
    return result

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    db.init_db()

    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)

    print('Loading BIANCHE...')
    lean_rows = load_bianche_lean(TARGET_LEANS)
    for lean in TARGET_LEANS:
        print(f'  {lean}: {len(lean_rows.get(lean, []))} rows loaded')

    lines = []
    lines.append('=' * 100)
    lines.append('  MP 假說驗證報告')
    lines.append('  假說: 廠務MP = ob_epph.section × (actual_EOLR / ob_header.eolr)')
    lines.append('  Actual EOLR: ' + ', '.join(f'{k}={v}' for k, v in sorted(LEAN_ACTUAL_EOLR.items())))
    lines.append('=' * 100)
    lines.append('')

    summary_rows = []   # for final stats

    for lean in TARGET_LEANS:
        actual_eolr = LEAN_ACTUAL_EOLR.get(lean)
        if actual_eolr is None:
            lines.append(f'SKIP {lean}: no actual EOLR defined')
            continue

        ref_rows = lean_rows.get(lean, [])
        if not ref_rows:
            lines.append(f'SKIP {lean}: no BIANCHE rows found')
            continue

        # Sample up to SAMPLE_PER_LEAN rows
        sample = ref_rows[:SAMPLE_PER_LEAN]

        lines.append(f'─── LEAN {lean}  (actual EOLR = {actual_eolr}) ─────────────────────────────────────────────────')
        lines.append('')
        header = (f"  {'ART':<12}  {'ref_cut':>7}  {'adj_cut':>7}  {'Δcut%':>6}"
                  f"  {'ref_sew':>7}  {'adj_sew':>7}  {'Δsew%':>6}"
                  f"  {'ref_asm':>7}  {'adj_asm':>7}  {'Δasm%':>6}"
                  f"  {'db_eolr':>7}  note")
        lines.append(header)
        lines.append('  ' + '─' * 95)

        for ref in sample:
            for art in ref['arts'][:2]:   # max 2 ARTs per model row
                epph = best_epph(art)
                if epph is None:
                    lines.append(f"  {art:<12}  (no ob_epph data)")
                    summary_rows.append({'lean': lean, 'art': art, 'pass': None, 'note': 'no_db'})
                    continue

                db_eolr, db_cut, db_sew, db_asm = epph

                scale = actual_eolr / db_eolr if db_eolr else 1.0

                adj_cut = round((db_cut or 0) * scale, 2)
                adj_sew = round((db_sew or 0) * scale, 2)
                adj_asm = round((db_asm or 0) * scale, 2)

                ref_cut    = ref['ref_cut']
                ref_stitch = ref['ref_stitch']
                ref_asm_v  = ref['ref_asm']

                def pct(adj, ref_val):
                    if ref_val and ref_val != 0:
                        return f'{((adj - ref_val) / ref_val * 100):+.1f}%'
                    return 'N/A'

                def fmt(v):
                    return f'{v:.2f}' if v is not None else '–'

                d_cut = pct(adj_cut, ref_cut)
                d_sew = pct(adj_sew, ref_stitch)
                d_asm = pct(adj_asm, ref_asm_v)

                # Pass = all deltas within ±10%
                def within10(adj, ref_val):
                    if ref_val is None or ref_val == 0:
                        return None
                    return abs(adj - ref_val) / ref_val <= 0.10

                cut_ok = within10(adj_cut, ref_cut)
                sew_ok = within10(adj_sew, ref_stitch)
                asm_ok = within10(adj_asm, ref_asm_v)

                checks = [v for v in [cut_ok, sew_ok, asm_ok] if v is not None]
                overall_pass = all(checks) if checks else None
                pass_label = 'PASS' if overall_pass else ('FAIL' if overall_pass is False else '?')

                line = (f"  {art:<12}  {fmt(ref_cut):>7}  {adj_cut:>7.2f}  {d_cut:>6}"
                        f"  {fmt(ref_stitch):>7}  {adj_sew:>7.2f}  {d_sew:>6}"
                        f"  {fmt(ref_asm_v):>7}  {adj_asm:>7.2f}  {d_asm:>6}"
                        f"  {db_eolr:>7}  [{pass_label}]")
                lines.append(line)

                summary_rows.append({
                    'lean': lean, 'art': art,
                    'pass': overall_pass,
                    'db_eolr': db_eolr, 'actual_eolr': actual_eolr, 'scale': scale,
                    'note': pass_label
                })

        lines.append('')

    # ── Summary ──────────────────────────────────────────────────────────────
    lines.append('=' * 100)
    lines.append('  SUMMARY')
    lines.append('=' * 100)
    total = len([r for r in summary_rows if r['pass'] is not None])
    passed = len([r for r in summary_rows if r['pass'] is True])
    failed = len([r for r in summary_rows if r['pass'] is False])
    no_data = len([r for r in summary_rows if r['pass'] is None])

    lines.append(f'  Total ARTs tested: {len(summary_rows)}')
    lines.append(f'  With DB data:      {total}')
    lines.append(f'  PASS (±10%):       {passed}  ({passed/total*100:.0f}% of testable)' if total else '  PASS: –')
    lines.append(f'  FAIL:              {failed}')
    lines.append(f'  No DB data:        {no_data}')
    lines.append('')
    lines.append('  Hypothesis verdict:')
    if total == 0:
        lines.append('  ─ Cannot evaluate: no DB data for sampled ARTs')
    elif passed / total >= 0.8:
        lines.append(f'  ✓ SUPPORTED  ({passed}/{total} ARTs within ±10% after EOLR scaling)')
    elif passed / total >= 0.5:
        lines.append(f'  ~ PARTIALLY  ({passed}/{total} ARTs pass — scaling direction is correct but precision varies)')
    else:
        lines.append(f'  ✗ NOT SUPPORTED  (only {passed}/{total} ARTs within ±10%)')
    lines.append('')
    lines.append('  Note: No data was modified. Read-only analysis only.')
    lines.append('=' * 100)

    output_text = '\n'.join(lines)
    print(output_text)

    with open(OUTPUT, 'w', encoding='utf-8') as f:
        f.write(output_text + '\n')
    print(f'\n[Written to {OUTPUT}]')


if __name__ == '__main__':
    main()
