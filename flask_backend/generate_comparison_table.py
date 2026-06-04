#!/usr/bin/env python3
"""
Generate comparison_table.xlsx: 計算值 vs 結果表 完整比對表
欄位: LEAN | 鞋型 | ART | 訂單 | 計算Cut | 結果表Cut | 計算Stitch | 結果表Stitch | 計算Asm | 結果表Asm | 狀態
狀態: OK / MISSING_IE / MISMATCH / NO_REF
空值或不一致欄位用紅色背景標示。

Usage:
    python generate_comparison_table.py
        --schedule <DS04.xlsx>
        --bianche  <廠務組織編制表.xlsx>
        --month    6.2026
        [--eolr    120]
        [--output  flask_backend/test_output/comparison_table.xlsx]
"""
import sys, os, re, argparse

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import database as db
from analyze_result_table import (
    extract_all_groups, build_csa_rows, get_mp, get_model_name,
    read_bianche_csa, compare_csa
)

# ── Colors ────────────────────────────────────────────────────────────────────
RED_FILL   = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
GREEN_FILL = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
YELLOW_FILL= PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
GREY_FILL  = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
HEADER_FILL= PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
WHITE_FONT = Font(color='FFFFFF', bold=True)
BOLD_FONT  = Font(bold=True)

THIN = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = [
    'LEAN', '鞋型', 'ART', '訂單',
    '計算Cut', '結果表Cut', '計算Stitch', '結果表Stitch',
    '計算Asm',  '結果表Asm', '狀態'
]

# Column widths
COL_WIDTHS = [8, 30, 12, 8, 10, 10, 12, 12, 10, 10, 12]


def status_for_row(r, matched):
    """Determine status label for a row."""
    if matched is None:
        return 'NO_REF'
    if not r.get('mp_ok'):
        return 'MISSING_IE'
    diff_cut = abs((r.get('cut') or 0) - (matched.get('cut') or 0))
    diff_s   = abs((r.get('stitch') or 0) - (matched.get('stitch') or 0))
    diff_a   = abs((r.get('asm') or 0) - (matched.get('asm') or 0))
    if diff_cut <= 0.5 and diff_s <= 0.5 and diff_a <= 0.5:
        return 'OK'
    return 'MISMATCH'


def match_ref_row(r, ref_rows):
    """Best-effort match CSA row against reference rows by ART or model name."""
    art = r['art']
    model = (r.get('model') or '').upper()
    for ref in ref_rows:
        if art in (ref.get('model') or ''):
            return ref
    # Try model name substring
    if model and len(model) > 4:
        for ref in ref_rows:
            ref_model = (ref.get('model') or '').upper()
            if model[:15] in ref_model or ref_model[:15] in model:
                return ref
    return None


def build_comparison_rows(schedule_path, bianche_path, month_sheet, eolr=120):
    """Build all comparison rows combining schedule, ob_epph, and reference."""
    if not HAS_OPENPYXL:
        return [], 'openpyxl not installed'

    # Load schedule
    if not os.path.exists(schedule_path):
        return [], f'Schedule not found: {schedule_path}'
    wb04 = openpyxl.load_workbook(schedule_path, data_only=True, read_only=True)
    schedule_groups = extract_all_groups(wb04)
    wb04.close()

    csa_rows = build_csa_rows(schedule_groups, eolr)

    # Load reference
    ref_by_lean = {}   # lean → [ref_rows]
    if bianche_path and os.path.exists(bianche_path):
        wb_ref = openpyxl.load_workbook(bianche_path, data_only=True)
        ref_all = read_bianche_csa(wb_ref, month_sheet)
        wb_ref.close()
        for r in ref_all:
            lean = r.get('lean', '')
            ref_by_lean.setdefault(lean, []).append(r)
    else:
        return [], f'Reference not found: {bianche_path}'

    # Match each CSA row to a reference
    comparison = []
    for r in csa_rows:
        sheet = r.get('sheet', '')
        lean = sheet.strip()

        # Try to find LEAN group from reference
        matched = None
        ref_rows = ref_by_lean.get(lean, [])
        if not ref_rows:
            # Search all leans
            for l, rows in ref_by_lean.items():
                m = match_ref_row(r, rows)
                if m:
                    lean = m['lean']
                    matched = m
                    break
        else:
            matched = match_ref_row(r, ref_rows)

        status = status_for_row(r, matched)
        comparison.append({
            'lean':       lean,
            'model':      r.get('model', ''),
            'art':        r.get('art', ''),
            'qty':        r.get('qty') or 0,
            'calc_cut':   r.get('cut'),
            'ref_cut':    matched.get('cut') if matched else None,
            'calc_stitch':r.get('stitch'),
            'ref_stitch': matched.get('stitch') if matched else None,
            'calc_asm':   r.get('asm'),
            'ref_asm':    matched.get('asm') if matched else None,
            'status':     status,
        })

    return comparison, None


def write_xlsx(rows, output_path, eolr=120):
    """Write comparison rows to colored Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '比對表'

    # Header row
    for col, (hdr, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 20

    # Data rows
    for row_idx, r in enumerate(rows, 2):
        status = r['status']

        # Determine row fill by status
        row_fill = None
        if status == 'OK':
            row_fill = GREEN_FILL
        elif status == 'MISSING_IE':
            row_fill = YELLOW_FILL
        elif status == 'MISMATCH':
            row_fill = None  # highlight individual cells
        elif status == 'NO_REF':
            row_fill = GREY_FILL

        values = [
            r['lean'],
            r['model'][:40] if r['model'] else '',
            r['art'],
            r['qty'],
            r['calc_cut'],
            r['ref_cut'],
            r['calc_stitch'],
            r['ref_stitch'],
            r['calc_asm'],
            r['ref_asm'],
            status,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center' if col != 2 else 'left')

            if row_fill:
                cell.fill = row_fill
            elif status == 'MISMATCH':
                # Highlight mismatched value pairs in red
                calc_col = {5: 'cut', 7: 'stitch', 9: 'asm'}
                ref_col  = {6: 'cut', 8: 'stitch', 10: 'asm'}

                if col in calc_col:
                    dept = calc_col[col]
                    cv = r[f'calc_{dept}']
                    rv = r[f'ref_{dept}']
                    if cv is not None and rv is not None and abs(cv - rv) > 0.5:
                        cell.fill = RED_FILL
                elif col in ref_col:
                    dept = ref_col[col]
                    cv = r[f'calc_{dept}']
                    rv = r[f'ref_{dept}']
                    if cv is not None and rv is not None and abs(cv - rv) > 0.5:
                        cell.fill = RED_FILL
                elif col in (1, 3, 4, 11):
                    pass  # no fill for meta cols in MISMATCH rows
            # Null values in red
            if val is None and col in (5, 7, 9):
                cell.fill = RED_FILL

    # Freeze header
    ws.freeze_panes = 'A2'

    # Add summary sheet
    ws2 = wb.create_sheet('摘要')
    status_counts = {}
    for r in rows:
        s = r['status']
        status_counts[s] = status_counts.get(s, 0) + 1

    ws2.append(['狀態', '筆數', '說明'])
    ws2.append(['OK',         status_counts.get('OK', 0),         '計算值與結果表一致'])
    ws2.append(['MISSING_IE', status_counts.get('MISSING_IE', 0), '無IE文件，計算值為空'])
    ws2.append(['MISMATCH',   status_counts.get('MISMATCH', 0),   '有IE文件但數值不一致'])
    ws2.append(['NO_REF',     status_counts.get('NO_REF', 0),     '結果表無此ART'])
    ws2.append(['合計',       len(rows), ''])

    for col in (1, 2, 3):
        ws2.column_dimensions[get_column_letter(col)].width = 16

    ws2.append([])
    ws2.append(['EOLR', eolr])
    ws2.append(['生成時間', '2026-06-04'])
    ws2.append(['說明', 'MISMATCH = ob_epph為整條產線MP，結果表為各組分配值，邏輯待Jim確認'])

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return status_counts


# ── CLI ───────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--schedule', default=r'C:\Users\user\OneDrive\Desktop\Biên chế\Jun\2026年6月份正式进度表 5 30.xlsx')
    parser.add_argument('--bianche',  default=r'C:\Users\user\OneDrive\Desktop\Biên chế\2026年6月份廠務组织編制_20260524.xlsx')
    parser.add_argument('--month',    default='6.2026')
    parser.add_argument('--eolr',     type=int, default=120)
    parser.add_argument('--output',   default='test_output/comparison_table.xlsx')
    args = parser.parse_args()

    if not HAS_OPENPYXL:
        print('openpyxl not installed')
        sys.exit(1)

    db.init_db()

    rows, err = build_comparison_rows(args.schedule, args.bianche, args.month, args.eolr)
    if err:
        print(f'Error: {err}')
        sys.exit(1)

    counts = write_xlsx(rows, args.output, args.eolr)
    print(f'Written: {args.output}')
    print(f'Total rows: {len(rows)}')
    for k, v in sorted(counts.items()):
        print(f'  {k}: {v}')
