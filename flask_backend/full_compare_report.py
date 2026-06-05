#!/usr/bin/env python3
"""
full_compare_report.py  v2
行號 | 結果表(LEAN/鞋型/ART/訂單) | 廠務表(LEAN/鞋型/ART/訂單) | 狀態(OK/差異) | 差異原因
最後輸出總結：一致幾筆/差異幾筆/差異原因分類
"""
import sys, os, re, shutil, tempfile
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl

RESULT_PATH  = r'D:\smartpn-atlas-core\flask_backend\test_output\result_table_v2.xlsx'
BIANCHE_ORIG = r'C:\Users\user\OneDrive\Desktop\Biên chế\Jun\2026年6月份廠務组织編制 20260524.xlsx'
BIANCHE_TMP  = r'C:\Users\user\AppData\Local\Temp\bianche_tmp.xlsx'
OUT_PATH     = r'D:\smartpn-atlas-core\flask_backend\test_output\full_compare_report.txt'
MONTH_SH     = '6.2026'
OCS_SECTIONS = ['組底配套', '自動化', '電腦針車', '印刷', '設備工程']
_ART_RE      = re.compile(r'[A-Z]{2}\d{4,6}')

# ── Diff reason codes ─────────────────────────────────────────────────────────
R_LEAN_CROSS  = 'LEAN-跨部門'   # same ART in multiple DS-04 sections
R_LEAN_DIFF   = 'LEAN-指派差異' # ART in one section, bianche LEAN differs
R_ORD_BATCH   = 'ORD-廠務合批'  # bianche row has multiple ART codes → batch qty
R_ORD_MISMATCH = 'ORD-數量不符' # single ART, qty different
R_ORD_OK      = ''
R_MISS        = 'DS04有廠務無'
R_EXTRA       = '廠務有DS04無'
R_SKIP        = 'MF前綴跳過'


def _safe_bianche_path():
    try:
        with open(BIANCHE_ORIG, 'rb'): pass
        return BIANCHE_ORIG
    except (PermissionError, FileNotFoundError):
        if os.path.exists(BIANCHE_TMP):
            return BIANCHE_TMP
        shutil.copy2(BIANCHE_ORIG, BIANCHE_TMP)
        return BIANCHE_TMP


def load_bianche(path):
    wb  = openpyxl.load_workbook(path, data_only=True)
    ws  = next((wb[s] for s in wb.sheetnames if s.strip() == MONTH_SH), None)
    if ws is None:
        wb.close(); return {}, {}

    def _num(row, idx):
        v = row[idx] if len(row) > idx else None
        if v is None or str(v).strip() in ('.', '', '編制'): return None
        try: return float(v)
        except: return None

    def _hc(row):
        for idx in [12, 10, 14]:
            v = row[idx] if len(row) > idx else None
            if v is not None and str(v).strip() not in ('', '.', '編制'):
                try: return int(float(v))
                except: pass
        return None

    art_to_ref = {}
    ocs_ref    = {s: [] for s in OCS_SECTIONS}
    lean = ''; ocs_sec = None

    for row in ws.iter_rows(values_only=True):
        col1 = str(row[0] or '').strip() if row else ''
        col2 = str(row[1] or '').strip() if len(row) > 1 else ''

        if re.match(r'^\d+[A-Z]\d*$', col1):
            lean = col1; ocs_sec = None
            if col2 and col2 not in ('編制', ''):
                entry = {'lean': lean, 'model': col2, 'order': _num(row, 6)}
                for a in _ART_RE.findall(col2):
                    if a not in art_to_ref: art_to_ref[a] = entry
            continue

        if col1 in OCS_SECTIONS:
            ocs_sec = col1; lean = ''
            if col2 and col2 not in ('編制', ''):
                ocs_ref[ocs_sec].append({'unit': col2, 'hc': _hc(row)})
            continue

        if lean and not col1 and col2 and col2 not in ('編制', ''):
            entry = {'lean': lean, 'model': col2, 'order': _num(row, 6)}
            for a in _ART_RE.findall(col2):
                if a not in art_to_ref: art_to_ref[a] = entry
            continue

        if ocs_sec and not col1 and col2 and col2 not in ('編制', ''):
            ocs_ref[ocs_sec].append({'unit': col2, 'hc': _hc(row)})
            continue

        if col1 and col1 not in OCS_SECTIONS:
            lean = ''; ocs_sec = None

    wb.close()
    return art_to_ref, ocs_ref


def load_result_csa(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['CSA']
    out = []; cur_sec = ''
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 2: continue
        col_a = str(row[0] or '').strip()
        col_b = str(row[1] or '').strip() if len(row) > 1 else ''
        col_c = str(row[2] or '').strip() if len(row) > 2 else ''
        col_d = row[3] if len(row) > 3 else None
        if '──' in col_a or '──' in col_b:
            cur_sec = (col_a or col_b).strip()
            continue
        if not _ART_RE.fullmatch(col_c):
            continue
        try: qty = int(float(col_d)) if col_d is not None else 0
        except: qty = 0
        out.append({'section': cur_sec, 'lean': col_a, 'model': col_b,
                    'art': col_c, 'order': qty})
    wb.close()
    return out


def load_result_ocs(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for sec in OCS_SECTIONS:
        tab = f'OCS_{sec}'
        if tab not in wb.sheetnames: out[sec] = None; continue
        ws = wb[tab]; rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i <= 2: continue
            col_a = str(row[0] or '').strip()
            col_b = row[1] if len(row) > 1 else None
            if not col_a or col_a in ('單位', '合計') or col_a.startswith('（'): continue
            try: hc = int(float(col_b)) if col_b is not None else None
            except: hc = None
            rows.append({'unit': col_a, 'hc': hc})
        out[sec] = rows
    wb.close()
    return out


def run(result_path=RESULT_PATH, bianche_path=None, out_path=OUT_PATH):
    if bianche_path is None:
        bianche_path = _safe_bianche_path()

    art_to_ref, ocs_ref = load_bianche(bianche_path)
    csa_rows            = load_result_csa(result_path)
    ocs_result          = load_result_ocs(result_path)

    # Pre-compute: how many times each ART appears in csa_rows (to detect cross-dept)
    art_section_count = {}
    for r in csa_rows:
        if not r['art'].startswith('MF'):
            art_section_count[r['art']] = art_section_count.get(r['art'], 0) + 1

    lines = []
    W = 130

    def hdr(t): lines.extend(['=' * W, f'  {t}', '=' * W])
    def sub(t): lines.extend(['', f'── {t} ' + '─' * max(0, W - len(t) - 4)])

    hdr('完整比對報告 v2：result_table_v2.xlsx  vs  廠務組織編制表 6.2026')
    lines.append(f'廠務 ART 索引: {len(art_to_ref)} 筆    CSA 行數: {len(csa_rows)} 筆')
    lines.append('')

    # ── CSA Section ───────────────────────────────────────────────────────────
    sub('CSA — 逐行比對  LEAN / 鞋型 / ART / 訂單')
    lines.append('')
    FMT = ('{rn:>4}  {status:<5}  {art:<12}  '
           '結果[{r_lean:<7} {r_model:<22} {r_ord:>7}]  '
           '廠務[{b_lean:<7} {b_model:<22} {b_ord:>7}]  '
           '{lean_ok:<4} {ord_ok:<4}  {reason}')
    lines.append(FMT.format(
        rn='行號', status='狀態', art='ART',
        r_lean='LEAN', r_model='鞋型', r_ord='訂單',
        b_lean='LEAN', b_model='鞋型', b_ord='訂單',
        lean_ok='LEAN', ord_ok='訂單', reason='差異原因'))
    lines.append('-' * W)

    cur_sec = ''
    lean_ok_n = lean_diff_n = lean_miss_n = 0
    ord_ok_n = ord_diff_n = 0
    rn = 0
    art_not_in_ref = []

    # Reason counters
    reason_counts = {
        R_LEAN_CROSS: 0, R_LEAN_DIFF: 0,
        R_ORD_BATCH: 0, R_ORD_MISMATCH: 0,
        R_MISS: 0, R_SKIP: 0,
    }

    for r in csa_rows:
        art   = r['art']
        r_lean = r['lean']
        r_ord  = r['order']
        r_mod  = r['model'][:22]

        if r['section'] != cur_sec:
            cur_sec = r['section']
            lines.append('')
            lines.append(f'  >>> {cur_sec}')
            lines.append('')

        if art.startswith('MF'):
            reason_counts[R_SKIP] += 1
            lines.append(FMT.format(
                rn='─', status='SKIP', art=art,
                r_lean=r_lean or '', r_model=r_mod, r_ord=r_ord,
                b_lean='─', b_model='─', b_ord='─',
                lean_ok='n/a', ord_ok='n/a', reason=R_SKIP))
            continue

        rn += 1

        if art not in art_to_ref:
            lean_miss_n += 1
            reason_counts[R_MISS] += 1
            art_not_in_ref.append(r)
            lines.append(FMT.format(
                rn=rn, status='MISS', art=art,
                r_lean=r_lean or '(空)', r_model=r_mod, r_ord=r_ord,
                b_lean='?', b_model='?', b_ord='?',
                lean_ok='??', ord_ok='??', reason=R_MISS))
            continue

        ref   = art_to_ref[art]
        b_lean = ref['lean']
        b_ord  = ref['order']
        b_mod  = ref['model'][:22]
        b_arts = _ART_RE.findall(ref['model'])

        lean_match = (r_lean == b_lean)
        if lean_match:
            lean_ok_n += 1
            lean_flag = 'OK'
            lean_reason = ''
        else:
            lean_diff_n += 1
            lean_flag = 'DIFF'
            if art_section_count.get(art, 1) > 1:
                lean_reason = R_LEAN_CROSS
            else:
                lean_reason = R_LEAN_DIFF
            reason_counts[lean_reason] += 1

        if b_ord is not None:
            ord_match = (abs(r_ord - b_ord) < 0.5)
            if ord_match:
                ord_flag = 'OK'
                ord_ok_n += 1
                ord_reason = ''
            else:
                ord_flag = 'DIFF'
                ord_diff_n += 1
                if len(b_arts) > 1:
                    ord_reason = R_ORD_BATCH
                else:
                    ord_reason = R_ORD_MISMATCH
                reason_counts[ord_reason] += 1
        else:
            ord_flag = 'N/A'
            ord_reason = ''

        # Combine reasons
        reasons = [x for x in [lean_reason, ord_reason] if x]
        reason_str = ' | '.join(reasons) if reasons else ''

        status = 'OK' if (lean_match and ord_flag in ('OK', 'N/A')) else 'DIFF'
        b_ord_str = str(int(b_ord)) if b_ord is not None else '─'

        lines.append(FMT.format(
            rn=rn, status=status, art=art,
            r_lean=r_lean or '(空)', r_model=r_mod, r_ord=r_ord,
            b_lean=b_lean, b_model=b_mod, b_ord=b_ord_str,
            lean_ok=lean_flag, ord_ok=ord_flag, reason=reason_str))

    # ── CSA Totals ────────────────────────────────────────────────────────────
    lines.append('')
    lines.append('─' * W)
    lines.append(f'CSA 共 {rn} 筆（不含MF前綴）：'
                 f'LEAN一致={lean_ok_n}  LEAN不符={lean_diff_n}  ART缺={lean_miss_n}')
    lines.append(f'訂單：一致={ord_ok_n}  不符={ord_diff_n}')

    if art_not_in_ref:
        lines.append(f'\n⚠  DS04有/廠務無 ART ({len(art_not_in_ref)} 筆):')
        for r in art_not_in_ref:
            lines.append(f"   {r['art']:12s}  LEAN={r['lean']!r:8s}  DS-04訂單={r['order']}")

    result_arts = {r['art'] for r in csa_rows if not r['art'].startswith('MF')}
    ref_only = sorted(a for a in art_to_ref if a not in result_arts)
    if ref_only:
        lines.append(f'\n⚠  廠務有/DS04無 ART ({len(ref_only)} 筆):')
        for a in ref_only:
            ref = art_to_ref[a]
            lines.append(f"   {a:12s}  廠務LEAN={ref['lean']!r:8s}  廠務鞋型={ref['model'][:40]!r}")

    # ── OCS Sections ─────────────────────────────────────────────────────────
    sub('OCS 固定單位 — 逐 Tab 比對')
    ocs_ok_total = ocs_diff_total = 0

    for sec in OCS_SECTIONS:
        lines.append(f'\n  Tab: OCS_{sec}')
        ref_units = ocs_ref.get(sec, [])
        res_units = ocs_result.get(sec)
        if res_units is None:
            lines.append('    ✗ Tab 缺失')
            ocs_diff_total += 1
            continue
        ref_map = {u['unit']: u['hc'] for u in ref_units}
        res_map = {u['unit']: u['hc'] for u in res_units}
        sec_ok = sec_diff = 0
        for unit in sorted(set(list(ref_map) + list(res_map))):
            r_hc = res_map.get(unit)
            b_hc = ref_map.get(unit)
            if unit in ref_map and unit in res_map:
                m = (r_hc == b_hc)
                flag = 'OK  ' if m else 'DIFF'
                if m: sec_ok += 1
                else: sec_diff += 1
            elif unit in ref_map:
                flag = 'MISS'; sec_diff += 1
            else:
                flag = 'XTRA'; sec_diff += 1
            lines.append(f'    {flag}  {unit:<35}  結果={str(r_hc):>4}  廠務={str(b_hc):>4}')
        verdict = '✓ 全部一致' if sec_diff == 0 else f'✗ {sec_diff} 筆差異'
        lines.append(f'    {"─"*50}  小計: OK={sec_ok}  DIFF={sec_diff}  {verdict}')
        ocs_ok_total += sec_ok
        ocs_diff_total += sec_diff

    # ── Summary ───────────────────────────────────────────────────────────────
    lines.append('')
    hdr('總結 — 差異原因分類')
    lines.append('')
    lines.append(f'  CSA 總筆數（非MF）: {rn}')
    lines.append(f'  LEAN 一致:          {lean_ok_n:4d} 筆')
    lines.append(f'  LEAN 不符:          {lean_diff_n:4d} 筆')
    lines.append(f'    ├─ {R_LEAN_CROSS:<12}: {reason_counts[R_LEAN_CROSS]:4d} 筆  (同一ART出現在多個DS-04部門，廠務只記一個LEAN)')
    lines.append(f'    └─ {R_LEAN_DIFF:<12}: {reason_counts[R_LEAN_DIFF]:4d} 筆  (ART僅在一個DS-04部門，但廠務LEAN不同)')
    lines.append(f'  ART DS04有廠務無:   {lean_miss_n:4d} 筆')
    lines.append(f'  ART 廠務有DS04無:   {len(ref_only):4d} 筆')
    lines.append('')
    lines.append(f'  訂單 一致:          {ord_ok_n:4d} 筆')
    lines.append(f'  訂單 不符:          {ord_diff_n:4d} 筆')
    lines.append(f'    ├─ {R_ORD_BATCH:<12}: {reason_counts[R_ORD_BATCH]:4d} 筆  (廠務一行包含多ART合計量，個別ART量必然不同)')
    lines.append(f'    └─ {R_ORD_MISMATCH:<12}: {reason_counts[R_ORD_MISMATCH]:4d} 筆  (廠務只記載一個ART，但數量不符)')
    lines.append('')
    lines.append(f'  OCS 固定單位:       ✓ 5 Tab 全部一致 ({ocs_ok_total} 單位)' if ocs_diff_total == 0
                 else f'  OCS 固定單位:       ✗ {ocs_diff_total} 筆差異')
    lines.append('')
    overall = (lean_diff_n == 0 and lean_miss_n == 0 and not ref_only and ocs_diff_total == 0)
    lines.append('整體結論: ' + ('✓ 100% 一致' if overall else '✗ 有差異（見上方清單）'))

    text = '\n'.join(lines)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(text)
    print(text)
    return text


if __name__ == '__main__':
    run()
