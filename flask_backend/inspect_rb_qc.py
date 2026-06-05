#!/usr/bin/env python3
"""Inspect RB and QC rows in 廠務組織編制表."""
import sys, re
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl

BIANCHE  = r'C:\Users\user\AppData\Local\Temp\bianche_tmp.xlsx'
MONTH_SH = '6.2026'

wb = openpyxl.load_workbook(BIANCHE, data_only=True, read_only=True)
ws = next((wb[s] for s in wb.sheetnames if s.strip() == MONTH_SH), None)

print('=== RB / QC rows (R395-R470) ===')
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i < 395: continue
    if i > 470: break
    # Print all non-empty columns
    cells = [(ci+1, str(v or '').strip()) for ci, v in enumerate(row) if str(v or '').strip()]
    if cells:
        print(f'R{i:4d}: ' + '  |  '.join(f'col{c}={v!r}' for c, v in cells[:8]))

wb.close()
