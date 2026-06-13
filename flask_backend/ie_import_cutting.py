"""
IE Process Import — Cutting/裁斷機區
Step 1: Create ie_process table
Step 2: Import Cutting sheets (zone=裁斷機) from all ARTs
Step 3: Output gap report xlsx
"""
import sqlite3, sys, re, os
sys.stdout.reconfigure(encoding='utf-8')

DB = os.path.join(os.path.dirname(__file__), 'data', 'atlas.db')
OUT_DIR = os.path.join(os.path.dirname(__file__), 'test_output')
os.makedirs(OUT_DIR, exist_ok=True)

# ── Step 1: Create ie_process table ──────────────────────────────────────────
def create_table(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ie_process (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            header_id   INTEGER NOT NULL,
            art         TEXT    NOT NULL,
            segment     TEXT    NOT NULL CHECK(segment IN ('cutting','stitching','assembly','stf')),
            zone        TEXT    NOT NULL,
            stage       INTEGER NOT NULL DEFAULT 1,
            seq         INTEGER,
            process_name TEXT,
            part_name   TEXT,
            tct         REAL,
            value_type  TEXT    CHECK(value_type IN ('formula','manual')),
            formula     TEXT,
            source_sheet TEXT,
            source_row  INTEGER,
            flag        TEXT
        )
    """)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_iep_hdr  ON ie_process(header_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_iep_art  ON ie_process(art)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_iep_seg  ON ie_process(segment, zone)")
    conn.commit()
    print("[Step 1] ie_process table ready.")

# ── Step 2: Import Cutting/裁斷機區 ──────────────────────────────────────────
FLAG_KEYWORDS = re.compile(r'atom|laser|emma|自动化', re.IGNORECASE)
MACHINE_COL7  = re.compile(r'^(ATOM|LASER|EMMA|Yinghui|yinghui)', re.IGNORECASE)

def is_numeric_seq(val):
    """Returns True if value looks like a row sequence number (integer string)."""
    if val is None:
        return False
    try:
        int(str(val).strip())
        return True
    except ValueError:
        return False

def import_cutting(conn):
    cur = conn.cursor()

    # Get all ARTs that have a cutting sheet
    cur.execute("""
        SELECT DISTINCT header_id, art
        FROM art_sheet_status
        WHERE sheet_type = 'cutting'
        ORDER BY header_id
    """)
    art_map = {row[0]: row[1] for row in cur.fetchall()}
    print(f"[Step 2] Found {len(art_map)} ARTs with cutting sheet type.")

    # Get all cutting-type sheet names from sheet_name_map
    cur.execute("SELECT raw_name FROM sheet_name_map WHERE type_key='cutting'")
    cutting_sheet_names = {r[0] for r in cur.fetchall()}

    inserted = 0
    arts_imported = set()

    for header_id, art in art_map.items():
        # Get all sheet names for this header_id that are cutting type
        cur.execute("""
            SELECT DISTINCT sheet_name
            FROM ie_sheet_data
            WHERE header_id = ?
        """, (header_id,))
        sheets_in_file = [r[0] for r in cur.fetchall()]
        cutting_sheets = [s for s in sheets_in_file if s in cutting_sheet_names]

        if not cutting_sheets:
            # Try fuzzy: sheet_name LIKE 'Cutting%' (trimmed)
            cutting_sheets = [s for s in sheets_in_file
                              if re.match(r'^\s*Cutting', s, re.IGNORECASE)]

        for sheet_name in cutting_sheets:
            # Fetch all rows for this sheet, pivoted by row number
            cur.execute("""
                SELECT row, col, value, formula, cell_type
                FROM ie_sheet_data
                WHERE header_id=? AND sheet_name=?
                ORDER BY row, col
            """, (header_id, sheet_name))

            # Group cells by row
            rows_data = {}
            for r, c, val, formula, cell_type in cur.fetchall():
                if r not in rows_data:
                    rows_data[r] = {}
                rows_data[r][c] = (val, formula, cell_type)

            # Carry-forward part_name from col 2
            current_part_name = None
            seq_counter = 0

            for row_num in sorted(rows_data.keys()):
                cols = rows_data[row_num]
                col2_val = cols.get(2, (None,))[0] if 2 in cols else None
                col3_val = cols.get(3, (None,))[0] if 3 in cols else None
                col4_val = cols.get(4, (None,))[0] if 4 in cols else None
                col7_val = cols.get(7, (None,))[0] if 7 in cols else None
                col8_info = cols.get(8, (None, None, 'manual')) if 8 in cols else (None, None, 'manual')
                col8_val, col8_formula, col8_type = col8_info

                # Update carry-forward part_name when col 2 has a non-header value
                if col2_val and col2_val.strip() and not is_numeric_seq(col2_val):
                    header_words = {'classification material', '材料类别', 'phân loại nguyên liệu',
                                    'classification', 'material'}
                    if col2_val.strip().lower() not in header_words:
                        current_part_name = col2_val.strip()

                # Process row = col 3 has an integer seq number AND col 4 has a name
                if not is_numeric_seq(col3_val):
                    continue
                if not col4_val or not str(col4_val).strip():
                    continue

                seq = int(str(col3_val).strip())
                process_name = str(col4_val).strip()

                # TCT
                tct_raw = col8_val
                tct = None
                if tct_raw is not None:
                    try:
                        tct = float(tct_raw)
                    except (ValueError, TypeError):
                        tct = None

                value_type = 'formula' if col8_formula else 'manual'
                formula_str = col8_formula

                # Flag detection: col 7 is machine type OR formula cross-references 自动化/ATOM/LASER/EMMA
                flag = None
                if col7_val and MACHINE_COL7.match(str(col7_val).strip()):
                    flag = '待分區'
                elif formula_str and FLAG_KEYWORDS.search(formula_str):
                    flag = '待分區'

                seq_counter += 1
                conn.execute("""
                    INSERT INTO ie_process
                      (header_id, art, segment, zone, stage, seq, process_name, part_name,
                       tct, value_type, formula, source_sheet, source_row, flag)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    header_id, art, 'cutting', '裁斷機', 1, seq,
                    process_name, current_part_name,
                    tct, value_type, formula_str,
                    sheet_name, row_num, flag
                ))
                inserted += 1

            if seq_counter > 0:
                arts_imported.add(art)

    conn.commit()
    print(f"[Step 2] Imported {inserted} process rows from {len(arts_imported)} ARTs.")
    return arts_imported, inserted

# ── Step 3: Gap report xlsx ───────────────────────────────────────────────────
def build_gap_report(conn, arts_imported):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[Step 3] openpyxl not available — installing...")
        import subprocess
        subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    cur = conn.cursor()

    # All ARTs with cutting sheet type
    cur.execute("SELECT DISTINCT header_id, art FROM art_sheet_status WHERE sheet_type='cutting'")
    all_cutting_arts = {r[1]: r[0] for r in cur.fetchall()}

    # Sheet 1: 完全缺 — has cutting sheet but zero process rows imported
    fully_missing = [(art, hid) for art, hid in all_cutting_arts.items()
                     if art not in arts_imported]

    # Sheet 2: 可疑 — imported but < 3 rows OR all TCT null
    cur.execute("""
        SELECT art, COUNT(*) as cnt, SUM(CASE WHEN tct IS NOT NULL THEN 1 ELSE 0 END) as tct_cnt
        FROM ie_process
        WHERE segment='cutting'
        GROUP BY art
    """)
    suspect = []
    for art, cnt, tct_cnt in cur.fetchall():
        if cnt < 3 or tct_cnt == 0:
            suspect.append((art, cnt, tct_cnt))

    # Sheet 3: 待分區 — flag='待分區' detail
    cur.execute("""
        SELECT art, source_sheet, source_row, process_name, tct, formula
        FROM ie_process
        WHERE flag='待分區'
        ORDER BY art, source_row
    """)
    pending = cur.fetchall()

    # Sheet 4: 統計
    cur.execute("SELECT COUNT(DISTINCT art) FROM ie_process WHERE segment='cutting'")
    success_arts = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM ie_process WHERE segment='cutting'")
    total_rows = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM ie_process WHERE segment='cutting' AND value_type='formula'")
    formula_cnt = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM ie_process WHERE segment='cutting' AND value_type='manual'")
    manual_cnt = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM ie_process WHERE segment='cutting' AND flag='待分區'")
    pending_cnt = cur.fetchone()[0]

    # Build xlsx
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='54463A')

    def make_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    # Sheet 1: 完全缺
    ws1 = wb.active
    ws1.title = '完全缺'
    make_header(ws1, ['ART', 'Header_ID', '說明'])
    for i, (art, hid) in enumerate(fully_missing, 2):
        ws1.cell(i, 1, art)
        ws1.cell(i, 2, hid)
        ws1.cell(i, 3, '有cutting sheet紀錄但無法抽出任何流程行')

    # Sheet 2: 可疑
    ws2 = wb.create_sheet('可疑')
    make_header(ws2, ['ART', '流程行數', 'TCT有值行數', '原因'])
    for i, (art, cnt, tct_cnt) in enumerate(suspect, 2):
        ws2.cell(i, 1, art)
        ws2.cell(i, 2, cnt)
        ws2.cell(i, 3, tct_cnt)
        reasons = []
        if cnt < 3: reasons.append('流程行數<3')
        if tct_cnt == 0: reasons.append('TCT全空')
        ws2.cell(i, 4, ' / '.join(reasons))

    # Sheet 3: 待分區
    ws3 = wb.create_sheet('待分區')
    make_header(ws3, ['ART', '來源Sheet', '來源行', '流程名稱', 'TCT', 'Formula'])
    for i, row in enumerate(pending, 2):
        for j, val in enumerate(row, 1):
            ws3.cell(i, j, val)

    # Sheet 4: 統計
    ws4 = wb.create_sheet('統計')
    stats = [
        ('成功導入ART數', success_arts),
        ('流程行總數', total_rows),
        ('formula行數', formula_cnt),
        ('manual行數', manual_cnt),
        ('formula比例', f'{formula_cnt/total_rows*100:.1f}%' if total_rows else '0%'),
        ('待分區行數', pending_cnt),
        ('完全缺ART數', len(fully_missing)),
        ('可疑ART數', len(suspect)),
    ]
    make_header(ws4, ['指標', '值'])
    for i, (k, v) in enumerate(stats, 2):
        ws4.cell(i, 1, k)
        ws4.cell(i, 2, v)

    # Auto-width
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    out_path = os.path.join(OUT_DIR, 'cutting裁斷區_缺口清單.xlsx')
    wb.save(out_path)
    print(f"[Step 3] Gap report saved: {out_path}")
    print(f"         完全缺: {len(fully_missing)} ARTs")
    print(f"         可疑:   {len(suspect)} ARTs")
    print(f"         待分區: {len(pending)} rows")
    return {
        'success_arts': success_arts, 'total_rows': total_rows,
        'formula_cnt': formula_cnt, 'manual_cnt': manual_cnt,
        'pending_cnt': pending_cnt, 'fully_missing': len(fully_missing),
        'suspect': len(suspect),
    }

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    conn = sqlite3.connect(DB)
    create_table(conn)
    arts_imported, total_inserted = import_cutting(conn)
    stats = build_gap_report(conn, arts_imported)
    conn.close()
    print()
    print("=" * 50)
    print("IMPORT SUMMARY — Cutting/裁斷機區")
    print("=" * 50)
    print(f"  成功導入ART數: {stats['success_arts']}")
    print(f"  流程行總數:    {stats['total_rows']}")
    print(f"  formula:      {stats['formula_cnt']} ({stats['formula_cnt']/stats['total_rows']*100:.1f}%)" if stats['total_rows'] else "  流程行總數: 0")
    print(f"  manual:       {stats['manual_cnt']}")
    print(f"  待分區:       {stats['pending_cnt']}")
    print(f"  完全缺ART:    {stats['fully_missing']}")
    print(f"  可疑ART:      {stats['suspect']}")
