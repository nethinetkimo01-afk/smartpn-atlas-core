#!/usr/bin/env python3
"""
auto_compare.py — Run after result_table rebuild.

Compares result_table_v2.xlsx (CSA + OCS fixed tabs) against 廠務組織編制表 6.2026.
Outputs summary to test_output/compare_result.txt.

Usage:
    python auto_compare.py [--result <path>] [--bianche <path>] [--out <path>]
"""
import sys, os, re, hashlib
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl

RESULT_DEFAULT  = r'D:\smartpn-atlas-core\flask_backend\test_output\result_table_v2.xlsx'
BIANCHE_DEFAULT = r'C:\Users\user\OneDrive\Desktop\Biên chế\Jun\2026年6月份廠務组织編制 20260524.xlsx'
OUT_DEFAULT     = r'D:\smartpn-atlas-core\flask_backend\test_output\compare_result.txt'
MONTH_SH        = '6.2026'
OCS_SECTIONS    = ['組底配套', '自動化', '電腦針車', '印刷', '設備工程']
_ART_RE         = re.compile(r'[A-Z]{2}\d{4,6}')


# ── Load 廠務組織編制表 ref ────────────────────────────────────────────────────
def _num(row, idx):
    v = row[idx] if len(row) > idx else None
    if v is None or str(v).strip() in ('.', '', '編制'):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

def _headcount(row):
    for idx in [12, 10, 14]:
        v = row[idx] if len(row) > idx else None
        if v is not None and str(v).strip() not in ('', '.', '編制'):
            try:
                return int(float(v))
            except (ValueError, TypeError):
                pass
    return None

def load_ref(bianche_path):
    wb = openpyxl.load_workbook(bianche_path, data_only=True)
    ws = None
    for sh in wb.sheetnames:
        if sh.strip() == MONTH_SH:
            ws = wb[sh]
            break

    csa_ref = []
    ocs_ref = {s: [] for s in OCS_SECTIONS}
    current_lean = ''
    current_ocs  = None

    for row in ws.iter_rows(values_only=True):
        col1 = str(row[0] or '').strip() if row else ''
        col2 = str(row[1] or '').strip() if len(row) > 1 else ''

        if re.match(r'^\d+[A-Z]\d*$', col1):
            current_lean = col1
            current_ocs  = None
            if col2 and col2 not in ('編制', ''):
                csa_ref.append({
                    'lean': current_lean, 'model': col2,
                    'arts': _ART_RE.findall(col2),
                    'order': _num(row, 6),
                })
            continue

        if col1 in OCS_SECTIONS:
            current_ocs  = col1
            current_lean = ''
            if col2 and col2 not in ('編制', ''):
                ocs_ref[current_ocs].append({'unit': col2, 'headcount': _headcount(row)})
            continue

        if current_lean and not col1 and col2 and col2 not in ('編制', ''):
            csa_ref.append({
                'lean': current_lean, 'model': col2,
                'arts': _ART_RE.findall(col2),
                'order': _num(row, 6),
            })
            continue

        if current_ocs and not col1 and col2 and col2 not in ('編制', ''):
            ocs_ref[current_ocs].append({'unit': col2, 'headcount': _headcount(row)})
            continue

        if col1 and col1 not in OCS_SECTIONS:
            current_lean = ''
            current_ocs  = None

    wb.close()

    art_to_ref = {}
    for r in csa_ref:
        for a in r['arts']:
            if a not in art_to_ref:
                art_to_ref[a] = r
    return art_to_ref, ocs_ref


# ── Load result table CSA ────────────────────────────────────────────────────
def load_result_csa(result_path):
    wb = openpyxl.load_workbook(result_path, data_only=True)
    ws = wb['CSA']
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 2:
            continue
        col_a = str(row[0] or '').strip()
        col_c = str(row[2] or '').strip() if len(row) > 2 else ''
        col_d = row[3] if len(row) > 3 else None
        if not _ART_RE.fullmatch(col_c):
            continue
        try:
            qty = int(float(col_d)) if col_d is not None else 0
        except (ValueError, TypeError):
            qty = 0
        rows.append({'lean': col_a, 'art': col_c, 'order': qty})
    wb.close()
    return rows


# ── Load result OCS fixed tabs ────────────────────────────────────────────────
def load_result_ocs(result_path):
    wb = openpyxl.load_workbook(result_path, data_only=True)
    result = {}
    for sec in OCS_SECTIONS:
        tab = f'OCS_{sec}'
        if tab not in wb.sheetnames:
            result[sec] = None
            continue
        ws = wb[tab]
        units = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i <= 2:
                continue
            col_a = str(row[0] or '').strip()
            col_b = row[1] if len(row) > 1 else None
            if not col_a or col_a in ('單位', '合計') or col_a.startswith('（'):
                continue
            try:
                hc = int(float(col_b)) if col_b is not None else None
            except (ValueError, TypeError):
                hc = None
            units.append({'unit': col_a, 'headcount': hc})
        result[sec] = units
    wb.close()
    return result


# ── Main compare function ────────────────────────────────────────────────────
def compare(result_path=None, bianche_path=None, out_path=None):
    result_path  = result_path  or RESULT_DEFAULT
    bianche_path = bianche_path or BIANCHE_DEFAULT
    out_path     = out_path     or OUT_DEFAULT

    lines = []
    lines.append('=' * 70)
    lines.append('非MP欄位比對報告：結果表 vs 廠務組織編制表 6.2026')
    lines.append('=' * 70)
    lines.append(f'結果表: {os.path.basename(result_path)}')
    lines.append(f'廠務表: {os.path.basename(bianche_path)}')
    lines.append('')

    art_to_ref, ocs_ref = load_ref(bianche_path)
    result_csa = load_result_csa(result_path)
    result_ocs = load_result_ocs(result_path)

    lines.append(f'廠務編制表 ART 索引: {len(art_to_ref)} 筆')
    lines.append(f'結果表 CSA 行數:     {len(result_csa)} 筆')
    lines.append('')

    # ── CSA LEAN check ────────────────────────────────────────────────────
    lean_ok = lean_wrong = lean_missing = 0
    lean_diff_lines = []
    art_not_in_ref  = []

    for r in result_csa:
        art  = r['art']
        lean = r['lean']
        if art.startswith('MF'):
            continue  # skip MF-prefix order codes

        if art not in art_to_ref:
            art_not_in_ref.append(r)
            lean_missing += 1
            continue

        ref_lean = art_to_ref[art]['lean']
        if lean == ref_lean:
            lean_ok += 1
        else:
            lean_wrong += 1
            lean_diff_lines.append(
                f'  LEAN不符  {art:12s}  結果表={lean!r:8s}  編制表={ref_lean!r}'
            )

    lines.append('── CSA LEAN 一致性 ──')
    lines.append(f'✓ 一致: {lean_ok}')
    if lean_wrong:
        lines.append(f'✗ LEAN不符: {lean_wrong} 筆')
        lines.extend(lean_diff_lines)
    if lean_missing:
        lines.append(f'✗ ART不在編制表: {lean_missing} 筆（排除MF前綴）')
        for r in art_not_in_ref[:20]:
            lines.append(f"  MISSING  {r['art']:12s}  LEAN={r['lean']!r}")
        if len(art_not_in_ref) > 20:
            lines.append(f'  ... 共 {len(art_not_in_ref)} 筆')

    result_arts = {r['art'] for r in result_csa if not r['art'].startswith('MF')}
    ref_only = sorted(a for a in art_to_ref if a not in result_arts)
    if ref_only:
        lines.append(f'✗ 廠務編制表有但結果表無: {len(ref_only)} 筆')
        for a in ref_only[:10]:
            lines.append(f"  REF_ONLY  {a:12s}  LEAN={art_to_ref[a]['lean']}")
        if len(ref_only) > 10:
            lines.append(f'  ... 共 {len(ref_only)} 筆')

    csa_ok = (lean_wrong == 0 and lean_missing == 0 and len(ref_only) == 0)

    # ── OCS fixed tabs ────────────────────────────────────────────────────
    lines.append('')
    lines.append('── OCS 固定單位一致性 ──')
    total_ocs_diffs = 0

    for sec in OCS_SECTIONS:
        ref_units = ocs_ref[sec]
        res_units = result_ocs.get(sec)

        if res_units is None:
            lines.append(f'  {sec}: ✗ Tab 缺失')
            total_ocs_diffs += 1
            continue

        ref_map = {u['unit']: u['headcount'] for u in ref_units}
        res_map = {u['unit']: u['headcount'] for u in res_units}
        diffs = 0

        for unit, ref_hc in ref_map.items():
            if unit not in res_map:
                lines.append(f'  {sec} MISSING  {unit!r}  ref={ref_hc}')
                diffs += 1
            elif res_map[unit] != ref_hc:
                lines.append(f'  {sec} HC_DIFF  {unit!r}  result={res_map[unit]}  ref={ref_hc}')
                diffs += 1
        for unit in res_map:
            if unit not in ref_map:
                lines.append(f'  {sec} EXTRA    {unit!r}  result={res_map[unit]}')
                diffs += 1

        if diffs == 0:
            lines.append(f'  {sec}: ✓ 完全一致 ({len(ref_units)} units)')
        total_ocs_diffs += diffs

    ocs_ok = (total_ocs_diffs == 0)

    # ── Summary ────────────────────────────────────────────────────────────
    lines.append('')
    lines.append('=' * 70)
    lines.append('總結')
    lines.append('=' * 70)
    csa_label = '✓ LEAN 100%一致' if csa_ok else f'✗ LEAN不符{lean_wrong}筆 / ART缺{lean_missing}筆 / 編制表多{len(ref_only)}筆'
    ocs_label = '✓ 100%一致' if ocs_ok else f'✗ {total_ocs_diffs} 筆差異'
    lines.append(f'CSA LEAN一致性: {csa_label}')
    lines.append(f'OCS 固定單位:   {ocs_label}')
    overall = csa_ok and ocs_ok
    verdict = '✓ 100% 一致' if overall else '✗ 有差異（見上方清單）'
    lines.append(f'\n整體結論: {verdict}')

    text = '\n'.join(lines)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(text)

    print(text)
    return overall, lean_wrong, lean_missing, len(ref_only), total_ocs_diffs


if __name__ == '__main__':
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument('--result',  default=RESULT_DEFAULT)
    p.add_argument('--bianche', default=BIANCHE_DEFAULT)
    p.add_argument('--out',     default=OUT_DEFAULT)
    args = p.parse_args()
    compare(args.result, args.bianche, args.out)
